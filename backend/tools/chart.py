"""Werkzeug ``create_chart`` – geprueftes Diagramm statt geratener Codeblock.

WARUM ES DAS GIBT (zwei Gruende, beide gemessen an echten Fehlerbildern):

1) VALIDIERUNG MIT RUECKMELDUNG. Bisher gab das Modell einen ```chartjs-Block
   als FREITEXT aus. Was daran kaputt ist, merkt erst der Browser – und dort
   steht dann "Chart-Daten ungueltig". Das Modell erfaehrt nichts davon und
   kann es nicht korrigieren; frontend/js/charts.js betreibt deshalb einen
   erheblichen Reparaturaufwand (Klammern schliessen, JS-Callbacks
   entfernen). Dieses Werkzeug prueft die Angaben, BEVOR etwas gerendert wird,
   und antwortet bei einem Fehler mit Klartext ("Serie 2 hat 5 Werte, es gibt
   aber 7 Kategorien") – das Modell ruft daraufhin im SELBEN Lauf korrigiert
   auf. Das ist der uebliche Repair-Loop fuer strukturierte Ausgaben.

2) DIE ZAHLEN LAUFEN NICHT DURCH DAS MODELL. Zwei Wege:
   - ``source``: das Werkzeug liest CSV/TSV/XLSX selbst und aggregiert. Das
     Modell nennt nur Datei und Spalten. Bisher musste es jeden Wert aus einem
     Werkzeug-Ergebnis abschreiben – bei mehreren hundert Punkten teuer und
     fehleranfaellig (data/instructions_default/tools.md warnt ausdruecklich
     davor, "ueber jeden Punkt einzeln nachzudenken").
   - Der Rueckgabewert ist ein kurzer MARKER ``[[JARVIS_CHART:<token>]]``.
     Die fertige Spezifikation liegt hier im Prozess; ``expand_markers()``
     setzt sie erst beim Ausliefern in die Antwort ein. Damit stehen die
     Zahlen nie im LLM-Kontext – auch nicht im gespeicherten Verlauf.
     Gleiches Muster wie ``[[JARVIS_DELIVER:…]]`` fuer Dateien.

ABGRENZUNG: Dieses Werkzeug erzeugt die INTERAKTIVE Web-Darstellung
(Chart.js). Ein herunterladbares PNG entsteht weiter ueber matplotlib/seaborn
per shell_execute – dort mit dem Stil ``backend/plotstyles/jarvis.mplstyle``.
Absichtlich kein PNG von hier: matplotlib liegt nicht im venv des Dienstes
(sondern im System-Python), ein Import waere im Backend-Prozess nicht
verfuegbar.

Die OPTIK setzt bewusst nicht dieses Werkzeug, sondern der Theme-Layer in
frontend/js/charts.js – nur dort sind Dark/Light, Branding-Farbe und die
lokalisierten Zahlenformate bekannt.
"""

import io
import csv
import json
import re
import secrets
import threading
from collections import OrderedDict
from pathlib import Path

from backend.tools.base import BaseTool

# ── Grenzen. Ueberschreitungen werden GEMELDET, nicht stillschweigend
# gekuerzt: eine abgeschnittene Datenreihe sieht wie eine vollstaendige aus.
MAX_SERIES = 12
MAX_POINTS = 1000
MAX_LABEL_LEN = 120
MAX_TABLE_ROWS = 200_000      # Notbremse gegen eine versehentlich riesige Datei

ALLOWED_TYPES = (
    "bar", "line", "pie", "doughnut", "radar", "polarArea", "bubble", "scatter",
)
# Diagrammarten ohne Kategorie-Achse: hier sind labels sinnlos bzw. die Daten
# sind Punktpaare.
_XY_TYPES = ("scatter", "bubble")
_SEGMENT_TYPES = ("pie", "doughnut", "polarArea")

AGGREGATES = ("none", "sum", "mean", "count", "min", "max")
SORTS = ("none", "label", "value_desc", "value_asc")

_MARKER_RE = re.compile(r"\[\[JARVIS_CHART:([0-9a-f]{16})\]\]")

# Fertige Spezifikationen bis zur Auslieferung. Prozessweit mit Deckel: der
# Marker wird typischerweise Sekunden spaeter eingeloest. Ein Eintrag bleibt
# nach dem Einloesen liegen (der Text wird mehrfach verarbeitet – Anzeige,
# Verlauf –, ein Verbrauch beim ersten Treffer wuerde den zweiten Durchlauf
# mit einem nackten Marker zuruecklassen).
_MAX_PENDING = 60
_pending: "OrderedDict[str, dict]" = OrderedDict()
_pending_lock = threading.Lock()


def register_spec(spec: dict) -> str:
    """Legt eine Chart.js-Spezifikation ab und gibt ihr Marker-Token zurueck."""
    token = secrets.token_hex(8)
    with _pending_lock:
        _pending[token] = spec
        while len(_pending) > _MAX_PENDING:
            _pending.popitem(last=False)
    return token


def get_spec(token: str):
    with _pending_lock:
        return _pending.get(token)


def expand_markers(text: str) -> str:
    """Ersetzt ``[[JARVIS_CHART:token]]`` durch den ```chartjs-Block.

    Wird auf den ANZEIGETEXT angewandt (nicht auf den LLM-Kontext): so
    rendert das Frontend das Diagramm, waehrend im Modell-Verlauf nur der
    kurze Marker steht. Ein unbekanntes Token wird ENTFERNT statt stehen
    gelassen – ein Marker im Klartext ist fuer den Benutzer sinnlos
    (z.B. nach einem Dienst-Neustart, wenn der Speicher leer ist)."""
    if not text or "[[JARVIS_CHART:" not in text:
        return text

    def _sub(m):
        spec = get_spec(m.group(1))
        if not spec:
            return ""
        try:
            return "\n```chartjs\n" + json.dumps(spec, ensure_ascii=False) + "\n```\n"
        except Exception:  # noqa: BLE001
            return ""

    out = _MARKER_RE.sub(_sub, text)
    return re.sub(r"\n{3,}", "\n\n", out).strip()


def strip_markers(text: str) -> str:
    """Entfernt die Marker ohne Ersatz – fuer Kanaele OHNE Web-Oberflaeche
    (WhatsApp, Telegram, Notify-API).

    Warum nicht auch dort expandieren: ein ```chartjs-Block ist dort blanker
    JSON-Text in der Nachricht. Der rohe Marker waere noch schlechter. Wer in
    diesen Kanaelen ein Diagramm braucht, erzeugt ein PNG per matplotlib –
    genau so steht es im Kanal-Prompt."""
    if not text or "[[JARVIS_CHART:" not in text:
        return text
    return re.sub(r"\n{3,}", "\n\n", _MARKER_RE.sub("", text)).strip()


# ── Zahlen aus Tabellen ─────────────────────────────────────────────────────

def parse_number(raw):
    """Wandelt einen Tabellenwert in eine Zahl. Erkennt deutsche UND englische
    Schreibweise, weil beide in denselben Exporten vorkommen.

    Regel bei GEMISCHTEN Trennzeichen ("1.234,56" / "1,234.56"): das
    RECHTESTE ist das Dezimaltrennzeichen. Bei nur einem Trennzeichen
    entscheidet das Muster – "1.234" und "1,234" sind Tausender (drei
    Nachstellen, keine weitere Gruppe), "1,5" ist eine Kommazahl. Diese
    Unterscheidung ist der Grund fuer die Funktion: ``float("1.234")``
    liefert 1.234 statt 1234 und verfaelscht damit still jede deutsche
    Tabelle um den Faktor 1000."""
    if raw is None:
        return None
    if isinstance(raw, bool):
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).strip()
    if not s:
        return None
    # Waehrung, Prozent, Einheiten, geschuetzte Leerzeichen
    s = s.replace(" ", "").replace(" ", "").replace(" ", "")
    s = s.replace("€", "").replace("$", "").replace("%", "").replace("'", "")
    neg = s.startswith("(") and s.endswith(")")     # Buchhaltung: (1.234)
    if neg:
        s = s[1:-1]
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        s = s.replace(",", "") if re.fullmatch(r"-?\d{1,3}(,\d{3})+", s) else s.replace(",", ".")
    elif "." in s:
        if re.fullmatch(r"-?\d{1,3}(\.\d{3})+", s):
            s = s.replace(".", "")
    try:
        v = float(s)
    except ValueError:
        return None
    return -v if neg else v


def csv_zeilen(path: Path, max_zeilen: int = MAX_TABLE_ROWS) -> list:
    """Rohe Zeilen einer CSV/TSV – EINE Stelle fuer Kodierung und Trennzeichen.

    Gibt Listen von Zeichenketten zurueck, ungefiltert und unkonvertiert. Wird
    von ``_read_table`` (Diagramme) UND von ``skills/office/tabellen.py``
    benutzt: seit 2026-08-24 nehmen die xlsx-Werkzeuge eine CSV als Quelle an,
    und zwei getrennte Erkennungen waeren beim naechsten Export-Format
    auseinandergelaufen (auf ECHT kommen deutsche Exporte mit ';' und cp1252).

    Wirft ``RuntimeError`` mit Klartext – die Aufrufer geben ihn weiter.
    """
    raw = path.read_bytes()
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise RuntimeError("Textkodierung nicht erkannt.")
    probe = "\n".join(text.splitlines()[:20])
    delim = "\t" if path.suffix.lower() == ".tsv" else ";"
    if path.suffix.lower() != ".tsv":
        try:
            delim = csv.Sniffer().sniff(probe, delimiters=";,\t|").delimiter
        except Exception:  # noqa: BLE001
            # Der Sniffer scheitert bei einer einspaltigen Datei. Dann
            # entscheidet die Haeufigkeit – ';' zuerst, weil deutsche Exporte
            # die Regel sind.
            delim = ";" if probe.count(";") >= probe.count(",") else ","
    zeilen = []
    for i, row in enumerate(csv.reader(io.StringIO(text), delimiter=delim), start=1):
        if i > max_zeilen:
            break
        zeilen.append(row)
    return zeilen


def _read_table(path: Path, sheet=None, header_row: int = 1):
    """Liest CSV/TSV/XLSX und gibt (spaltennamen, zeilen) zurueck.
    Zeilen sind Listen in Spaltenreihenfolge."""
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except Exception as e:  # noqa: BLE001
            raise RuntimeError(
                f"Excel-Dateien brauchen openpyxl (nicht verfuegbar: {e}). "
                "Speichere die Tabelle als CSV oder lies sie per shell_execute."
            ) from e
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        try:
            if sheet:
                if sheet not in wb.sheetnames:
                    raise RuntimeError(
                        f"Tabellenblatt '{sheet}' gibt es nicht. Vorhanden: "
                        + ", ".join(wb.sheetnames)
                    )
                ws = wb[sheet]
            else:
                ws = wb[wb.sheetnames[0]]
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if i > MAX_TABLE_ROWS:
                    break
                rows.append(list(row))
        finally:
            wb.close()
    else:
        rows = csv_zeilen(path)

    rows = [r for r in rows if r and any(c is not None and str(c).strip() != "" for c in r)]
    if not rows:
        raise RuntimeError("Die Datei enthaelt keine Daten.")
    hr = max(1, int(header_row or 1))
    if hr > len(rows):
        raise RuntimeError(f"header_row={hr} liegt hinter dem Dateiende ({len(rows)} Zeilen).")
    header = [("" if c is None else str(c).strip()) for c in rows[hr - 1]]
    # Namenlose Spalten bekommen ihre Position als Namen, damit sie
    # ansprechbar bleiben (sonst ist eine Spalte ohne Kopf unerreichbar).
    header = [h or f"Spalte{i + 1}" for i, h in enumerate(header)]
    return header, rows[hr:]


def _col_index(header, name):
    """Spaltenindex per Name (exakt, dann ohne Gross/Kleinschreibung, dann
    Teiltreffer) oder per 1-basierter Position ("3")."""
    if name is None:
        return -1
    n = str(name).strip()
    if not n:
        return -1
    if n in header:
        return header.index(n)
    low = [h.lower() for h in header]
    if n.lower() in low:
        return low.index(n.lower())
    treffer = [i for i, h in enumerate(low) if n.lower() in h]
    if len(treffer) == 1:
        return treffer[0]
    if re.fullmatch(r"\d+", n):
        i = int(n) - 1
        if 0 <= i < len(header):
            return i
    return -1


def _aggregate(values, how):
    vals = [v for v in values if v is not None]
    if how == "count":
        return float(len(values))
    if not vals:
        return None
    if how == "sum":
        return float(sum(vals))
    if how == "mean":
        return float(sum(vals) / len(vals))
    if how == "min":
        return float(min(vals))
    if how == "max":
        return float(max(vals))
    return float(vals[0])


class _Korrigierbar(Exception):
    """Fehler, den das Modell durch einen neuen Aufruf beheben kann."""


def _series_from_source(src: dict):
    """Baut (labels, series, hinweise) aus einer Datenquelle."""
    if not isinstance(src, dict):
        raise _Korrigierbar("source muss ein Objekt sein (file, label_column, value_columns).")
    datei = str(src.get("file") or src.get("path") or "").strip()
    if not datei:
        raise _Korrigierbar("source.file fehlt – gib den Pfad der CSV-/XLSX-Datei an.")
    p = Path(datei).expanduser()
    if not p.is_absolute():
        p = (Path(__file__).resolve().parent.parent.parent / datei).resolve()
    if not p.exists():
        raise _Korrigierbar(f"Datei nicht gefunden: {datei}")
    if p.is_dir():
        raise _Korrigierbar(f"{datei} ist ein Verzeichnis, keine Tabelle.")

    header, rows = _read_table(p, src.get("sheet"), int(src.get("header_row") or 1))

    lab_name = src.get("label_column") or src.get("labels")
    li = _col_index(header, lab_name)
    if lab_name and li < 0:
        raise _Korrigierbar(
            f"Spalte '{lab_name}' gibt es nicht. Vorhandene Spalten: " + ", ".join(header)
        )

    wanted = src.get("value_columns") or src.get("value_column") or src.get("values")
    if isinstance(wanted, str):
        wanted = [w.strip() for w in wanted.split(",") if w.strip()]
    if not wanted:
        raise _Korrigierbar(
            "source.value_columns fehlt – nenne mindestens eine Wertespalte. "
            "Vorhandene Spalten: " + ", ".join(header)
        )
    if not isinstance(wanted, list):
        raise _Korrigierbar("source.value_columns muss eine Liste von Spaltennamen sein.")
    vi = []
    for w in wanted:
        i = _col_index(header, w)
        if i < 0:
            raise _Korrigierbar(
                f"Wertespalte '{w}' gibt es nicht. Vorhandene Spalten: " + ", ".join(header)
            )
        vi.append((header[i], i))
    if len(vi) > MAX_SERIES:
        raise _Korrigierbar(f"Hoechstens {MAX_SERIES} Wertespalten – angefragt: {len(vi)}.")

    how = str(src.get("aggregate") or "none").lower()
    if how not in AGGREGATES:
        raise _Korrigierbar(f"aggregate muss einer von {', '.join(AGGREGATES)} sein.")

    hinweise = []
    # Gruppieren in der Reihenfolge des ersten Auftretens (stabil und
    # nachvollziehbar; ein sortiertes dict waere fuer Monatsnamen falsch).
    gruppen = OrderedDict()
    for r in rows:
        lab = ""
        if li >= 0:
            lab = "" if li >= len(r) or r[li] is None else str(r[li]).strip()
        if li >= 0 and not lab:
            continue                      # Zeile ohne Kategorie ueberspringen
        key = lab[:MAX_LABEL_LEN] if li >= 0 else len(gruppen)
        eintrag = gruppen.setdefault(key, [[] for _ in vi])
        for k, (_n, idx) in enumerate(vi):
            eintrag[k].append(parse_number(r[idx]) if idx < len(r) else None)

    if not gruppen:
        raise _Korrigierbar("Kein einziger verwertbarer Datensatz gefunden (Spalten leer?).")

    labels = [str(k) for k in gruppen.keys()]
    reihen = []
    for k, (nm, _idx) in enumerate(vi):
        werte = []
        for key in gruppen:
            spalte = gruppen[key][k]
            werte.append(_aggregate(spalte, how) if how != "none" else
                         next((v for v in spalte if v is not None), None))
        reihen.append({"label": nm, "data": werte})

    # Sortieren / begrenzen – beides ausdruecklich benannt, damit der Benutzer
    # weiss, dass er nicht alles sieht.
    sort = str(src.get("sort") or "none").lower()
    if sort not in SORTS:
        raise _Korrigierbar(f"sort muss einer von {', '.join(SORTS)} sein.")
    if sort != "none":
        idx = list(range(len(labels)))
        if sort == "label":
            idx.sort(key=lambda i: labels[i].lower())
        else:
            def _key(i):
                v = reihen[0]["data"][i]
                return v if isinstance(v, (int, float)) else float("-inf")
            idx.sort(key=_key, reverse=(sort == "value_desc"))
        labels = [labels[i] for i in idx]
        for s in reihen:
            s["data"] = [s["data"][i] for i in idx]

    top = src.get("top_n")
    if top:
        try:
            n = int(top)
        except (TypeError, ValueError):
            raise _Korrigierbar("top_n muss eine Zahl sein.") from None
        if 0 < n < len(labels):
            weg = len(labels) - n
            labels = labels[:n]
            for s in reihen:
                s["data"] = s["data"][:n]
            hinweise.append(f"Nur die ersten {n} von {n + weg} Kategorien dargestellt (top_n).")

    if len(labels) > MAX_POINTS:
        raise _Korrigierbar(
            f"{len(labels)} Kategorien sind zu viele (Grenze {MAX_POINTS}). "
            "Nutze aggregate/top_n oder fasse die Daten vorher zusammen."
        )
    hinweise.append(
        f"{len(labels)} Kategorien aus {p.name}"
        + (f", Aggregat: {how}" if how != "none" else "")
    )
    return labels, reihen, hinweise


class CreateChartTool(BaseTool):
    @property
    def name(self) -> str:
        return "create_chart"

    @property
    def description(self) -> str:
        return (
            "Erzeugt ein geprueftes, interaktives Diagramm im Chat (Balken, Linie, Torte, "
            "Streuung, Radar). BEVORZUGTER WEG fuer alle Diagramme aus Zahlen – statt einen "
            "```chartjs-Block selbst zu schreiben. Zwei Arten, die Daten zu uebergeben: "
            "(a) 'series' mit fertigen Werten, (b) 'source' mit Pfad zu einer CSV-/XLSX-Datei "
            "plus Spaltennamen – dann liest und aggregiert das Werkzeug die Datei selbst "
            "(Weg (b) IMMER vorziehen, wenn eine Datei vorliegt: keine abgeschriebenen Zahlen). "
            "Farben, Schrift und Zahlenformate setzt das System – keine Style-Angaben noetig. "
            "Der Rueckgabewert enthaelt eine Marker-Zeile, die UNVERAENDERT in die Antwort muss. "
            "Fuer ein herunterladbares PNG stattdessen matplotlib per shell_execute verwenden."
        )

    def parameters_schema(self) -> dict:
        return {
            "type": "object",
            "properties": {
                "type": {
                    "type": "string",
                    "enum": list(ALLOWED_TYPES),
                    "description": "Diagrammart. bar=Vergleich, line=Verlauf, pie/doughnut=Anteile, scatter=Zusammenhang.",
                },
                "title": {"type": "string", "description": "Ueberschrift des Diagramms."},
                "labels": {
                    "type": "array", "items": {"type": "string"},
                    "description": "Kategorien der x-Achse (nicht bei scatter/bubble).",
                },
                "series": {
                    "type": "array",
                    "description": "Datenreihen: [{label, data:[Zahlen]}]. Bei scatter/bubble: data:[{x,y[,r]}].",
                    "items": {
                        "type": "object",
                        "properties": {
                            "label": {"type": "string"},
                            "data": {"type": "array", "items": {}},
                        },
                        "required": ["data"],
                    },
                },
                "source": {
                    "type": "object",
                    "description": "Datenquelle statt 'series': liest die Datei selbst.",
                    "properties": {
                        "file": {"type": "string", "description": "Pfad zur CSV/TSV/XLSX-Datei."},
                        "label_column": {"type": "string", "description": "Spalte mit den Kategorien."},
                        "value_columns": {
                            "type": "array", "items": {"type": "string"},
                            "description": "Eine oder mehrere Wertespalten (je Spalte eine Datenreihe).",
                        },
                        "aggregate": {
                            "type": "string", "enum": list(AGGREGATES),
                            "description": "Zusammenfassung je Kategorie (sum/mean/count/min/max). Vorgabe none.",
                        },
                        "sheet": {"type": "string", "description": "Tabellenblatt (nur XLSX)."},
                        "header_row": {"type": "integer", "description": "Zeile mit den Spaltennamen (Vorgabe 1)."},
                        "sort": {"type": "string", "enum": list(SORTS), "description": "Sortierung."},
                        "top_n": {"type": "integer", "description": "Nur die ersten n Kategorien."},
                    },
                    "required": ["file", "value_columns"],
                },
                "x_title": {"type": "string", "description": "Achsentitel unten (mit Einheit)."},
                "y_title": {"type": "string", "description": "Achsentitel links (mit Einheit)."},
                "horizontal": {"type": "boolean", "description": "Balken waagerecht (lange Kategorienamen)."},
                "stacked": {"type": "boolean", "description": "Reihen stapeln (Anteile an einer Summe)."},
                "target_line": {
                    "type": "number",
                    "description": "Waagerechte Ziel-/Schwellenlinie auf der Wert-Achse.",
                },
                "target_label": {"type": "string", "description": "Beschriftung der Ziellinie."},
            },
            "required": ["type"],
        }

    # ── Pruefung + Aufbau ──────────────────────────────────────────────────
    def _baue_spec(self, kwargs) -> tuple:
        typ = str(kwargs.get("type") or "").strip()
        if typ not in ALLOWED_TYPES:
            raise _Korrigierbar(
                f"'type' fehlt oder ist unbekannt. Erlaubt: {', '.join(ALLOWED_TYPES)}."
            )

        labels = kwargs.get("labels")
        series = kwargs.get("series")
        source = kwargs.get("source")
        hinweise = []

        if source:
            if series:
                hinweise.append("'series' wurde ignoriert – 'source' hat Vorrang.")
            labels, series, hs = _series_from_source(source)
            hinweise.extend(hs)
        else:
            if not series:
                raise _Korrigierbar(
                    "Es fehlen die Daten: entweder 'series' (fertige Werte) oder 'source' "
                    "(Datei + Spalten) angeben."
                )
            if isinstance(series, dict):
                series = [series]
            if not isinstance(series, list):
                raise _Korrigierbar("'series' muss eine Liste von {label, data} sein.")
            if len(series) > MAX_SERIES:
                raise _Korrigierbar(f"Hoechstens {MAX_SERIES} Datenreihen (uebergeben: {len(series)}).")

        # Datenreihen pruefen und Zahlen normieren
        sauber = []
        for i, s in enumerate(series, start=1):
            if not isinstance(s, dict):
                raise _Korrigierbar(f"Datenreihe {i} ist kein Objekt {{label, data}}.")
            daten = s.get("data")
            if daten is None:
                raise _Korrigierbar(f"Datenreihe {i} ('{s.get('label', '')}') hat kein 'data'.")
            if not isinstance(daten, list) or not daten:
                raise _Korrigierbar(f"'data' der Datenreihe {i} muss eine nicht-leere Liste sein.")
            if len(daten) > MAX_POINTS:
                raise _Korrigierbar(
                    f"Datenreihe {i} hat {len(daten)} Werte (Grenze {MAX_POINTS}). "
                    "Fasse die Daten zusammen oder nutze source mit aggregate/top_n."
                )
            if typ in _XY_TYPES:
                punkte = []
                for k, p in enumerate(daten, start=1):
                    if not isinstance(p, dict) or ("x" not in p) or ("y" not in p):
                        raise _Korrigierbar(
                            f"{typ} braucht Punktpaare: data[{k - 1}] der Reihe {i} muss "
                            "{\"x\": Zahl, \"y\": Zahl} sein"
                            + (" (bei bubble zusaetzlich \"r\")." if typ == "bubble" else ".")
                        )
                    x, y = parse_number(p.get("x")), parse_number(p.get("y"))
                    if x is None or y is None:
                        raise _Korrigierbar(
                            f"data[{k - 1}] der Reihe {i}: x und y muessen Zahlen sein."
                        )
                    punkt = {"x": x, "y": y}
                    if typ == "bubble":
                        r = parse_number(p.get("r"))
                        punkt["r"] = r if r is not None else 6
                    punkte.append(punkt)
                sauber.append({"label": str(s.get("label") or f"Reihe {i}")[:MAX_LABEL_LEN],
                               "data": punkte})
            else:
                werte = [parse_number(v) for v in daten]
                if all(v is None for v in werte):
                    raise _Korrigierbar(
                        f"Datenreihe {i} enthaelt keine Zahl – 'data' braucht Zahlenwerte "
                        "(Texte wie \"k.A.\" bitte als null uebergeben)."
                    )
                sauber.append({"label": str(s.get("label") or f"Reihe {i}")[:MAX_LABEL_LEN],
                               "data": werte})

        # Kategorien
        if typ not in _XY_TYPES:
            if not labels:
                raise _Korrigierbar(
                    "'labels' fehlt – nenne die Kategorien der x-Achse "
                    "(gleiche Anzahl wie Werte je Datenreihe)."
                )
            if isinstance(labels, str):
                labels = [x.strip() for x in labels.split(",") if x.strip()]
            if not isinstance(labels, list):
                raise _Korrigierbar("'labels' muss eine Liste von Texten sein.")
            labels = [str(x)[:MAX_LABEL_LEN] for x in labels]
            for i, s in enumerate(sauber, start=1):
                if len(s["data"]) != len(labels):
                    raise _Korrigierbar(
                        f"Datenreihe {i} ('{s['label']}') hat {len(s['data'])} Werte, "
                        f"es gibt aber {len(labels)} Kategorien. Beides muss gleich lang "
                        "sein (fehlende Werte als null)."
                    )
            if typ in _SEGMENT_TYPES and len(sauber) > 1:
                hinweise.append(
                    "Kreisdiagramme zeigen nur die erste Datenreihe – fuer einen Vergleich "
                    "mehrerer Reihen ist 'bar' die bessere Wahl."
                )

        # ── Chart.js-Konfiguration. ABSICHTLICH ohne Farben/Schriften:
        # das macht der Theme-Layer im Browser (Dark/Light + Branding).
        spec = {
            "type": typ,
            "data": {"datasets": sauber},
            "options": {"plugins": {}, "scales": {}},
        }
        if typ not in _XY_TYPES:
            spec["data"]["labels"] = labels

        titel = str(kwargs.get("title") or "").strip()
        if titel:
            spec["options"]["plugins"]["title"] = {"display": True, "text": titel[:200]}

        if typ not in _SEGMENT_TYPES and typ != "radar":
            xt = str(kwargs.get("x_title") or "").strip()
            yt = str(kwargs.get("y_title") or "").strip()
            if xt:
                spec["options"]["scales"].setdefault("x", {})["title"] = {"display": True, "text": xt[:120]}
            if yt:
                spec["options"]["scales"].setdefault("y", {})["title"] = {"display": True, "text": yt[:120]}
            if kwargs.get("horizontal") and typ == "bar":
                spec["options"]["indexAxis"] = "y"
            if kwargs.get("stacked"):
                spec["options"]["scales"].setdefault("x", {})["stacked"] = True
                spec["options"]["scales"].setdefault("y", {})["stacked"] = True

            ziel = kwargs.get("target_line")
            if ziel is not None and ziel != "":
                z = parse_number(ziel)
                if z is None:
                    raise _Korrigierbar("target_line muss eine Zahl sein.")
                # Achse der Linie folgt der Ausrichtung: bei waagerechten
                # Balken liegt die Wert-Achse auf x.
                waagerecht = spec["options"].get("indexAxis") == "y"
                ann = {
                    "type": "line",
                    "borderColor": "rgba(239,68,68,0.9)",
                    "borderWidth": 2,
                    "borderDash": [6, 4],
                }
                ann["xMin" if waagerecht else "yMin"] = z
                ann["xMax" if waagerecht else "yMax"] = z
                zl = str(kwargs.get("target_label") or "").strip()
                if zl:
                    ann["label"] = {"display": True, "content": zl[:80],
                                    "position": "end", "backgroundColor": "rgba(239,68,68,0.9)"}
                spec["options"]["plugins"]["annotation"] = {"annotations": {"ziel": ann}}

        if not spec["options"]["scales"]:
            spec["options"].pop("scales")
        if not spec["options"]["plugins"]:
            spec["options"].pop("plugins")
        return spec, hinweise

    async def execute(self, **kwargs) -> str:
        # Vom Dispatch gesetzt; wird hier nur konsumiert, nicht ausgewertet
        # (die Pfadfreigabe entscheidet agent.py, siehe dort).
        kwargs.pop("_privileged", None)
        kwargs.pop("_status_callback", None)
        try:
            spec, hinweise = self._baue_spec(kwargs)
        except _Korrigierbar as e:
            # Die Klartext-Rueckmeldung IST der Repair-Loop: das Modell sieht
            # sie als Werkzeug-Ergebnis und ruft im selben Lauf korrigiert auf.
            return (
                f"FEHLER_KORRIGIERBAR: {e} "
                "Rufe create_chart danach mit korrigierten Angaben erneut auf."
            )
        except RuntimeError as e:
            return f"FEHLER: {e}"
        except Exception as e:  # noqa: BLE001
            return f"FEHLER: Diagramm konnte nicht erstellt werden: {e}"

        token = register_spec(spec)
        anzahl = len(spec["data"]["datasets"])
        punkte = sum(len(d["data"]) for d in spec["data"]["datasets"])
        info = f"{spec['type']}, {anzahl} Datenreihe(n), {punkte} Werte"
        if hinweise:
            info += " – " + " ".join(hinweise)
        return (
            f"DIAGRAMM_ERSTELLT ({info}).\n\n"
            "Gib in deiner finalen Antwort GENAU EINE eigene Zeile mit diesem Marker aus "
            "(unveraendert, ohne Codeblock, ohne Anfuehrungszeichen) – daraus entsteht das "
            "Diagramm:\n\n"
            f"[[JARVIS_CHART:{token}]]\n\n"
            "Schreibe die Zahlen NICHT zusaetzlich in die Antwort; ein kurzer Satz zur "
            "Aussage des Diagramms genuegt."
        )
