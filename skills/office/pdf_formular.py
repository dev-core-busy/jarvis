"""Formular-PDFs auslesen: eine Seite je Datensatz, Ergebnis als Tabelle.

WARUM ES DIESES MODUL GIBT (Vorfall ECHT, 2026-08-12 bis 2026-08-19)
====================================================================
Eine Benutzerin bat mehrfach darum, aus ``Einsender_KIM_Anbindung.pdf`` die
Adressdaten zu ziehen: 54 Seiten, je Seite EIN ausgefuelltes Formular mit
Name / Strasse / PLZ+Ort / Telefon / Fax / E-Mail / BSNR / LANR. Sieben Laeufe
an drei Tagen, kein brauchbares Ergebnis. Nachgemessen an den echten Logs:

    Anhang-Text im Prompt:            82.018 Zeichen
    llm_max_tokens auf ECHT:           8.192  (Denk- UND Antworttoken)
    -> zwei Laeufe endeten mit steps=0 und "max_tokens erreicht"
    Ein Lauf mit 15 Schritten endete mit dem Satz
    "Lass mich das Script Schritt fuer Schritt aufbauen." – ohne Ergebnis.

Der EINE Lauf, der eine Datei erzeugte, lieferte 16 von 54 Zeilen – und darin
falsche Adressen: fuer die erste Praxis stand dort ``54321 Absenderstadt``.
Richtig ist ``12345 Musterstadt`` – uebernommen worden war die PLZ des ABSENDERS
aus dem Briefkopf.
Solche Werte sind das eigentliche Problem: sie sehen plausibel aus.

DIE DREI URSACHEN – jede fuer sich reicht zum Scheitern
=======================================================
1. **Der Wert steht UEBER seinem Label, nicht daneben.** Gemessen ueber alle
   Seiten: das Label ``Strasse, Hausnummer:`` sitzt rund 6,5 pt TIEFER als der
   eingetragene Wert, weil der Eintrag auf der Formularlinie sitzt. Jedes
   zeilenbasierte Verfahren (``pdftotext -layout``, "nimm die naechste Zeile")
   ordnet deshalb zwangslaeufig falsch zu. Genau dieser Fehler steht im
   Protokoll: ``Name: -> NAECHSTE 'Am Beispielweg'`` – das ist die Strasse.
2. **Die Textebene ist beschaedigt.** Sie liefert ``12395 Musterstadt`` statt
   ``12345``, ``55 4S 33`` statt ``55 44 33``, ``llgner`` statt
   ``Ilgner``. Bei Adressdaten sind Ziffern der Inhalt – BSNR, LANR, PLZ,
   Telefon. Eine beschaedigte Ziffer ist kein Schoenheitsfehler.
3. **Die Menge.** 54 Formulare als Fliesstext im Prompt sind eine Aufgabe, die
   ein Modell "im Kopf" abarbeiten soll, waehrend derselbe Prompt sein
   Token-Budget bereits zu grossen Teilen aufbraucht.

DIE REGEL, DIE DIESES MODUL DURCHSETZT
======================================
Dieselbe wie in ``tabellen.py``: **die Daten gehen NIE durch das
Sprachmodell.** Das Modell nennt die Datei, die Zuordnung passiert hier im
Backend ueber die GEOMETRIE der Seite, und zurueck kommt eine fertige Tabelle
plus eine Bilanz, wie viel gefunden wurde.

WIE DIE ZUORDNUNG FUNKTIONIERT – die Schablone
==============================================
Ein Formular-PDF besteht aus vielen gleich aufgebauten Seiten. Gemessen an der
echten Datei streuen die Feldpositionen ueber 54 Seiten um **weniger als 1 %
der Seitenhoehe**. Die OCR verliest dagegen die Beschriftungen je Seite
unterschiedlich (``LANR(s):`` wurde als ``LANRIS)``, ``LANAI)``, ``LANR6S)``
gelesen, ``Anmerkungen:`` als ``Ah markNgerE``).

Daraus folgt: **die Position ist verlaesslich, der Labeltext nicht.** Also
wird in einem ersten Durchgang eine Schablone gelernt – wo liegt welches Feld,
und wie heisst es nach dem Mehrheitsentscheid ueber alle Seiten. Erst im
zweiten Durchgang werden die Werte an diesen Positionen geholt. Eine Seite,
auf der die Beschriftung unlesbar war, liefert ihren Wert trotzdem.
"""

from __future__ import annotations

import csv
import os
import re
import shutil
import subprocess
import tempfile
import statistics
import unicodedata
from collections import Counter
from difflib import SequenceMatcher
from pathlib import Path

from PIL import Image

from backend.tools.base import BaseTool


class FormularFehler(Exception):
    """Klartext-Fehler, der bis zum Modell durchgereicht wird."""


# ── Grenzen ──────────────────────────────────────────────────────────────
# Alle als Funktion bzw. modulweite Konstante mit ENV-Ueberschreibung: ein
# beim Import eingefrorener Wert waere bis zum Dienstneustart unveraenderlich
# (gleiche Begruendung wie documents.retention_days()).
def _zahl(name: str, vorgabe: int, unten: int, oben: int) -> int:
    try:
        v = int(os.environ.get(name, "").strip() or vorgabe)
    except (TypeError, ValueError):
        return vorgabe
    return max(unten, min(oben, v))


def max_seiten() -> int:
    """Deckel gegen ein versehentlich riesiges PDF. OCR kostet ~2 s je Seite."""
    return _zahl("JARVIS_PDFFORM_MAX_SEITEN", 300, 1, 2000)


DPI = 200                 # gemessen ausreichend; 300 brachte keine besseren Werte
MIN_CONF = 30.0           # unter 30 liefert Tesseract fast nur Rauschen
VORSCHAU_ZEILEN = 8
PROBE_SEITEN = 2        # Stichprobe fuer die Wahl der Textquelle
RENDER_BLOCK = 8        # so viele Seiten je poppler-Aufruf
ERGEBNIS_MAX = 16000


# ── Wortliste: eine Schnittstelle, zwei Quellen ──────────────────────────
# Beide Quellen liefern dieselbe Struktur {t,x,y,w,h}, damit die gesamte
# Geometrie-Logik darunter nur EINMAL existiert. Die Einheiten duerfen sich
# unterscheiden (pt bzw. Pixel) – gerechnet wird durchgehend RELATIV zur
# Seitenhoehe bzw. -breite.
# Tesseract parallelisiert sich intern ueber OpenMP und belegt dabei ALLE
# Kerne. Werden mehrere Seiten gleichzeitig erkannt, kaempfen die Instanzen
# um dieselben Kerne. Gemessen an 8 Seiten dieses PDFs: 69,1 s ohne, 4,5 s
# mit dieser Begrenzung – Faktor 15, hochgerechnet 466 s gegen 31 s fuer das
# ganze Dokument. Deshalb wird tesseract direkt mit EIGENER Umgebung
# aufgerufen statt ueber pytesseract: die Variable global zu setzen wuerde
# auch andere Rechenarbeit des Dienstes treffen (Embeddings, Bildmodelle).
_OCR_UMGEBUNG = {"OMP_THREAD_LIMIT": "1"}


def _woerter_ocr_datei(bildpfad: str, lang: str) -> list[dict]:
    """Eine Bilddatei erkennen -> Woerter mit Koordinaten (TSV von Tesseract)."""
    umgebung = dict(os.environ)
    umgebung.update(_OCR_UMGEBUNG)
    fertig = subprocess.run(
        ["tesseract", bildpfad, "stdout", "-l", lang, "--psm", "4", "tsv"],
        capture_output=True, text=True, env=umgebung, check=False,
    )
    if fertig.returncode != 0:
        raise FormularFehler(
            f"Texterkennung fehlgeschlagen: {(fertig.stderr or '').strip()[:200]}")
    aus = []
    for row in csv.DictReader(fertig.stdout.splitlines(),
                              delimiter="\t", quoting=csv.QUOTE_NONE):
        try:
            conf = float(row.get("conf") or -1)
        except (TypeError, ValueError):
            continue
        t = (row.get("text") or "").strip()
        if conf < MIN_CONF or not t:
            continue
        try:
            aus.append({"t": t, "x": int(row["left"]), "y": int(row["top"]),
                        "w": int(row["width"]), "h": int(row["height"])})
        except (KeyError, TypeError, ValueError):
            continue
    return aus


def _ocr_sprache() -> str:
    """Deutsch bevorzugt – und zwar ALLEIN.

    'deu+eng' ist nicht nur ein Drittel langsamer, es liest an diesem Dokument
    auch schlechter (gemessen: 46 von 54 Strassen gegen 54 von 54). Zwei
    Sprachmodelle gleichzeitig erhoehen die Auswahl an Deutungen, und bei
    Ziffern ist das ein Nachteil.
    """
    try:
        fertig = subprocess.run(["tesseract", "--list-langs"],
                                capture_output=True, text=True, check=False)
        vorhanden = {z.strip() for z in (fertig.stdout or "").splitlines()[1:]}
    except Exception:  # noqa: BLE001
        return "deu"
    for kandidat in ("deu", "eng"):
        if kandidat in vorhanden:
            return kandidat
    return "deu"


def _woerter_textebene(seite) -> list[dict]:
    aus = []
    for w in seite.extract_words():
        t = (w.get("text") or "").strip()
        if not t:
            continue
        aus.append({"t": t, "x": float(w["x0"]), "y": float(w["top"]),
                    "w": float(w["x1"]) - float(w["x0"]),
                    "h": float(w["bottom"]) - float(w["top"])})
    return aus


# ── Geometrie ────────────────────────────────────────────────────────────
def zeilen_bilden(woerter: list[dict], tol_faktor: float = 0.5) -> list[dict]:
    """Woerter zu Zeilen gruppieren – ueber die vertikale MITTE, nicht ueber
    die Oberkante: Label und Wert sind unterschiedlich gross gesetzt, ihre
    Oberkanten liegen deshalb weiter auseinander als ihre Mitten.

    Die Toleranz ist bewusst ENG. Beschriftung und eingetragener Wert stehen
    im Formular nur 11 bis 19 px auseinander (gemessen ueber alle Seiten,
    Zeilenabstand 37 px). Eine grosszuegige Toleranz wirft beide in eine
    Zeile – dann steht der Wert der STRASSE neben der Beschriftung 'Name:'
    und landet im falschen Feld. Zusammengehoerige Bruchstuecke eines
    mehrzeiligen Wertes werden spaeter ueber die Zuordnung wieder vereint,
    ein einmal falsch verschmolzenes Paar dagegen nie.
    """
    if not woerter:
        return []
    hoehen = sorted(w["h"] for w in woerter)
    tol = max(2.0, hoehen[len(hoehen) // 2] * tol_faktor)
    zeilen: list[dict] = []
    for w in sorted(woerter, key=lambda w: w["y"] + w["h"] / 2):
        m = w["y"] + w["h"] / 2
        if zeilen and abs(m - zeilen[-1]["m"]) <= tol:
            z = zeilen[-1]
            z["ws"].append(w)
            z["m"] = sum(x["y"] + x["h"] / 2 for x in z["ws"]) / len(z["ws"])
        else:
            zeilen.append({"m": m, "ws": [w]})
    for z in zeilen:
        z["ws"].sort(key=lambda w: w["x"])
        z["text"] = " ".join(w["t"] for w in z["ws"])
        z["x0"] = z["ws"][0]["x"]
    return zeilen


_ENDET_DOPPELP = re.compile(r":$")


def label_in_zeile(zeile: dict, max_label_x: float):
    """(name, rest_woerter, x_ende_label) – Label ist alles bis zum LINKESTEN
    Wort, das auf ':' endet. ``max_label_x`` verhindert, dass eine Kopfzeile
    weit rechts ('Telefon: +49 …' im Briefkopf) zum Formularfeld wird."""
    ws = zeile["ws"]
    if not ws or ws[0]["x"] > max_label_x:
        return None
    for i, w in enumerate(ws):
        if _ENDET_DOPPELP.search(w["t"]):
            name = " ".join(x["t"] for x in ws[:i + 1]).rstrip(":").strip()
            if not name:
                return None
            return name, ws[i + 1:], w["x"] + w["w"]
    return None


def _normal(s: str) -> str:
    s = unicodedata.normalize("NFKD", (s or "").lower())
    return re.sub(r"[^a-z0-9]", "", s)


def wertspalte_lernen(seiten_zeilen, breiten, max_label_x, bin_breite=0.03):
    """Wo beginnt die Spalte der EINGETRAGENEN Werte (relativ zur Breite)?

    Noetig, weil auf derselben Zeilenhoehe noch ganz andere Dinge stehen
    koennen. In der echten Datei sitzt rechts unten ein Praxisstempel; ohne
    diese Schranke wandert dessen Text ins Feld 'Anmerkungen' – gemessen:
    133 echte Werte beginnen bei x/Breite ~0,31, die 23 Stempelzeilen bei
    ~0,70. Der dichteste Bin gewinnt.
    """
    starts, label_starts = [], []
    for zeilen, breite in zip(seiten_zeilen, breiten):
        for z in zeilen:
            tr = label_in_zeile(z, max_label_x)
            if tr:
                label_starts.append(z["x0"] / breite)
                if tr[1]:
                    starts.append(tr[1][0]["x"] / breite)
    if not starts:
        return 0.28, 1.0, 0.20
    haeufig = Counter(int(s / bin_breite) for s in starts).most_common(1)[0][0]
    im_bin = [s for s in starts if int(s / bin_breite) == haeufig]
    mitte = statistics.median(im_bin)
    x_min = max(0.0, mitte - 0.06)
    # Grenze fuer FORTSETZUNGSZEILEN eines mehrzeiligen Wertes: die Mitte
    # zwischen Beschriftungsspalte und Wertspalte. Ein langer Eintrag beginnt
    # weiter links als die Spalte (gemessen: 'Urologische' bei 21 % der
    # Blattbreite, Spalte bei 31 %) – waere die Grenze die Spalte selbst,
    # ginge die erste Zeile des Namens verloren. Eine Beschriftung (10 %)
    # bleibt trotzdem draussen.
    x_fort = (statistics.median(label_starts) + x_min) / 2 if label_starts else x_min * 0.7
    return x_min, mitte + 0.22, x_fort


def _spalten_messen(seiten_zeilen, breiten, grobe_grenze):
    """(x der Beschriftungsspalte, x der Wertspalte) – beide relativ."""
    label_x, wert_x = [], []
    for zeilen, breite in zip(seiten_zeilen, breiten):
        for z in zeilen:
            tr = label_in_zeile(z, grobe_grenze)
            if tr:
                label_x.append(z["x0"] / breite)
                if tr[1]:
                    wert_x.append(tr[1][0]["x"] / breite)
    if not label_x:
        return 0.10, 0.30
    lx = statistics.median(label_x)
    wx = statistics.median(wert_x) if wert_x else lx + 0.20
    if wx <= lx + 0.05:
        wx = lx + 0.20
    return lx, wx


def schablone_lernen(seiten_zeilen, hoehen, max_label_x, tol=0.008):
    """Label-Positionen ueber ALLE Seiten clustern -> Feldliste.

    Der kanonische Name ist die haeufigste gelesene Schreibweise: bei 54
    Seiten korrigiert die Mehrheit die Verlesungen der Einzelseiten.

    FALLSTRICK KETTENBILDUNG: Verglichen wird gegen den MEDIAN des Clusters,
    nicht gegen dessen letzten Eintrag. Sonst wandert ein Cluster mit jedem
    Fund ein Stueck weiter und verschluckt das naechste Feld – gemessen an
    der echten Datei wurden so aus 'Name', 'Strasse', 'PLZ/Ort' und 'Telefon'
    (Abstand je 0,018) ein einziges Feld namens 'Name', und die Werte lagen
    danach um eine Zeile verschoben in der Tabelle.
    """
    funde = []
    for si, (zeilen, hoehe) in enumerate(zip(seiten_zeilen, hoehen)):
        if not hoehe:
            continue
        for z in zeilen:
            tr = label_in_zeile(z, max_label_x)
            if tr:
                funde.append({"seite": si, "y": z["m"] / hoehe, "name": tr[0]})
    funde.sort(key=lambda f: f["y"])
    cluster: list[dict] = []
    for f in funde:
        passt = False
        if cluster:
            c = cluster[-1]
            nah = abs(f["y"] - c["mitte"]) <= tol
            # ZWEITES Kriterium: der Name. Bei einem schief eingezogenen Scan
            # ueberlappen die Hoehenbereiche benachbarter Felder ueber die
            # Seiten hinweg – dann trennt die Position allein nicht mehr, und
            # 'Name' verschluckte im Test 'Strasse'. Zwei Funde gehoeren nur
            # zusammen, wenn sie auch dieselbe Beschriftung tragen; eine
            # Verlesung derselben Beschriftung ist sich immer noch aehnlich.
            aehnlich = SequenceMatcher(
                None, _normal(f["name"]),
                _normal(Counter(c["namen"]).most_common(1)[0][0])).ratio()
            passt = nah and aehnlich >= 0.55
        if passt:
            c = cluster[-1]
            c["ys"].append(f["y"])
            c["namen"].append(f["name"])
            c["seiten"].add(f["seite"])
            c["mitte"] = statistics.median(c["ys"])
        else:
            cluster.append({"ys": [f["y"]], "namen": [f["name"]],
                            "seiten": {f["seite"]}, "mitte": f["y"]})
    felder = []
    for c in cluster:
        name = Counter(c["namen"]).most_common(1)[0][0]
        felder.append({"name": name, "y": statistics.median(c["ys"]),
                       "seiten": len(c["seiten"]),
                       "varianten": sorted(set(c["namen"]))})
    # Zwei Cluster mit demselben kanonischen Namen direkt untereinander sind
    # ein auseinandergerissenes Feld – zusammenlegen, sonst stuende dieselbe
    # Spalte zweimal in der Tabelle.
    verschmolzen: list[dict] = []
    for f in felder:
        if verschmolzen and _normal(verschmolzen[-1]["name"]) == _normal(f["name"]) \
                and abs(f["y"] - verschmolzen[-1]["y"]) <= tol * 2:
            v = verschmolzen[-1]
            v["y"] = (v["y"] + f["y"]) / 2
            v["seiten"] += f["seiten"]
            v["varianten"] = sorted(set(v["varianten"]) | set(f["varianten"]))
        else:
            verschmolzen.append(f)
    return verschmolzen


# Hoechstens so viele Woerter duerfen links der Wertspalte stehen, damit der
# Rest der Zeile noch als Wert gilt. Ein verlesenes Label ist ein bis zwei
# Woerter ('Name:' wurde auf Seite 44 der echten Datei zu 'Nana'); ein Satz
# aus dem Fliesstext hat mehr.
MAX_WOERTER_LINKS = 3

# Ab dieser Laenge darf ein wiederkehrender Wert ueber AEHNLICHKEIT als
# Kopf-/Fusszeile gelten (darunter zaehlt nur exakte Gleichheit).
BOILERPLATE_MIN_LAENGE = 25


def wertfragmente(zeilen, max_label_x, links_grenze, wert_grenze=None):
    """Alle Textstuecke, die NICHT Beschriftung sind – je mit eigener Hoehe.

    ZWEI Grenzen, und das ist kein Zufall:

    * ``max_label_x`` (grosszuegig) entscheidet, ob eine Zeile eine
      Beschriftung TRAEGT. Sie muss weit reichen, weil die Texterkennung
      Wortanfaenge verschluckt – auf Seite 1 der echten Datei fiel
      'Postleitzahl,' aus, das Feld hiess dort nur noch 'Ort:' und begann
      erst bei 19 % der Blattbreite.
    * ``links_grenze`` (eng) entscheidet, ob eine Zeile OHNE Doppelpunkt
      Fliesstext der linken Spalte ist – oder die Fortsetzung eines langen
      Wertes, die weiter links beginnt als die Wertspalte (gemessen: 21 %
      gegen 31 %).

    Mit nur EINER Grenze geht zwangslaeufig eines von beiden verloren.
    """
    frag = []
    for z in zeilen:
        tr = label_in_zeile(z, max_label_x)
        if tr:
            rest = tr[1]
        elif _beginnt_links(z, links_grenze):
            # Kein Doppelpunkt und links beginnend – normalerweise Fliesstext.
            # ABER: die Texterkennung verliert Doppelpunkte ('Name:' wurde zu
            # 'Nana'). Dann steht der Wert trotzdem in der Zeile und ginge
            # ersatzlos verloren. Gerettet wird er nur, wenn links davon
            # wenige Woerter stehen – ein ganzer Satz bleibt Fliesstext.
            rest = []
            if wert_grenze is not None:
                links = [w for w in z["ws"] if w["x"] < wert_grenze]
                rechts = [w for w in z["ws"] if w["x"] >= wert_grenze]
                if rechts and len(links) <= MAX_WOERTER_LINKS:
                    rest = rechts
        else:
            rest = z["ws"]
        if not rest:
            continue
        frag.append({"ws": rest,
                     "m": sum(w["y"] + w["h"] / 2 for w in rest) / len(rest),
                     "x0": min(w["x"] for w in rest)})
    frag.sort(key=lambda f: f["m"])
    return frag


def _beginnt_links(zeile, max_label_x) -> bool:
    return bool(zeile["ws"]) and zeile["ws"][0]["x"] <= max_label_x


def _fragmente_text(fragmente, breite) -> str:
    """Mehrere Bruchstuecke zu einem Wert zusammensetzen.

    Neu nach Zeilen gruppiert und je Zeile nach x sortiert: bei einem
    zweizeiligen Praxisnamen liegt das erste Wort ('Urologische', x=349)
    TIEFER als die Fortsetzung rechts daneben ('Gemein.-Praxis …', x=503).
    Reine Sortierung nach Hoehe drehte den Namen um.
    """
    ws = [w for f in fragmente for w in f["ws"]]
    if not ws:
        return ""
    teile = []
    for z in zeilen_bilden(ws, tol_faktor=0.9):
        stueck = _ab_spaltenbeginn(z if "breite" in z else dict(z, breite=breite), 0.0, 1.0)
        if stueck:
            teile.append(stueck)
    return " ".join(teile).strip()


def seitenversatz(zeilen, hoehe, felder, max_label_x, tol=0.02) -> float:
    """Wie weit ist DIESE Seite gegen die Schablone verschoben (in Pixeln)?

    Ein Scanner zieht Blaetter nicht auf den Punkt ein. Ohne Ausgleich sucht
    die Schablone an einer Stelle, an der auf dieser Seite schon das naechste
    Feld steht – bei 38 px Zeilenabstand genuegen 16 px Versatz, damit die
    Werte um ein Feld verrutschen. Gemessen wird der Median ueber alle
    Beschriftungen, die sich auf der Seite eindeutig wiederfinden; ohne
    Treffer bleibt es bei 0.
    """
    diffs = []
    for z in zeilen:
        tr = label_in_zeile(z, max_label_x)
        if not tr:
            continue
        name = _normal(tr[0])
        beste, bester_abstand = None, tol * hoehe
        for f in felder:
            if SequenceMatcher(None, name, _normal(f["name"])).ratio() < 0.75:
                continue
            d = z["m"] - f["y"] * hoehe
            if abs(d) < bester_abstand:
                beste, bester_abstand = d, abs(d)
        if beste is not None:
            diffs.append(beste)
    return statistics.median(diffs) if diffs else 0.0


def wert_holen(fragmente, label_ys, feld, hoehe, breite, fenster, vergeben,
               x_max=1.0, versatz=0.0):
    """Wert EINES Feldes: die Bruchstuecke UEBER seiner Beschriftung.

    Das ist die zentrale Regel dieses Moduls, und sie ist gemessen: der
    eingetragene Wert sitzt auf der Formularlinie und steht damit 11 bis 19 px
    HOEHER als seine Beschriftung, bei 37 px Zeilenabstand. Wer stattdessen
    'gleiche Zeile' oder 'naechste Zeile darunter' annimmt, ordnet jedes Feld
    um eins verschoben zu – genau daran sind die bisherigen Versuche
    gescheitert.

    ``vergeben`` verhindert, dass zwei Felder sich dasselbe Bruchstueck teilen.
    """
    ziel = feld["y"] * hoehe + versatz
    oben, unten = fenster
    # Kein Einsammeln ueber eine FREMDE Beschriftung hinweg. Die eigene zaehlt
    # ausdruecklich NICHT: der Wert steht ja UEBER ihr – wer sie als Schranke
    # nimmt, verwirft genau den Wert, den er sucht (und die Tabelle bleibt
    # ausgerechnet bei den wichtigsten Feldern leer).
    schranke = max((y for y in label_ys if y < ziel - oben), default=-1.0)
    treffer = []
    for i, f in enumerate(fragmente):
        if i in vergeben:
            continue
        d = ziel - f["m"]
        if not (-unten <= d <= oben) or f["m"] <= schranke:
            continue
        # Was jenseits der Wertspalte BEGINNT, gehoert zu einem anderen
        # Seitenblock (Praxisstempel, zweite Spalte) – nicht zum Feld.
        if f["x0"] / max(1e-9, breite) > x_max:
            continue
        treffer.append((i, f))
    if not treffer:
        return "", "leer"
    for i, _ in treffer:
        vergeben.add(i)
    return _fragmente_text([f for _, f in treffer], breite), "treffer"


# Ab dieser horizontalen Luecke (Anteil der Blattbreite) beginnt ein neuer
# Seitenblock. Gemessen: Wortabstaende innerhalb eines Wertes liegen unter
# 2 %, der Praxisstempel rechts unten steht 22 % neben dem Feldinhalt.
LUECKE_MAX = 0.08


def _ab_spaltenbeginn(zeile, x_min, x_max) -> str:
    """Wert = ab dem ersten Wort, das IN der Wertspalte beginnt, bis Zeilenende.

    Die obere Grenze gilt nur fuer den ANFANG. Ein eingetragener Wert darf
    beliebig weit nach rechts laufen – der Name auf Seite 1 der echten Datei
    reicht bis 81 % der Blattbreite, waehrend die Spalte bei 31 % beginnt.
    Wuerde x_max auf jedes Wort angewendet, endete der Wert mitten im Text
    ('Herr M. Ilgner, FA fuer Augenheilkunde,' – der Rest fehlt). Der
    Praxisstempel rechts unten faellt trotzdem heraus: er BEGINNT jenseits
    der Spalte.
    """
    breite = max(1e-9, zeile["breite"])
    for i, w in enumerate(zeile["ws"]):
        rel = w["x"] / breite
        if x_min <= rel <= x_max:
            teile = [w["t"]]
            vor = w
            for x in zeile["ws"][i + 1:]:
                # Grosse Luecke = anderer Block auf gleicher Hoehe (Stempel,
                # zweite Spalte). Ohne diesen Abbruch wandert er in den Wert:
                # 'Praxis nimmt am LDT-Versandt teil. 987654321'.
                if (x["x"] - (vor["x"] + vor["w"])) / breite > LUECKE_MAX:
                    break
                teile.append(x["t"])
                vor = x
            return " ".join(teile).strip()
    return ""


# ── Seiten einlesen ──────────────────────────────────────────────────────
def _pdf_seiten_woerter(pfad: Path, quelle: str, seiten_wahl):
    """[(woerter, breite, hoehe, seitennr)] – aus OCR oder aus der Textebene.

    ``quelle='auto'`` entscheidet ueber die vorhandene Qualitaetspruefung aus
    backend.tools.knowledge: eine GESUNDE Textebene ist exakt und kostet keine
    Rechenzeit, eine beschaedigte liefert falsche Ziffern (siehe Modulkopf).
    """
    try:
        import pdfplumber  # noqa: PLC0415
    except ImportError as e:  # noqa: BLE001
        raise FormularFehler(
            "pdfplumber fehlt. Nachinstallieren: 'sudo bash deploy/sandbox_python.sh' "
            "bzw. im venv 'pip install pdfplumber'."
        ) from e

    with pdfplumber.open(str(pfad)) as pdf:
        gesamt = len(pdf.pages)
        indizes = _seiten_indizes(seiten_wahl, gesamt)
        if len(indizes) > max_seiten():
            raise FormularFehler(
                f"{len(indizes)} Seiten angefordert, Deckel liegt bei {max_seiten()}. "
                f"Grenze den Bereich ueber 'seiten' ein (z.B. '1-50')."
            )
        masse = {i: (float(pdf.pages[i].width), float(pdf.pages[i].height)) for i in indizes}
        gewaehlt = quelle
        if quelle == "auto":
            gewaehlt = _quelle_waehlen(pdf, indizes, pfad)
        if gewaehlt == "text":
            aus = []
            for i in indizes:
                b, h = masse[i]
                aus.append((_woerter_textebene(pdf.pages[i]), b, h, i + 1))
            return aus, gewaehlt, gesamt

    # OCR laeuft ausserhalb des pdfplumber-Kontexts – die Bilder kommen von
    # poppler, nicht von pdfplumber.
    return _ocr_seiten(pfad, indizes), gewaehlt, gesamt


def _seiten_indizes(wahl, gesamt: int) -> list[int]:
    """'1-20', '3', '1,5,9-12' oder leer = alle. 1-basiert nach aussen."""
    if not wahl or not str(wahl).strip():
        return list(range(gesamt))
    aus: list[int] = []
    for teil in str(wahl).replace(" ", "").split(","):
        if not teil:
            continue
        if "-" in teil:
            a, _, b = teil.partition("-")
            try:
                von, bis = int(a), int(b)
            except ValueError as e:
                raise FormularFehler(f"Seitenangabe '{teil}' nicht lesbar (erwartet z.B. '1-20').") from e
            aus.extend(range(von - 1, bis))
        else:
            try:
                aus.append(int(teil) - 1)
            except ValueError as e:
                raise FormularFehler(f"Seitenangabe '{teil}' nicht lesbar.") from e
    gueltig = sorted({i for i in aus if 0 <= i < gesamt})
    if not gueltig:
        raise FormularFehler(f"Keine gueltige Seite in '{wahl}' (das PDF hat {gesamt} Seiten).")
    return gueltig


# Kennzahlen fuer den VERGLEICH zweier Fassungen desselben Dokuments.
# Bewusst KEINE absolute Schadensschwelle: eine solche Heuristik trifft
# zwangslaeufig auch echte Fachdaten (ICD-10 'O61.0', Pflegekategorie 'A4S1'
# – die Lehre aus der Durchsicht vom 2026-08-13). Verglichen werden immer
# Textebene GEGEN Texterkennung derselben Seiten; was beiden gemeinsam ist,
# faellt dabei heraus.
_GEMISCHT = re.compile(r"\b(?=[^\s]*\d)(?=[^\s]*[A-Za-z])[A-Za-z0-9]{2,}\b")
_LANGE_NUMMER = re.compile(r"\b\d{7,}\b")
_PLZ = re.compile(r"\b\d{5}\b")
_MAIL = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")


def textguete(text: str) -> tuple[float, int]:
    """(gemischte Tokens je 1000 Zeichen, verwertbare Treffer).

    'Gemischt' heisst: ein Wort enthaelt Ziffern UND Buchstaben. Eine kaputte
    Zeichentabelle erzeugt davon viele ('55 4S 33', 'OL.O7.2026'); als
    ABSOLUTES Mass waere die Zahl wertlos, im Vergleich zweier Fassungen
    desselben Textes ist sie aussagekraeftig.

    'Verwertbar' sind zusammenhaengende lange Nummern (BSNR, LANR),
    Postleitzahlen und E-Mail-Adressen – genau das, was aus einem
    Adressformular gebraucht wird. Die beschaedigte Fassung zerreisst sie
    ('12345 6789' statt '123456789') und verliert damit Treffer.
    """
    if not text or not text.strip():
        return 999.0, 0
    kilo = max(0.001, len(text) / 1000)
    gemischt = len(_GEMISCHT.findall(text)) / kilo
    # Lange Nummern zaehlen doppelt: eine zusammenhaengende BSNR/LANR ist der
    # staerkste Beleg fuer eine intakte Zeichentabelle. Die beschaedigte
    # Fassung zerreisst sie ('12345 6789'), und die Bruchstuecke saehen als
    # Einzeltreffer sonst genauso gut aus wie das Original.
    treffer = (2 * len(_LANGE_NUMMER.findall(text)) + len(_PLZ.findall(text))
               + 2 * len(_MAIL.findall(text)))
    return gemischt, treffer


def _quelle_waehlen(pdf, indizes: list[int], pfad: Path) -> str:
    """Textebene oder Texterkennung? GEMESSEN an einer Stichprobe.

    WARUM NICHT NUR EINE HEURISTIK: Die erste Fassung fragte
    ``backend.tools.knowledge.pdf_text_verdacht`` und fiel bei einem
    Importfehler still auf 'text' zurueck. Ergebnis am echten PDF: eine um
    eine Ziffer falsche Postleitzahl und eine zerrissene BSNR – falsche
    Adressdaten, die plausibel aussehen. Eine Pruefung, die im Zweifel die
    unsichere Quelle waehlt, ist schlimmer als keine.

    Deshalb werden BEIDE Wege an wenigen Seiten wirklich gelesen und
    verglichen. Das kostet ein paar Sekunden und entscheidet anhand dessen,
    worauf es ankommt: zusammenhaengende Ziffernbloecke (BSNR, LANR, PLZ) und
    lesbare E-Mail-Adressen.
    """
    probe = indizes[:PROBE_SEITEN]
    text = "\n".join((pdf.pages[i].extract_text() or "") for i in probe)
    # Praktisch kein Text -> gescanntes PDF, da hilft nur Texterkennung.
    if len(text.strip()) < 200 * len(probe):
        return "ocr"
    try:
        ocr_text = "\n".join(
            " ".join(w["t"] for w in woerter)
            for woerter, _, _, _ in _ocr_seiten(pfad, probe))
    except Exception:  # noqa: BLE001
        # Ohne Texterkennung bleibt nur die Textebene – aber der Aufrufer
        # erfaehrt es ueber den Bericht.
        return "text"
    g_text, t_text = textguete(text)
    g_ocr, t_ocr = textguete(ocr_text)
    # Texterkennung gewinnt, wenn sie spuerbar mehr verwertbare Angaben
    # liefert ODER deutlich weniger zerhackte Woerter enthaelt. Beides
    # RELATIV – absolute Schwellen waeren fuer fremde Dokumente geraten.
    if t_ocr > t_text * 1.1 or g_ocr < g_text * 0.8:
        return "ocr"
    return "text"


def _ocr_seiten(pfad: Path, indizes: list[int]):
    try:
        from pdf2image import convert_from_path  # noqa: PLC0415
    except ImportError as e:  # noqa: BLE001
        raise FormularFehler(
            "Fuer die Texterkennung fehlt pdf2image. Nachinstallieren: "
            "'sudo bash deploy/sandbox_python.sh' bzw. im venv "
            "'pip install pdf2image' (dazu die Programme 'pdftoppm' und "
            "'tesseract')."
        ) from e
    if not shutil.which("tesseract"):
        raise FormularFehler(
            "Das Programm 'tesseract' ist nicht installiert. Ohne es kann ein "
            "PDF mit beschaedigter Textebene nicht zuverlaessig gelesen werden "
            "(apt install tesseract-ocr tesseract-ocr-deu)."
        )

    from concurrent.futures import ThreadPoolExecutor  # noqa: PLC0415

    lang = _ocr_sprache()
    # Zwei Kerne bleiben frei, damit eine Extraktion den Webdienst nicht
    # aushungert (gleiche Regel wie beim Embedding-Modell).
    threads = max(1, min((os.cpu_count() or 2) - 2, 8))

    # BLOCKWEISE rendern, nicht Seite fuer Seite: poppler oeffnet und parst das
    # PDF bei jedem Aufruf komplett neu. Bei 54 Seiten waren das 54 Durchlaeufe.
    # Gerendert wird in DATEIEN (paths_only) statt in den Speicher – 54 Seiten
    # bei 200 dpi sind ueber 200 MB Bilddaten, und tesseract liest ohnehin von
    # der Platte.
    ergebnisse = []
    with tempfile.TemporaryDirectory(prefix="pdfform_") as tmp:
        for start in range(0, len(indizes), RENDER_BLOCK):
            block = indizes[start:start + RENDER_BLOCK]
            pfade = convert_from_path(
                str(pfad), dpi=DPI, first_page=block[0] + 1, last_page=block[-1] + 1,
                output_folder=tmp, fmt="png", paths_only=True,
            )
            paare = [(i, pfade[i - block[0]]) for i in block
                     if 0 <= i - block[0] < len(pfade)]

            def eine(paar):
                i, bild = paar
                with Image.open(bild) as im:
                    breite, hoehe = im.size
                return _woerter_ocr_datei(bild, lang), float(breite), float(hoehe), i + 1

            with ThreadPoolExecutor(max_workers=threads) as pool:
                ergebnisse.extend(pool.map(eine, paare))
            for bild in pfade:
                try:
                    os.unlink(bild)
                except OSError:
                    pass
    return [e for e in ergebnisse if e]


# ── Auswertung ───────────────────────────────────────────────────────────
def _boilerplate(spalten: list[str], zeilen: list[dict], mindest_anteil=0.8):
    """Felder, deren nicht-leerer Wert auf fast allen Seiten DERSELBE ist,
    sind Kopf-/Fusszeile (Umsatzsteuer-ID, Geschaeftsfuehrung) und kein
    Formularinhalt. Leere Felder bleiben – ein unausgefuelltes Formularfeld
    ist eine Aussage ueber das Formular.

    Verglichen wird AEHNLICHKEIT, nicht Gleichheit: dieselbe Fusszeile wird
    von der Texterkennung je Seite leicht anders gelesen ('Andreas Giebisch,
    Arnd Liman,' / 'Andreas Giebisch, ‚Arnd Liman,'). Ein Test auf exakte
    Gleichheit findet sie deshalb nicht.
    """
    raus = []
    n = len(zeilen)
    if n < 3:
        return raus
    for sp in spalten:
        gefuellt = [(z.get(sp) or "").strip() for z in zeilen if (z.get(sp) or "").strip()]
        if len(gefuellt) < n * mindest_anteil:
            continue
        kerne = [_normal(w)[:40] for w in gefuellt]
        haeufig, anzahl = Counter(kerne).most_common(1)[0]
        if not haeufig:
            continue
        if anzahl >= len(gefuellt) * mindest_anteil:
            raus.append((sp, Counter(gefuellt).most_common(1)[0][0]))
            continue
        # AEHNLICHKEIT nur bei LANGEN Werten. Kurze Eintraege, die sich bloss
        # in einer Ziffer unterscheiden ('Person 1' / 'Person 2', 'Weg 1' /
        # 'Weg 2'), erreichen sonst muehelos 0,8 und die halbe Tabelle
        # verschwindet als vermeintliche Fusszeile – im eigenen Test genau so
        # passiert. Eine echte Fusszeile ist lang.
        if min(len(k) for k in kerne) < BOILERPLATE_MIN_LAENGE:
            continue
        aehnlich = sum(1 for k in kerne
                       if SequenceMatcher(None, k, haeufig).ratio() >= 0.8)
        if aehnlich >= len(gefuellt) * mindest_anteil:
            raus.append((sp, Counter(gefuellt).most_common(1)[0][0]))
    return raus


def _dichter_block(felder: list[dict], faktor: float = 4.0):
    """Nur den zusammenhaengenden Feldblock behalten.

    Ein Formular hat seine Felder dicht beieinander (hier: Abstand 0,018 bis
    0,037 der Seitenhoehe). Briefkopf und Fusszeile stehen weit davon
    entfernt (Luecken von 0,16). Getrennt wird an Luecken, die ein Vielfaches
    des ueblichen Feldabstands betragen; von den entstehenden Bloecken
    gewinnt der mit den meisten Feldern.
    """
    if len(felder) < 4:
        return felder, []
    ys = [f["y"] for f in felder]
    abstaende = [b - a for a, b in zip(ys, ys[1:])]
    if not abstaende:
        return felder, []
    typisch = statistics.median(abstaende)
    if typisch <= 0:
        return felder, []
    bloecke, aktuell = [], [felder[0]]
    for vor, f, d in zip(felder, felder[1:], abstaende):
        if d > typisch * faktor:
            bloecke.append(aktuell)
            aktuell = [f]
        else:
            aktuell.append(f)
    bloecke.append(aktuell)
    if len(bloecke) == 1:
        return felder, []
    groesster = max(bloecke, key=len)
    draussen = [f for b in bloecke if b is not groesster for f in b]
    # Nur eingreifen, wenn der Hauptblock wirklich dominiert – sonst waere
    # ein zweispaltiges Formular halbiert.
    if len(groesster) < len(felder) * 0.6:
        return felder, []
    return groesster, draussen


def formular_auswerten(pfad: Path, quelle="auto", seiten_wahl=None,
                       mit_boilerplate=False):
    seiten, benutzte_quelle, gesamt = _pdf_seiten_woerter(pfad, quelle, seiten_wahl)
    if not seiten:
        raise FormularFehler("Aus dem PDF liess sich keine einzige Seite lesen.")

    seiten_zeilen, hoehen, breiten, nummern = [], [], [], []
    for woerter, breite, hoehe, nr in seiten:
        zs = zeilen_bilden(woerter)
        for z in zs:
            z["breite"] = breite
        seiten_zeilen.append(zs)
        hoehen.append(hoehe)
        breiten.append(breite)
        nummern.append(nr)

    # ── Die Grenze zwischen Beschriftungs- und Wertspalte wird GELERNT ──
    # Sie fest zu setzen geht schief: liegt sie zu weit rechts, gelten die
    # eingetragenen Werte als Beschriftungsspalte und die Tabelle bleibt leer;
    # zu weit links, und ein mehrzeiliger Wert verliert seine erste Zeile.
    # Gemessen an der echten Datei: Beschriftungen beginnen bei 11 % der
    # Blattbreite, Werte bei 31 %, Fortsetzungszeilen bei 21 %.
    grob = statistics.median([b * 0.33 for b in breiten])
    label_x0, wert_x0 = _spalten_messen(seiten_zeilen, breiten, grob)
    spanne = wert_x0 - label_x0
    grenze_rel = label_x0 + spanne * 0.60      # bis hierhin darf eine Beschriftung beginnen
    links_rel = label_x0 + spanne * 0.20       # links davon ist Fliesstext
    max_label_abs = [b * grenze_rel for b in breiten]

    felder = schablone_lernen(
        seiten_zeilen, hoehen,
        max_label_x=statistics.median(max_label_abs),
    )
    if not felder:
        raise FormularFehler(
            "Auf den Seiten wurde keine einzige Feldbeschriftung gefunden (erwartet "
            "wird 'Beschriftung:' am linken Rand). Das PDF ist vermutlich kein "
            "Formular mit einem Datensatz je Seite – benutze office_read."
        )

    x_min, x_max, x_fort = wertspalte_lernen(
        seiten_zeilen, breiten, statistics.median(max_label_abs))
    x_fort = min(x_fort, label_x0 + (wert_x0 - label_x0) * 0.5)

    # Nur Felder, die auf einem nennenswerten Teil der Seiten ERKANNT wurden.
    # Bewusst niedrig: die Schablone holt den Wert auch dort, wo die
    # Beschriftung unlesbar war – genau dafuer gibt es sie.
    schwelle = max(2, int(len(seiten_zeilen) * 0.25))
    kern = [f for f in felder if f["seiten"] >= schwelle]
    if not kern:
        kern = felder
    kern, ausserhalb = _dichter_block(kern)

    # Fenster fuer die Zuordnung: nach OBEN knapp unter dem Zeilenabstand
    # (sonst greift ein Feld auf den Wert des Feldes darueber), nach UNTEN
    # nur eine Kleinigkeit (ein Wert kann minimal tiefer sitzen als seine
    # Beschriftung).
    ys = sorted(f["y"] for f in kern)
    abstand = statistics.median([b - a for a, b in zip(ys, ys[1:])]) if len(ys) > 1 else 0.02
    fenster_rel = (abstand * 0.85, abstand * 0.15)

    # MEHRFACH VORKOMMENDE BESCHRIFTUNGEN: enthaelt eine Seite zwei Formulare,
    # traegt die Schablone 'Name' zweimal. Beide in dasselbe Feld zu schreiben
    # hiesse, den ersten Datensatz stillschweigend zu ueberschreiben – genau
    # der Datenverlust, den dieses Modul verhindern soll. Also eindeutige
    # Spalten, und der Bericht sagt es.
    doppelt = [n for n, c in Counter(f["name"] for f in kern).items() if c > 1]
    if doppelt:
        gesehen: Counter = Counter()
        for f in kern:
            gesehen[f["name"]] += 1
            if gesehen[f["name"]] > 1:
                f["name"] = f"{f['name']} ({gesehen[f['name']]})"

    zeilen_aus = []
    herkunft = Counter()
    for zs, hoehe, breite, nr in zip(seiten_zeilen, hoehen, breiten, nummern):
        satz = {"Seite": nr}
        frag = wertfragmente(zs, statistics.median(max_label_abs),
                             breite * links_rel, breite * (wert_x0 - 0.03))
        label_ys = [z["m"] for z in zs
                    if label_in_zeile(z, statistics.median(max_label_abs))]
        vergeben: set[int] = set()
        fenster = (fenster_rel[0] * hoehe, fenster_rel[1] * hoehe)
        versatz = seitenversatz(zs, hoehe, kern, statistics.median(max_label_abs))
        for f in kern:
            wert, woher = wert_holen(frag, label_ys, f, hoehe, breite,
                                     fenster, vergeben, x_max, versatz)
            satz[f["name"]] = wert
            herkunft[woher] += 1
        zeilen_aus.append(satz)

    spalten = [f["name"] for f in kern]
    ausgeblendet = [] if mit_boilerplate else _boilerplate(spalten, zeilen_aus)
    raus_namen = {s for s, _ in ausgeblendet}
    spalten = [s for s in spalten if s not in raus_namen]
    for z in zeilen_aus:
        for s in raus_namen:
            z.pop(s, None)

    return {
        "spalten": ["Seite"] + spalten,
        "zeilen": zeilen_aus,
        "quelle": benutzte_quelle,
        "seiten_gesamt": gesamt,
        "felder": kern,
        "ausgeblendet": ausgeblendet,
        "ausserhalb": [f["name"] for f in ausserhalb],
        "doppelte_felder": doppelt,
        "herkunft": dict(herkunft),
        "wertspalte": (x_min, x_max),
    }


# ── Werkzeug ─────────────────────────────────────────────────────────────
def _tabelle_text(spalten, zeilen, max_zeilen):
    breiten = {s: min(28, max(len(s), *(len(str(z.get(s, ""))) for z in zeilen[:max_zeilen]) or [len(s)]))
               for s in spalten}
    kopf = " | ".join(s[:breiten[s]].ljust(breiten[s]) for s in spalten)
    trenn = "-+-".join("-" * breiten[s] for s in spalten)
    aus = [kopf, trenn]
    for z in zeilen[:max_zeilen]:
        aus.append(" | ".join(str(z.get(s, ""))[:breiten[s]].ljust(breiten[s]) for s in spalten))
    return "\n".join(aus)


def _xlsx_schreiben(spalten, zeilen, ziel_name):
    from openpyxl import Workbook  # noqa: PLC0415
    from skills.office.main import _new_path, _ok  # noqa: PLC0415

    disk, fname, dl = _new_path(ziel_name, "xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Formulardaten"
    ws.append(spalten)
    for z in zeilen:
        ws.append([z.get(s, "") for s in spalten])
    # Spaltenbreite nach Inhalt, damit die Datei ohne Nacharbeit lesbar ist.
    for i, s in enumerate(spalten, 1):
        laenge = max([len(str(s))] + [len(str(z.get(s, ""))) for z in zeilen[:200]])
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(45, max(10, laenge + 2))
    ws.freeze_panes = "A2"
    wb.save(str(disk))
    return disk, fname, dl, _ok


class FormularExtraktTool(BaseTool):
    # Vom Dispatch generisch ausgewertet (sandbox.authorize_fs("read", …)).
    pfad_parameter = ("path",)
    ergebnis_max = ERGEBNIS_MAX

    @property
    def name(self) -> str:
        return "pdf_formular_extrakt"

    @property
    def description(self) -> str:
        return (
            "Liest ein FORMULAR-PDF aus, bei dem jede Seite einen Datensatz "
            "enthaelt ('Name:', 'Strasse:', 'Telefon:' …), und liefert eine "
            "Tabelle mit einer Zeile je Seite – auf Wunsch direkt als .xlsx. "
            "BENUTZE DAS immer, wenn aus einem mehrseitigen PDF gleichartige "
            "Angaben (Adressen, Kontakte, Antraege) gesammelt werden sollen. "
            "Die Zuordnung Beschriftung->Wert passiert ueber die Position auf "
            "der Seite; Text abzutippen oder aus dem Anhang-Text zu raten "
            "liefert nachweislich falsche Werte. Erkennt selbsttaetig, ob die "
            "Textebene brauchbar ist, und schaltet sonst auf Texterkennung um."
        )

    def parameters_schema(self) -> dict:
        return {
            "type": "OBJECT",
            "properties": {
                "path": {"type": "STRING", "description": "Dateiname aus data/documents/, /api/documents/-URL oder Serverpfad der PDF."},
                "seiten": {"type": "STRING", "description": "Optional, z.B. '1-20' oder '1,5,9-12'. Weglassen = alle Seiten."},
                "quelle": {"type": "STRING", "enum": ["auto", "text", "ocr"], "description": "'auto' (Vorgabe) prueft die Textebene und schaltet bei Beschaedigung auf Texterkennung um."},
                "als_datei": {"type": "BOOLEAN", "description": "true (Vorgabe) schreibt zusaetzlich eine .xlsx und gibt den Download-Link zurueck."},
                "dateiname": {"type": "STRING", "description": "Name der Ergebnisdatei ohne Endung."},
                "mit_kopfzeilen": {"type": "BOOLEAN", "description": "true behaelt Felder, die auf allen Seiten denselben Wert haben (Briefkopf/Fusszeile). Vorgabe false."},
            },
            "required": ["path"],
        }

    async def execute(self, path: str = "", seiten: str = "", quelle: str = "auto",
                      als_datei: bool = True, dateiname: str = "",
                      mit_kopfzeilen: bool = False, **kwargs) -> str:
        from skills.office.main import _resolve_existing  # noqa: PLC0415

        unbekannt = [k for k in kwargs if not k.startswith("_")]
        if unbekannt:
            # Ein Werkzeug, das unbekannte Parameter wortlos verwirft, meldet
            # Erfolg fuer etwas, das nie passiert ist (Vorfall 2026-08-19).
            return (f"Fehler: unbekannte Parameter {unbekannt}. Erlaubt sind: "
                    f"path, seiten, quelle, als_datei, dateiname, mit_kopfzeilen.")

        p = _resolve_existing(path)
        if not p:
            return f"Fehler: Datei nicht gefunden: {path}"
        if p.suffix.lower() != ".pdf":
            return f"Fehler: '{p.suffix}' ist kein PDF. Dieses Werkzeug liest nur .pdf."
        if quelle not in ("auto", "text", "ocr"):
            return f"Fehler: quelle='{quelle}' unbekannt (auto|text|ocr)."

        import asyncio  # noqa: PLC0415
        try:
            # Rendern und OCR sind rechenintensiv und blockierend – im
            # Event-Loop wuerden sie den ganzen Dienst anhalten (Vorfall
            # 2026-08-11: 20 s Freeze durch einen blockierenden Aufruf).
            erg = await asyncio.to_thread(
                formular_auswerten, p, quelle, seiten, mit_kopfzeilen)
        except FormularFehler as e:
            return f"Fehler: {e}"
        except Exception as e:  # noqa: BLE001
            return f"Fehler beim Auswerten von '{p.name}': {e}"

        spalten, zeilen = erg["spalten"], erg["zeilen"]
        gefuellt = sum(1 for z in zeilen for s in spalten
                       if s != "Seite" and str(z.get(s, "")).strip())
        moeglich = max(1, len(zeilen) * max(1, len(spalten) - 1))

        bericht = [
            f"{len(zeilen)} Seiten ausgewertet (PDF hat {erg['seiten_gesamt']}), "
            f"{len(spalten) - 1} Felder erkannt.",
            f"Textquelle: " + ("Texterkennung (OCR) – die Textebene des PDFs ist "
                               "beschaedigt" if erg["quelle"] == "ocr" else
                               "Textebene des PDFs"),
            f"Ausgefuellt: {gefuellt} von {moeglich} moeglichen Werten "
            f"({gefuellt * 100 // moeglich} %). Leere Felder waren im Formular leer.",
            "Felder: " + " · ".join(s for s in spalten if s != "Seite"),
        ]
        if erg.get("doppelte_felder"):
            bericht.append(
                "ACHTUNG: Diese Beschriftungen kommen je Seite MEHRFACH vor ("
                + ", ".join(erg["doppelte_felder"][:6])
                + ") – vermutlich stehen mehrere Datensaetze auf einer Seite. "
                "Sie wurden zu getrennten Spalten '… (2)', '… (3)' gemacht; "
                "eine Zeile der Tabelle enthaelt dann mehrere Datensaetze.")
        if erg.get("ausserhalb"):
            bericht.append(
                "Ausserhalb des Formularblocks und daher nicht uebernommen: "
                + ", ".join(erg["ausserhalb"][:8])
                + " (Briefkopf/Fusszeile).")
        if erg["ausgeblendet"]:
            bericht.append(
                "Als Briefkopf/Fusszeile ausgeblendet (auf allen Seiten gleich): "
                + "; ".join(f"{n} = {w[:40]}" for n, w in erg["ausgeblendet"])
                + ". Mit mit_kopfzeilen=true bleiben sie drin.")
        varianten = [f for f in erg["felder"] if len(f["varianten"]) > 1]
        if varianten and erg["quelle"] == "ocr":
            bericht.append(
                "Bei diesen Feldern hat die Texterkennung die Beschriftung je "
                "Seite unterschiedlich gelesen; der haeufigste Name wurde "
                "gewaehlt: " + ", ".join(f["name"] for f in varianten[:6]))
        fehlt = erg["herkunft"].get("fehlt", 0)
        if fehlt:
            bericht.append(
                f"ACHTUNG: {fehlt} Feldpositionen liessen sich auf ihrer Seite nicht "
                f"wiederfinden – dort ist der Wert leer, obwohl im PDF etwas stehen "
                f"kann. Betroffene Seiten stichprobenartig im PDF pruefen.")
        bericht.append(
            "Die Werte stammen unveraendert aus dem PDF. Tippe sie NICHT ab und "
            "rechne nichts hinzu – bei Zweifeln die Seitenzahl nennen.")

        vorschau = _tabelle_text(spalten, zeilen, VORSCHAU_ZEILEN)
        rest = len(zeilen) - VORSCHAU_ZEILEN
        kopf = "\n".join(bericht) + "\n\n" + vorschau
        if rest > 0:
            kopf += f"\n… {rest} weitere Zeilen (nur die Vorschau ist gekuerzt, die Datei enthaelt alle)."

        if not als_datei:
            return kopf

        try:
            disk, fname, dl, _ok = _xlsx_schreiben(
                spalten, zeilen, dateiname or (p.stem + "_Extrakt"))
        except Exception as e:  # noqa: BLE001
            return kopf + f"\n\n[Die .xlsx konnte nicht geschrieben werden: {e}]"
        return _ok(dl, fname, disk, extra=kopf)


def get_pdf_formular_tools():
    return [FormularExtraktTool()]
