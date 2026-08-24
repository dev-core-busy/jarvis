"""Tabellen-Werkzeuge: eine BESTEHENDE .xlsx ansehen und bearbeiten.

WARUM ES DIESES MODUL GIBT (Vorfall ECHT, 2026-08-19)
=====================================================
Ein Short-Track "Tabellen zusammenfuehren" sollte aus einer Master- und einer
Slave-Tabelle eine erweiterte Master-Tabelle bauen. Herausgekommen sind fuenf
Dateien, davon eine LEER, eine mit ZWEI Zeilen und eine mit den Daten eines
einzigen Jahres. Nachgemessen an der echten Eingabedatei:

    13 Blaetter | 362.195 Zellen | 32.779 Formeln | 763 KB
    office_read erzeugt daraus 1.265.130 Zeichen Text
      -> Deckel im Werkzeug (20.000):        1,58 % bleiben
      -> Deckel im Agenten  ( 5.000):        0,40 % bleiben
    Das Modell sah Blatt '2004', angeschnitten mitten in einer Summenzeile.

Danach hatte es nur `office_create_excel`, um ein Ergebnis zu erzeugen – und
das baut eine Tabelle NEU AUF aus Werten, die das Modell als JSON-Argument
tippt. Fuer 362.195 Zellen ist das nicht "ungenau", sondern strukturell
unmoeglich; und die 32.779 Formeln des Masters waeren selbst dann verloren,
wenn jede Zahl stimmte.

DIE REGEL, DIE DIESES MODUL DURCHSETZT
======================================
**Die Daten gehen NIE durch das Sprachmodell.** Das Modell bekommt die
STRUKTUR (Blaetter, Kopfzeilen, Typen, ein paar Beispielzeilen) und beschreibt
dann die TRANSFORMATION (welche Spalte, welcher Schluessel). Ausgefuehrt wird
sie hier im Backend auf der echten Datei.

Daraus folgen drei Eigenschaften, die jedes Werkzeug hier einhaelt:

1. **Die Ausgabe ist BEGRENZT und die Begrenzung wird BEZIFFERT.** Nie ein
   stilles "…". Wer nicht weiss, wie viel ihm fehlt, antwortet auf einem
   Ausschnitt und haelt ihn fuer das Ganze – genau das ist im Vorfall passiert,
   und die Kuerzungsmeldung log das Modell zusaetzlich an (sie nannte "5.000
   von 20.014", weil `office_read` vorher schon still von 1.265.130 auf 20.000
   gekuerzt hatte).
2. **Fehlschlaege sind LAUT.** Der leere Aufruf im Vorfall hiess
   ``{" sheets": "..."}`` – mit fuehrendem Leerzeichen. Das landete in
   ``**kwargs``, wurde wortlos verworfen, und das Werkzeug meldete "erstellt".
   Ein Werkzeug, das bei einer unbrauchbaren Eingabe Erfolg meldet, nimmt dem
   Modell die einzige Chance, sich zu korrigieren.
3. **Das Layout bleibt.** Bearbeitet wird die GEOEFFNETE Originaldatei, nicht
   ein Neubau. Formeln, Spaltenbreiten, verbundene Zellen und Zahlenformate
   ueberleben damit.

GRENZE, DIE MAN KENNEN MUSS
===========================
openpyxl verliert beim Oeffnen-und-Speichern **Diagramme, Bilder und
Pivot-Tabellen**. Das laesst sich hier nicht reparieren – deshalb wird es
ERKANNT und im Ergebnis ausdruecklich gemeldet, statt es zu verschweigen.
"""

from __future__ import annotations

import re
from pathlib import Path

from backend.tools.base import BaseTool


# ── Deckel ──────────────────────────────────────────────────────────────────
# Alle Grenzen stehen hier beieinander und tauchen im ERGEBNISTEXT auf, sobald
# sie greifen. Eine Grenze, die man nur im Quelltext sieht, ist fuer den
# Aufrufer eine unsichtbare Luege.
SPALTEN_IM_KOPF = 60        # so viele Kopf-Spalten je Blatt werden benannt
# ZEICHEN-Deckel fuer die Spaltennamen. Die Namen sind STRUKTUR, nicht Daten –
# und Struktur ist genau das, was diese Werkzeuge liefern sollen. Gemessen an
# einer echten Datei (2026-08-24, ECHT): ein Blatt mit 254 Laborcode-Spalten
# ergab bei der alten festen Grenze "+182 weitere benannte Spalten" – die 182
# Namen bekam man NIRGENDS her, waehrend der Hinweis "Mit 'spalten' gezielt
# auswaehlen" genau sie verlangte. Ein Zirkelschluss, aus dem das Modell nur
# durch Durchprobieren herauskam: 24 Leseaufrufe, 534 s, kein Ergebnis. Der
# ganze inspect-Text war dabei 2.630 von 14.000 erlaubten Zeichen lang – es war
# reichlich Platz, die Kappung kam allein von der festen Zahl.
KOPF_TEXT_MAX = 4000
BEISPIEL_ZEILEN = 3         # Beispielzeilen je Blatt in `xlsx_inspect`
BEISPIEL_SPALTEN = 25       # Spalten je Beispielzeile
LESE_ZEILEN_VORGABE = 30    # `xlsx_read_range` ohne Angabe
LESE_ZEILEN_MAX = 200       # harte Obergrenze, auch auf Wunsch
ZELLTEXT_MAX = 120          # ein einzelner Zellwert in der Anzeige
AUSGABE_MAX = 14000         # Notbremse fuer den gesamten Ergebnistext


def _kurz(wert, grenze: int = ZELLTEXT_MAX) -> str:
    """Ein Zellwert als kurzer Text. Leere Zelle = leerer String."""
    if wert is None:
        return ""
    s = str(wert)
    if len(s) > grenze:
        return s[: grenze - 1] + "…"
    return s


def _spaltenbuchstabe(n: int) -> str:
    """1 -> A, 27 -> AA (openpyxl hat das, aber ohne Import-Kosten hier)."""
    s = ""
    while n > 0:
        n, rest = divmod(n - 1, 26)
        s = chr(65 + rest) + s
    return s


# Excel kennt hoechstens 16.384 Spalten (XFD) – also nie mehr als drei
# Buchstaben. Beide Grenzen sind NOETIG und nicht kosmetisch: ohne sie gilt
# JEDES Wort als Spaltenbezeichnung. Im ersten Testlauf loeste der Spaltenname
# "Unbekannt" zu Spalte 4.498.495.991.152 auf, und der Merge starb beim
# Speichern mit "Invalid column index" – bei einem kuerzeren Wort haette er
# stattdessen klaglos in eine voellig falsche Spalte geschrieben.
_MAX_SPALTE = 16384


def _buchstabe_zu_index(s: str) -> int | None:
    """A -> 1, AA -> 27. None, wenn es keine gueltige Spaltenangabe ist."""
    s = (s or "").strip().upper()
    if not s or len(s) > 3 or not s.isalpha() or not s.isascii():
        return None
    n = 0
    for z in s:
        n = n * 26 + (ord(z) - 64)
    return n if 1 <= n <= _MAX_SPALTE else None


def _kopf_text(paare) -> tuple[str, int]:
    """Spaltennamen als 'A=Name | B=Name | …' – bis KOPF_TEXT_MAX Zeichen.

    Liefert ``(text, nicht_gezeigt)``. Gedeckelt wird nach ZEICHEN und nicht
    nach Anzahl: eine Tabelle mit 254 kurzen Kuerzeln passt vollstaendig hinein
    (rund 2,3 KB), eine mit 1.000 langen Ueberschriften wird ehrlich gekuerzt.
    Der Aufrufer MUSS die Restzahl ausgeben – eine unvollstaendige Spaltenliste,
    die sich fuer vollstaendig ausgibt, ist schlimmer als eine kurze.
    """
    teile: list[str] = []
    laenge = 0
    for nr, (i, k) in enumerate(paare):
        t = f"{_spaltenbuchstabe(i)}={_kurz(k, 40)}"
        if teile and laenge + len(t) + 3 > KOPF_TEXT_MAX:
            return " | ".join(teile), len(paare) - nr
        teile.append(t)
        laenge += len(t) + 3
    return " | ".join(teile), 0


def _deckeln(text: str) -> str:
    """Notbremse fuer die Gesamtausgabe – MIT Angabe, was fehlt.

    Greift praktisch nie (die Einzeldeckel oben sind enger), aber eine Tabelle
    mit 3.000 Blaettern gibt es irgendwo. Stilles Abschneiden ist genau der
    Fehler, der diesen Vorfall verursacht hat.
    """
    if len(text) <= AUSGABE_MAX:
        return text
    return (text[:AUSGABE_MAX]
            + f"\n\n[… gekuerzt: {AUSGABE_MAX} von {len(text)} Zeichen gezeigt. "
              f"Frage mit xlsx_read_range gezielt nach dem fehlenden Teil, "
              f"statt den Rest zu erraten.]")


# ── Datei oeffnen ───────────────────────────────────────────────────────────

# Endungen, die als TABELLE gelesen werden, obwohl sie keine Mappe sind.
# Vorgabe des Nutzers 2026-08-24: "CSV muss natuerlich auch erlaubt werden."
# Der Anlass steht in CLAUDE.md – die Ablage "Tabellen zusammenfuehren" erlaubt
# `csv`, und `xlsx_merge` wies die Slave-CSV mit "ist keine Excel-Datei" ab.
# Das Modell zog daraus die Folgerung "der Slave-Juli ist LEER" und setzte 27
# Spalten auf 0: eine Fehlermeldung, die nicht gelesen wurde, ist schlimmer als
# eine fehlende Funktion.
_CSV_ENDUNGEN = (".csv", ".tsv", ".txt")


def _csv_wert(roh):
    """Textzelle in eine Zahl wandeln – aber nur, wenn nichts verloren geht.

    Eine CSV kennt keine Typen; ohne Umwandlung landet "3282" als TEXT in der
    Master-Zelle, Excel richtet es links aus und jede Summenformel darueber
    rechnet es als 0. Umgekehrt darf die Umwandlung keine Bedeutung loeschen:

    * **Fuehrende Null ist Inhalt, nicht Formatierung** – PLZ ``02625``,
      Kundennummer ``00083``, Artikelnummer. Als Zahl waeren sie zerstoert.
    * Sehr lange Ziffernfolgen (IBAN, Belegnummer) verlieren als float ihre
      letzten Stellen – sie bleiben Text.
    * Deutsche Schreibweise laeuft ueber ``chart.parse_number``: ``float("1.234")``
      ergaebe 1.234 statt 1234 und verfaelschte jede Tabelle um Faktor 1000.
    """
    if roh is None:
        return None
    s = str(roh).strip()
    if not s:
        return None
    if len(s) > 1 and s[0] == "0" and s[1] not in ",.":
        return roh
    if len(re.sub(r"\D", "", s)) > 15:
        return roh
    try:
        from backend.tools.chart import parse_number  # noqa: PLC0415
    except Exception:  # noqa: BLE001
        return roh
    v = parse_number(s)
    if v is None:
        return roh
    try:
        return int(v) if float(v).is_integer() and abs(v) < 2 ** 53 else v
    except Exception:  # noqa: BLE001
        return roh


def _csv_mappe(p):
    """Baut aus einer CSV/TSV eine Mappe IM SPEICHER.

    Damit arbeiten `_kz`, `_kopfzeile`, `iter_rows` und alle Werkzeuge
    unveraendert weiter – die Alternative waere ein zweiter Codepfad je
    Werkzeug gewesen, und der laeuft erfahrungsgemaess auseinander.

    Das Blatt heisst wie die Datei (ohne Endung), damit `blatt=`-Angaben und
    Fehlermeldungen etwas Sinnvolles nennen.
    """
    try:
        from openpyxl import Workbook  # noqa: PLC0415
        from backend.tools.chart import csv_zeilen  # noqa: PLC0415
    except Exception as e:  # noqa: BLE001
        raise TabellenFehler(f"CSV-Unterstuetzung nicht verfuegbar ({e}).") from e
    try:
        zeilen = csv_zeilen(p)
    except Exception as e:  # noqa: BLE001
        raise TabellenFehler(f"CSV nicht lesbar: {e}") from e
    if not zeilen:
        raise TabellenFehler(f"'{p.name}' enthaelt keine Zeilen.")
    wb = Workbook()
    ws = wb.active
    ws.title = (p.stem or "CSV")[:31]
    for row in zeilen:
        ws.append([_csv_wert(c) for c in row])
    return wb


def _oeffnen(path: str, schreibend: bool = False):
    """Loest den Pfad auf und oeffnet die Mappe. Gibt (workbook, Path) zurueck.

    Wirft ``TabellenFehler`` mit Klartext – die Aufrufer geben den unveraendert
    an das Modell weiter, damit es sich korrigieren kann.

    ``schreibend`` laedt OHNE ``read_only`` (nur so laesst sich speichern) und
    mit ``data_only=False``, damit Formeln als Formeln erhalten bleiben. Mit
    ``data_only=True`` wuerde jede Formel beim Speichern durch ihren zuletzt
    berechneten Wert ERSETZT – die Mappe waere still zu einer Wertetabelle
    geworden.
    """
    from skills.office.main import _resolve_existing  # noqa: PLC0415

    p = _resolve_existing(path)
    if not p:
        raise TabellenFehler(f"Datei nicht gefunden: {path!r}")
    if p.suffix.lower() in _CSV_ENDUNGEN:
        # SCHREIBEND geht nicht: eine CSV hat kein Layout, keine Formeln und
        # keine Blaetter – "bearbeiten und Formeln behalten" ist dort keine
        # Zusage, die man halten kann. Die Meldung nennt den Weg, statt nur
        # abzulehnen.
        if schreibend:
            raise TabellenFehler(
                f"'{p.name}' ist eine CSV – sie kann als QUELLE dienen "
                f"(xlsx_inspect, xlsx_read_range, xlsx_merge als 'slave'), aber "
                f"nicht bearbeitet werden: eine CSV hat kein Layout und keine "
                f"Formeln. Die zu befuellende Tabelle (master/path) muss eine "
                f".xlsx sein."
            )
        return _csv_mappe(p), p
    if p.suffix.lower() not in (".xlsx", ".xlsm"):
        raise TabellenFehler(
            f"'{p.name}' ist weder eine Excel-Datei (.xlsx/.xlsm) noch eine "
            f"Texttabelle (.csv/.tsv), sondern '{p.suffix or 'ohne Endung'}'. "
            f"Fuer .xls muss die Datei erst konvertiert werden."
        )
    try:
        from openpyxl import load_workbook  # noqa: PLC0415
    except Exception as e:  # noqa: BLE001
        raise TabellenFehler(f"openpyxl nicht verfuegbar ({e}).") from e

    try:
        wb = load_workbook(str(p), read_only=not schreibend, data_only=False)
    except Exception as e:  # noqa: BLE001
        raise TabellenFehler(f"Datei konnte nicht geoeffnet werden: {e}") from e
    return wb, p


class TabellenFehler(Exception):
    """Eingabefehler, der dem Modell im Klartext gemeldet wird."""


def _blatt(wb, name: str):
    """Waehlt ein Blatt. Ohne Angabe das erste; unbekannter Name = Fehler.

    Der Fehler NENNT die vorhandenen Blattnamen. Ohne diese Liste raet das
    Modell weiter (im Vorfall hat es sich Blattnamen ausgedacht).
    """
    namen = [ws.title for ws in wb.worksheets]
    if not namen:
        raise TabellenFehler("Die Mappe enthaelt kein einziges Blatt.")
    if not (name or "").strip():
        return wb.worksheets[0]
    gesucht = str(name).strip()
    for ws in wb.worksheets:
        if ws.title == gesucht:
            return ws
    for ws in wb.worksheets:              # tolerant: Gross/Klein, Leerzeichen
        if ws.title.strip().lower() == gesucht.lower():
            return ws
    raise TabellenFehler(
        f"Blatt {gesucht!r} gibt es nicht. Vorhandene Blaetter: "
        + ", ".join(repr(n) for n in namen[:40])
        + (f" … (+{len(namen) - 40} weitere)" if len(namen) > 40 else "")
    )


# ── Kopfzeile und Spalten-Aufloesung ────────────────────────────────────────

def _kopfzeile(ws, zeile: int = 1) -> list[str]:
    """Liest die Kopfzeile als Liste von Texten (Index 0 = Spalte A)."""
    for i, row in enumerate(ws.iter_rows(min_row=zeile, max_row=zeile,
                                         values_only=True), start=zeile):
        return ["" if c is None else str(c).strip() for c in row]
    return []


# Wie viele Zeilen am Blattanfang fuer die Kopfzeilen-Erkennung angesehen werden.
KOPF_SUCHTIEFE = 12


def _hat_beschriftungen(row) -> bool:
    """Enthaelt die Zeile echte Spaltennamen?

    Echt heisst: Text, der WEDER eine Formel (``=B1``) NOCH eine reine Zahl ist.
    Die Formel-Bedingung stammt aus der echten Datei von ECHT: in den Blaettern
    2019-2026 ist die wiederholte Kopfzeile in Zeile 3 keine Beschriftung,
    sondern ein VERWEIS auf Zeile 1 (``=B1``, ``=C1``, …). Weil die Mappe mit
    ``data_only=False`` geoeffnet wird (sonst verlieren wir beim Speichern alle
    Formeln), steht dort der Formeltext – als Spaltenname unbrauchbar.
    """
    gefuellt = [c for c in row if c is not None and str(c).strip() != ""]
    if len(gefuellt) < 2:
        return False
    echte = [c for c in gefuellt
             if isinstance(c, str)
             and not c.lstrip().startswith("=")
             and not c.strip().replace(".", "").replace(",", "").isdigit()]
    return len(echte) / len(gefuellt) >= 0.5


def _kopfzeile_raten(ws) -> tuple[int, int]:
    """Ermittelt (Beschriftungszeile, erste Datenzeile).

    DASS DAS ZWEI VERSCHIEDENE DINGE SIND, hat erst die echte Datei gezeigt:
    in Blatt '2019' stehen die Spaltennamen in Zeile 1, Zeile 3 wiederholt sie
    als FORMEL (``=B1``), und die Daten beginnen in Zeile 4. Wer beides
    gleichsetzt, liest entweder ``=B1`` als Spaltennamen (Zeile 3) oder haelt
    die Zeilen 2 und 3 fuer Daten (Zeile 1).

    WARUM DAS NOETIG IST – an der echten Datei von ECHT gemessen: die 13 Blaetter
    sind UNTERSCHIEDLICH gebaut.

        Blatt 2004  Z1 = Kopfzeile,          Daten ab Z2
        Blatt 2015  Z1 = Nummerncodes, Z2 leer, Z3 = Kopfzeile,  Daten ab Z4
        Blatt 2019  Z1 = Kopfzeile, Z2 leer, Z3 = Kopfzeile (wiederholt, mit
                    Abschnittsnamen in Spalte A), Daten ab Z4

    Mit der festen Vorgabe "Zeile 1" liest man in Blatt 2015 also eine Liste von
    Nummern als Spaltennamen – und findet anschliessend keinen einzigen
    Schluessel. Das Modell muesste das von sich aus bemerken und ``kopfzeile``
    mitgeben; darauf ist kein Verlass.

    DER ANKER IST DER DATENANFANG, NICHT DIE BESCHRIFTUNG. Gesucht wird die erste
    ueberwiegend NUMERISCHE Zeile; die Kopfzeile ist die letzte nicht-leere Zeile
    darueber. Das trifft alle drei Bauformen oben – und bei zwei Kopfzeilen
    (2019/2026) die UNTERE, was richtig ist: sie steht unmittelbar ueber den
    Daten, und die Zeilen dazwischen wuerden sonst als Datenzeilen mitgelesen.

    ENTSCHEIDEND IST DER TYP, NICHT DER AUSSEHEN-EINDRUCK: gezaehlt wird nur
    ``int``/``float``. In Blatt 2015 stehen in Zeile 1 Werte wie
    ``"00000000083"`` – als TEXT gespeichert. Wer sie als Zahl zaehlt, haelt
    Zeile 1 fuer den Datenanfang und landet wieder bei der falschen Kopfzeile.
    """
    zeilen: list[tuple[int, tuple]] = []
    try:
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=KOPF_SUCHTIEFE,
                                             max_col=40, values_only=True), start=1):
            zeilen.append((i, row))
    except Exception:  # noqa: BLE001
        return 1, 2

    erste_daten = 0
    for i, row in zeilen:
        gefuellt = [c for c in row if c is not None and str(c).strip() != ""]
        if len(gefuellt) < 3:          # zu duenn, um den Datenanfang zu belegen
            continue
        zahlen = [c for c in gefuellt
                  if isinstance(c, (int, float)) and not isinstance(c, bool)]
        if len(zahlen) / len(gefuellt) >= 0.5:
            erste_daten = i
            break

    if erste_daten <= 1:
        # Keine Datenzeile gefunden (reine Texttabelle) oder die Daten beginnen
        # schon in Zeile 1 – dann ist Zeile 1 die beste verfuegbare Annahme.
        return 1, 2

    # Die BESCHRIFTUNGSZEILE ist die unterste Zeile oberhalb der Daten, die
    # wirklich Namen enthaelt. Eine Formel-Wiederholung (``=B1``) wird dabei
    # uebersprungen und die darueberliegende echte Beschriftung genommen.
    ueber = list(reversed(zeilen[: erste_daten - 1]))
    for i, row in ueber:
        if _hat_beschriftungen(row):
            return i, erste_daten
    # Keine brauchbare Beschriftung gefunden: die letzte nicht-leere Zeile ist
    # immer noch die beste Annahme (besser als blind Zeile 1).
    for i, row in ueber:
        if any(c is not None and str(c).strip() != "" for c in row):
            return i, erste_daten
    return 1, erste_daten


def _kz(ws, vorgabe) -> tuple[int, int, bool]:
    """Loest die Kopfzeilen-Angabe auf: (Beschriftungszeile, erste Datenzeile,
    automatisch_erkannt).

    ``0`` bzw. eine fehlende Angabe heisst AUTOMATISCH. Das ist ein bewusster
    Sentinel und keine Falsyness-Pruefung: Zeile 0 gibt es in Excel nicht, und
    ein ausdrueckliches ``kopfzeile: 1`` muss von "nicht angegeben"
    unterscheidbar bleiben – sonst wuerde die Erkennung eine bewusste Vorgabe
    ueberstimmen.

    Bei ausdruecklicher Angabe beginnen die Daten unmittelbar darunter: wer die
    Zeile selbst nennt, meint den einfachen Aufbau.
    """
    try:
        n = int(vorgabe or 0)
    except Exception:  # noqa: BLE001
        n = 0
    if n >= 1:
        return n, n + 1, False
    kopf, daten = _kopfzeile_raten(ws)
    return kopf, daten, True


def _kz_text(zeile: int, automatisch: bool, daten_ab: int = 0) -> str:
    t = f"Zeile {zeile}" + (" (automatisch erkannt)" if automatisch
                            else " (vorgegeben)")
    # Den Datenanfang nur nennen, wenn er NICHT direkt darunter liegt – sonst
    # ist es Rauschen. Weicht er ab, ist er die wichtigere der beiden Zahlen.
    if daten_ab and daten_ab != zeile + 1:
        t += f", Daten ab Zeile {daten_ab}"
    return t


def _spalte_aufloesen(kopf: list[str], bez) -> int | None:
    """Findet eine Spalte ueber Kopfzeilen-Namen, Buchstabe oder 1-basierte Zahl.

    Bewusst DREI Schreibweisen: das Modell benutzt mal 'Monat', mal 'B', mal 2 –
    und jede Fehlinterpretation schreibt Daten in die falsche Spalte. Ein
    Vergleich, der alle drei kennt, ist billiger als ein falsch befuellter
    Master.

    Reihenfolge ist Absicht: der KOPFZEILEN-NAME gewinnt. Eine Tabelle mit einer
    Spalte namens "B" gibt es; wer sie meint, meint fast nie Spalte 2.

    DIE BREITENGRENZE IST DER EIGENTLICHE SCHUTZ, nicht die Laengenpruefung in
    ``_buchstabe_zu_index``. Ein Wort aus ein bis drei Buchstaben IST eine
    syntaktisch gueltige Spaltenangabe: "Ort" ergibt Spalte 10.628. Ohne die
    Grenze wuerde ein Tippfehler im Spaltennamen also nicht als Fehler auffallen,
    sondern still in eine Spalte weit ausserhalb der Tabelle schreiben – und
    genau solche stillen Treffer haben den Vorfall vom 2026-08-19 ausgemacht.
    Ein Verweis JENSEITS der Kopfzeile ist praktisch immer ein missverstandener
    Name; er wird deshalb abgelehnt, damit der Aufrufer eine Fehlermeldung mit
    der echten Kopfzeile bekommt.
    """
    if bez is None:
        return None
    s = str(bez).strip()
    if not s:
        return None
    for i, k in enumerate(kopf, start=1):          # 1) exakter Kopfzeilen-Name
        if k == s:
            return i
    for i, k in enumerate(kopf, start=1):          # 2) tolerant
        if k.strip().lower() == s.lower():
            return i

    breite = len(kopf)
    idx = _buchstabe_zu_index(s)                   # 3) Spaltenbuchstabe
    if idx is None and s.isdigit():                # 4) 1-basierte Nummer
        n = int(s)
        idx = n if 1 <= n <= _MAX_SPALTE else None
    if idx is None:
        return None
    if breite and idx > breite:
        return None
    return idx


def _spalten_bericht(kopf: list[str], fehlend: list[str]) -> str:
    """Fehlermeldung fuer nicht gefundene Spalten – MIT der echten Kopfzeile."""
    vorhanden = [k for k in kopf if k][:SPALTEN_IM_KOPF]
    rest = max(0, len([k for k in kopf if k]) - len(vorhanden))
    return (
        "Diese Spalten gibt es nicht: " + ", ".join(repr(f) for f in fehlend)
        + ".\nVorhandene Spalten (Kopfzeile): "
        + ", ".join(repr(v) for v in vorhanden)
        + (f" … (+{rest} weitere)" if rest else "")
    )


# ── Verlustwarnung ──────────────────────────────────────────────────────────

def _verluste(wb) -> list[str]:
    """Was ein openpyxl-Rundlauf NICHT ueberlebt – erkannt, nicht verschwiegen.

    openpyxl liest Diagramme, Bilder und Pivot-Tabellen nicht vollstaendig ein
    und schreibt sie folglich nicht zurueck. Das ist eine Eigenschaft der
    Bibliothek und hier nicht reparierbar. Es zu VERSCHWEIGEN waere aber die
    schlimmere Variante: der Benutzer oeffnet die Datei und seine Auswertung
    ist weg, ohne dass irgendwo etwas davon stand.
    """
    hin = []
    try:
        d = sum(len(getattr(ws, "_charts", []) or []) for ws in wb.worksheets)
        if d:
            hin.append(f"{d} Diagramm(e)")
    except Exception:  # noqa: BLE001
        pass
    try:
        b = sum(len(getattr(ws, "_images", []) or []) for ws in wb.worksheets)
        if b:
            hin.append(f"{b} Bild(er)")
    except Exception:  # noqa: BLE001
        pass
    try:
        p = sum(len(getattr(ws, "_pivots", []) or []) for ws in wb.worksheets)
        if p:
            hin.append(f"{p} Pivot-Tabelle(n)")
    except Exception:  # noqa: BLE001
        pass
    return hin


def _verlust_hinweis(verluste: list[str]) -> str:
    if not verluste:
        return ""
    return ("\n\n⚠ HINWEIS_AN_NUTZER: Die Originaldatei enthaelt "
            + ", ".join(verluste)
            + ". Diese Elemente gehen beim Bearbeiten VERLOREN (Grenze der "
              "verwendeten Bibliothek). Formeln, Spaltenbreiten, verbundene "
              "Zellen und Zahlenformate bleiben erhalten.")


def _weiterarbeiten_hinweis(fname: str) -> str:
    """Sagt, wie der NAECHSTE Schritt auf diesem Stand aufsetzt.

    GEMELDET VON ECHT (2026-08-24): ein Lauf schrieb die Juli-Werte in fuenf
    Batches und rief `xlsx_edit` fuenfmal auf – jedes Mal mit ``path`` auf der
    ORIGINAL-Quelle. Jeder Aufruf legt aber ueber ``_new_path()`` eine eigene
    Ergebnisdatei an. Ergebnis: fuenf Dateien mit gleichem Namen, jede mit nur
    46-53 der 194 Werte, keine davon brauchbar; die Reihe baute nicht
    aufeinander auf.

    Der Hinweis ist deshalb keine Kosmetik, sondern die Bedingung dafuer, dass
    mehrschrittiges Bearbeiten ueberhaupt ein vollstaendiges Ergebnis liefert.
    Er nennt die ``/api/documents/``-URL, weil ``path`` genau die annimmt.
    """
    return ("\n\nWEITER AN DIESEM STAND: fuer zusaetzliche Aenderungen 'path' auf "
            "/api/documents/%s setzen. Wer erneut die Ausgangsdatei bearbeitet, "
            "verliert die eben geschriebenen Aenderungen und erzeugt einen "
            "zweiten Teilstand." % fname)


# ═══════════════════════════════════════════════════════════════════════════
# xlsx_inspect – STRUKTUR statt Inhalt
# ═══════════════════════════════════════════════════════════════════════════

class InspectTool(BaseTool):
    # Vom Dispatch generisch ausgewertet: jeder hier genannte Parameter wird
    # vor der Ausfuehrung durch sandbox.authorize_fs("read", …) geprueft.
    # Als ATTRIBUT und nicht als Liste im Dispatch, damit ein kuenftiges
    # Tabellen-Werkzeug automatisch mit abgesichert ist (dieselbe Lehre wie
    # bei den MCP-Gates: eine Whitelist erwischt neue Quellen von selbst).
    pfad_parameter = ("path",)
    ergebnis_max = 16000

    @property
    def name(self) -> str:
        return "xlsx_inspect"

    @property
    def description(self) -> str:
        return (
            "ERSTER SCHRITT bei jeder Tabellen-Aufgabe: zeigt den AUFBAU einer "
            "vorhandenen .xlsx ODER .csv/.tsv – Blaetter, Zeilen- und "
            "Spaltenzahl, Kopfzeile, "
            "Datentyp je Spalte und einige Beispielzeilen. Die Ausgabe ist "
            "klein und unabhaengig von der Dateigroesse. Benutze das statt "
            "office_read, sobald es um Tabellendaten geht: office_read macht "
            "aus einer grossen Mappe Text und kuerzt ihn auf einen Bruchteil."
        )

    def parameters_schema(self) -> dict:
        return {
            "type": "OBJECT",
            "properties": {
                "path": {"type": "STRING", "description": "Dateiname, /api/documents/-URL oder Pfad. .xlsx/.xlsm oder .csv/.tsv (Texttabelle, nur lesend)."},
                "kopfzeile": {"type": "INTEGER", "description": "Zeilennummer der Kopfzeile. Weglassen = je Blatt automatisch erkennen (empfohlen)."},
            },
            "required": ["path"],
        }

    async def execute(self, path: str = "", kopfzeile: int = 0, **kwargs) -> str:
        unbekannt = _unbekannte(kwargs)
        if unbekannt:
            return unbekannt
        try:
            wb, p = _oeffnen(path)
        except TabellenFehler as e:
            return f"Fehler: {e}"

        groesse = p.stat().st_size
        # Die Blatt-Beschreibungen werden ZUERST gesammelt und erst danach
        # zusammengesetzt – Warnung und Wegweiser gehoeren an den ANFANG.
        # An der echten Datei gemessen: 13 Blaetter ergeben 14.134 Zeichen, der
        # Deckel liegt bei 14.000 – die am Ende angehaengte Warnung wurde also
        # abgeschnitten. Genau derselbe Fehler wie bei office_read: ein Hinweis
        # am Textende ueberlebt die Kuerzung nicht, die er erklaeren soll.
        zeilen: list[str] = []
        abweichend: list[str] = []
        for ws in wb.worksheets:
            n_z = ws.max_row or 0
            n_s = ws.max_column or 0
            # JE BLATT erkennen, nicht einmal fuer die Mappe: die 13 Blaetter der
            # echten Datei haben ihre Kopfzeile teils in Z1, teils in Z3.
            kz, ab, auto = _kz(ws, kopfzeile)
            if auto and (kz != 1 or ab != 2):
                abweichend.append(f"'{ws.title}' -> Kopfzeile {kz}, Daten ab {ab}")
            zeilen.append("")
            zeilen.append(f"# Blatt '{ws.title}'  {n_z} Zeilen x {n_s} Spalten")

            kopf = _kopfzeile(ws, kz)
            benannt = [(i, k) for i, k in enumerate(kopf, start=1) if k]
            if benannt:
                txt, rest = _kopf_text(benannt)
                zeilen.append(f"  Kopfzeile {_kz_text(kz, auto, ab)}: {txt}"
                              + (f"  … (+{rest} weitere benannte Spalten, "
                                 f"{len(benannt)} von {n_s} insgesamt)" if rest else ""))
            else:
                zeilen.append(f"  Kopfzeile {_kz_text(kz, auto, ab)}: "
                              f"keine Beschriftungen gefunden")

            # Beispielzeilen direkt unter der Kopfzeile.
            gezeigt_z = 0
            for row in ws.iter_rows(min_row=ab, max_row=ab + BEISPIEL_ZEILEN - 1,
                                    max_col=BEISPIEL_SPALTEN, values_only=True):
                gezeigt_z += 1
                werte = " | ".join(_kurz(c, 30) for c in row)
                zeilen.append(f"  Zeile {ab + gezeigt_z - 1}: {werte}"
                              + (f"  … (+{n_s - BEISPIEL_SPALTEN} Spalten)"
                                 if n_s > BEISPIEL_SPALTEN else ""))
            if not gezeigt_z:
                zeilen.append("  (keine Datenzeilen unter der Kopfzeile)")

        n_blaetter = len(wb.worksheets)   # VOR close() – danach nicht mehr lesbar
        wb.close()

        # Formeln/Layout: ein ZWEITER Durchlauf waere teuer, deshalb nur die
        # Frage "gibt es sie ueberhaupt" – und die entscheidet, ob der Neubau
        # per office_create_excel ueberhaupt in Frage kommt.
        hinweis = self._formel_hinweis(p)

        kopf_block = [f"Datei: {p.name.split('__', 1)[-1]}  "
                      f"({groesse // 1024} KB, {n_blaetter} Blatt/Blaetter)"]
        if abweichend:
            # AUSDRUECKLICH BENENNEN. Ein Blatt, dessen Kopfzeile nicht in Zeile 1
            # steht, ist der haeufigste Grund fuer "kein Schluessel hat getroffen" –
            # und die Erkennung arbeitet zwar automatisch, kann aber danebenliegen.
            # Wer die Zahlen sieht, kann sie in den anderen Werkzeugen ueberstimmen.
            kopf_block.append(
                "ACHTUNG – nicht jedes Blatt ist gleich gebaut: "
                + "; ".join(abweichend[:12])
                + (f" (+{len(abweichend) - 12} weitere)" if len(abweichend) > 12 else "")
                + ". Die anderen Werkzeuge erkennen das ebenfalls selbst; mit "
                  "'kopfzeile' kannst du es ueberstimmen, wenn die Erkennung "
                  "danebenliegt."
            )
        kopf_block.append(
            "NAECHSTER SCHRITT: einzelne Bereiche mit xlsx_read_range ansehen, "
            "Daten mit xlsx_merge zusammenfuehren oder Zellen mit xlsx_edit "
            "schreiben. Baue die Tabelle NICHT mit office_create_excel neu auf – "
            "dabei gehen Formeln und Layout verloren."
        )
        if hinweis:
            kopf_block.append(hinweis)
        return _deckeln("\n".join(kopf_block + zeilen))

    def _formel_hinweis(self, p: Path) -> str:
        """Zaehlt Formeln stichprobenartig ueber den XML-Rohtext.

        Bewusst ueber das ZIP und nicht ueber openpyxl: ein zweites vollstaendiges
        Laden der Mappe kostet bei 360.000 Zellen mehrere Sekunden, waehrend das
        blosse Zaehlen von '<f>'-Elementen im Blatt-XML in Millisekunden fertig
        ist. Die Zahl ist eine Groessenordnung, keine Bilanz – und genau so wird
        sie auch formuliert.
        """
        try:
            import zipfile  # noqa: PLC0415
            n = 0
            with zipfile.ZipFile(str(p)) as z:
                for eintrag in z.namelist():
                    if eintrag.startswith("xl/worksheets/sheet") and eintrag.endswith(".xml"):
                        n += z.read(eintrag).count(b"<f>") + z.read(eintrag).count(b"<f ")
            if n:
                return (f"Diese Mappe enthaelt rund {n} Formel(n). Sie bleiben nur "
                        f"erhalten, wenn du die Datei mit xlsx_edit/xlsx_merge "
                        f"BEARBEITEST. office_create_excel wuerde sie durch feste "
                        f"Werte ersetzen.")
        except Exception:  # noqa: BLE001
            pass
        return ""


def _unbekannte(kwargs: dict) -> str:
    """LAUTER Fehlschlag bei unbekannten Parametern – der Kern des Vorfalls.

    Am 2026-08-19 rief das Modell ``office_create_excel`` mit dem Parameter
    ``" sheets"`` auf – mit FUEHRENDEM LEERZEICHEN. Der landete in ``**kwargs``,
    wurde wortlos verworfen, es wurde nichts geschrieben, und das Werkzeug
    meldete "✅ erstellt". Der Benutzer bekam eine leere Datei als Ergebnis
    angeboten, und das Modell hatte keine Moeglichkeit, den Fehler zu bemerken.

    Deshalb: ein unbekannter Parameter ist ein FEHLER, und die Meldung nennt
    ihn beim Namen, damit die Korrektur im naechsten Schritt moeglich ist.
    """
    uebrig = [k for k in kwargs if not k.startswith("_")]
    if not uebrig:
        return ""
    return ("Fehler: unbekannte Parameter " + ", ".join(repr(k) for k in uebrig)
            + ". Achte auf die genaue Schreibweise (auch auf fuehrende "
              "Leerzeichen im Parameternamen) und benutze nur die im Schema "
              "genannten Felder.")


# ═══════════════════════════════════════════════════════════════════════════
# xlsx_read_range – gezielt lesen, begrenzt, mit bezifferter Kuerzung
# ═══════════════════════════════════════════════════════════════════════════

class ReadRangeTool(BaseTool):
    pfad_parameter = ("path",)
    ergebnis_max = 16000

    @property
    def name(self) -> str:
        return "xlsx_read_range"

    @property
    def description(self) -> str:
        return (
            "Liest einen BEGRENZTEN Ausschnitt einer .xlsx oder .csv/.tsv "
            "(Blatt, Zeilen-, "
            "Spaltenbereich) als Text. Sagt immer dazu, wie viele Zeilen und "
            "Spalten es insgesamt gibt und wie viele gezeigt werden. Fuer den "
            "Ueberblick zuerst xlsx_inspect benutzen."
        )

    def parameters_schema(self) -> dict:
        return {
            "type": "OBJECT",
            "properties": {
                "path": {"type": "STRING", "description": "Dateiname, /api/documents/-URL oder Pfad. .xlsx/.xlsm oder .csv/.tsv (Texttabelle, nur lesend)."},
                "blatt": {"type": "STRING", "description": "Blattname (Standard: erstes Blatt)."},
                "ab_zeile": {"type": "INTEGER", "description": "Erste zu lesende Zeile (1-basiert, Standard 1)."},
                "zeilen": {"type": "INTEGER", "description": f"Anzahl Zeilen (Standard {LESE_ZEILEN_VORGABE}, hoechstens {LESE_ZEILEN_MAX})."},
                "spalten": {"type": "ARRAY", "items": {"type": "STRING"}, "description": "Nur diese Spalten (Kopfzeilen-Name, Buchstabe oder Nummer). Leer = alle."},
                "kopfzeile": {"type": "INTEGER", "description": "Zeilennummer der Kopfzeile. Weglassen = automatisch erkennen (empfohlen)."},
            },
            "required": ["path"],
        }

    async def execute(self, path: str = "", blatt: str = "", ab_zeile: int = 1,
                      zeilen: int = 0, spalten=None, kopfzeile: int = 0,
                      **kwargs) -> str:
        unbekannt = _unbekannte(kwargs)
        if unbekannt:
            return unbekannt
        try:
            wb, _p = _oeffnen(path)
            ws = _blatt(wb, blatt)
        except TabellenFehler as e:
            return f"Fehler: {e}"

        try:
            start = max(1, int(ab_zeile or 1))
        except Exception:  # noqa: BLE001
            start = 1
        try:
            n = int(zeilen or LESE_ZEILEN_VORGABE)
        except Exception:  # noqa: BLE001
            n = LESE_ZEILEN_VORGABE
        gewuenscht = max(1, n)
        n = min(gewuenscht, LESE_ZEILEN_MAX)

        gesamt_z = ws.max_row or 0
        gesamt_s = ws.max_column or 0

        # Spaltenauswahl gegen die Kopfzeile aufloesen – die stand bis zum
        # Feinschliff fest auf Zeile 1 und traf damit in Blatt '2015' der echten
        # Datei eine Liste von Nummerncodes statt der Spaltennamen.
        kz, _ab, kz_auto = _kz(ws, kopfzeile)
        kopf = _kopfzeile(ws, kz)
        auswahl: list[int] | None = None
        if spalten:
            if not isinstance(spalten, (list, tuple)):
                wb.close()
                return ("Fehler: 'spalten' muss eine Liste sein, z.B. "
                        "[\"Monat\", \"Umsatz\"] oder [\"A\", \"C\"].")
            auswahl, fehlend = [], []
            for b in spalten:
                i = _spalte_aufloesen(kopf, b)
                (auswahl if i else fehlend).append(i if i else b)
            if fehlend:
                wb.close()
                return "Fehler: " + _spalten_bericht(kopf, fehlend)

        # Die SPALTENNAMEN vollstaendig (bis KOPF_TEXT_MAX) – nicht auf
        # BEISPIEL_SPALTEN gekappt. Sonst verlangt der Hinweis unten ("Mit
        # 'spalten' gezielt auswaehlen") Namen, die dieselbe Ausgabe gerade
        # verschwiegen hat. Die DATENZEILEN bleiben gekappt: dort ist die
        # Begrenzung richtig, denn Werte gehoeren nicht durch das Modell.
        _kopf_paare = [(i, (kopf[i - 1] if i <= len(kopf) else "") or "")
                       for i in (auswahl or range(1, gesamt_s + 1))]
        _kopf_str, _kopf_rest = _kopf_text(_kopf_paare)
        kopf_txt = ("Spalten (Kopfzeile " + _kz_text(kz, kz_auto, _ab) + "): "
                    + _kopf_str
                    + (f"  … (+{_kopf_rest} weitere Spalten nicht benannt)"
                       if _kopf_rest else ""))

        aus = [f"Blatt '{ws.title}' – {gesamt_z} Zeilen x {gesamt_s} Spalten insgesamt.",
               kopf_txt, ""]

        max_s = gesamt_s if auswahl else min(gesamt_s, BEISPIEL_SPALTEN)
        gelesen = 0
        for i, row in enumerate(ws.iter_rows(min_row=start, max_row=start + n - 1,
                                             max_col=max_s, values_only=True),
                                start=start):
            if auswahl:
                werte = [row[k - 1] if k - 1 < len(row) else None for k in auswahl]
            else:
                werte = list(row)
            aus.append(f"{i}: " + " | ".join(_kurz(c, 40) for c in werte))
            gelesen += 1
        wb.close()

        # Die Bilanz ist Pflicht, nicht Zierrat: ohne sie haelt das Modell den
        # Ausschnitt fuer die ganze Tabelle (der Vorfall in einem Satz).
        ende = start + gelesen - 1
        aus.append("")
        aus.append(f"Gezeigt: Zeile {start} bis {ende} von {gesamt_z} "
                   f"({gelesen} Zeilen)."
                   + (f" NICHT gezeigt: {gesamt_z - ende} weitere Zeilen."
                      if gesamt_z > ende else "")
                   + (f" Angefordert waren {gewuenscht} Zeilen, das Werkzeug "
                      f"liefert hoechstens {LESE_ZEILEN_MAX}."
                      if gewuenscht > LESE_ZEILEN_MAX else ""))
        if not auswahl and gesamt_s > max_s:
            aus.append(f"NICHT gezeigt: {gesamt_s - max_s} weitere Spalten. "
                       f"Mit 'spalten' gezielt auswaehlen.")
        aus.append("Wenn du diese Daten weiterverarbeiten willst, tippe sie NICHT "
                   "ab – benutze xlsx_merge oder xlsx_edit.")
        return _deckeln("\n".join(aus))


# ═══════════════════════════════════════════════════════════════════════════
# xlsx_merge – Slave-Daten in den Master schreiben, Layout bleibt
# ═══════════════════════════════════════════════════════════════════════════

class MergeTool(BaseTool):
    pfad_parameter = ("master", "slave")
    ergebnis_max = 16000

    @property
    def name(self) -> str:
        return "xlsx_merge"

    @property
    def description(self) -> str:
        return (
            "Traegt Daten aus einer zweiten Tabelle (slave, .xlsx ODER .csv) in "
            "eine bestehende .xlsx (master) ein und liefert das Ergebnis als "
            "neue Datei. Der "
            "Master wird GEOEFFNET und beschrieben – Formeln, Spaltenbreiten, "
            "verbundene Zellen und Formate bleiben erhalten. Zugeordnet wird "
            "ueber 'schluessel' (gemeinsame Spalten). Die Daten laufen nicht "
            "durch das Modell: du benennst nur Spalten, nicht Werte."
        )

    def parameters_schema(self) -> dict:
        return {
            "type": "OBJECT",
            "properties": {
                "master": {"type": "STRING", "description": "Master-Tabelle, muss .xlsx sein (sie wird beschrieben und gibt Layout und Zielspalten vor)."},
                "slave": {"type": "STRING", "description": "Tabelle mit den einzutragenden Daten – .xlsx ODER .csv/.tsv."},
                "ziel": {"type": "STRING", "description": "Dateiname des Ergebnisses (ohne Pfad), z.B. 'Master_erweitert'."},
                "master_blatt": {"type": "STRING", "description": "Blatt im Master (Standard: erstes)."},
                "slave_blatt": {"type": "STRING", "description": "Blatt im Slave (Standard: erstes)."},
                "schluessel": {"type": "ARRAY", "items": {"type": "STRING"}, "description": "Spalten, ueber die zugeordnet wird (muessen in BEIDEN vorkommen), z.B. [\"Jahr\", \"Monat\"]."},
                "spalten": {"type": "ARRAY", "items": {"type": "STRING"}, "description": "Welche Slave-Spalten uebernommen werden. Leer = alle gleichnamigen ausser den Schluesseln."},
                "modus": {"type": "STRING", "description": "'aktualisieren' (nur vorhandene Master-Zeilen, Standard), 'anfuegen' (nur neue Zeilen) oder 'beides'."},
                "kopfzeile": {"type": "INTEGER", "description": "Zeilennummer der Kopfzeile in BEIDEN Tabellen. Weglassen = je Tabelle automatisch erkennen (empfohlen)."},
                "leere_uebernehmen": {"type": "BOOLEAN", "description": "Leere Slave-Werte in den Master schreiben (Standard false = vorhandenen Master-Wert stehen lassen)."},
            },
            "required": ["master", "slave", "ziel", "schluessel"],
        }

    async def execute(self, master: str = "", slave: str = "", ziel: str = "",
                      master_blatt: str = "", slave_blatt: str = "",
                      schluessel=None, spalten=None, modus: str = "aktualisieren",
                      kopfzeile: int = 0, leere_uebernehmen: bool = False,
                      **kwargs) -> str:
        unbekannt = _unbekannte(kwargs)
        if unbekannt:
            return unbekannt
        if not ziel:
            return "Fehler: 'ziel' ist Pflicht (Dateiname des Ergebnisses)."
        if not schluessel or not isinstance(schluessel, (list, tuple)):
            return ("Fehler: 'schluessel' ist Pflicht und muss eine Liste sein – "
                    "die Spalte(n), ueber die Master- und Slave-Zeilen einander "
                    "zugeordnet werden, z.B. [\"Jahr\", \"Monat\"]. Wenn du nicht "
                    "weisst, welche Spalten das sind, sieh dir beide Dateien "
                    "zuerst mit xlsx_inspect an.")

        art = str(modus or "aktualisieren").strip().lower()
        if art not in ("aktualisieren", "anfuegen", "beides"):
            return (f"Fehler: 'modus' war {modus!r}. Erlaubt sind "
                    f"'aktualisieren', 'anfuegen' oder 'beides'.")
        try:
            # Master SCHREIBEND (kein read_only) – nur so bleibt das Layout.
            wb_m, p_m = _oeffnen(master, schreibend=True)
            ws_m = _blatt(wb_m, master_blatt)
            wb_s, _p_s = _oeffnen(slave)
            ws_s = _blatt(wb_s, slave_blatt)
        except TabellenFehler as e:
            return f"Fehler: {e}"

        # JE SEITE erkennen. Master und Slave sind haeufig verschieden gebaut –
        # eine gemeinsame Zeilennummer waere fuer eine der beiden falsch. Eine
        # ausdrueckliche Angabe gilt weiterhin fuer beide.
        kz_m, ab_m, auto_m = _kz(ws_m, kopfzeile)
        kz_s, ab_s, auto_s = _kz(ws_s, kopfzeile)
        kopf_m = _kopfzeile(ws_m, kz_m)
        kopf_s = _kopfzeile(ws_s, kz_s)

        # ── Schluessel in beiden Tabellen aufloesen ────────────────────────
        k_m, k_s, fehlt_m, fehlt_s = [], [], [], []
        for b in schluessel:
            im = _spalte_aufloesen(kopf_m, b)
            i_s = _spalte_aufloesen(kopf_s, b)
            (k_m.append(im) if im else fehlt_m.append(b))
            (k_s.append(i_s) if i_s else fehlt_s.append(b))
        if fehlt_m or fehlt_s:
            wb_m.close(); wb_s.close()
            teile = []
            if fehlt_m:
                teile.append("Im MASTER: " + _spalten_bericht(kopf_m, fehlt_m))
            if fehlt_s:
                teile.append("Im SLAVE: " + _spalten_bericht(kopf_s, fehlt_s))
            return "Fehler: Schluesselspalten nicht gefunden.\n" + "\n".join(teile)

        # ── Welche Spalten uebernommen werden ─────────────────────────────
        schluessel_namen = {str(b).strip().lower() for b in schluessel}
        paare: list[tuple[int, int, str]] = []   # (slave_idx, master_idx, name)
        if spalten:
            if not isinstance(spalten, (list, tuple)):
                wb_m.close(); wb_s.close()
                return "Fehler: 'spalten' muss eine Liste von Spaltennamen sein."
            fehlend = []
            for b in spalten:
                i_s = _spalte_aufloesen(kopf_s, b)
                im = _spalte_aufloesen(kopf_m, b)
                if not i_s:
                    fehlend.append(f"{b} (im Slave)")
                elif not im:
                    fehlend.append(f"{b} (im Master)")
                else:
                    paare.append((i_s, im, str(b)))
            if fehlend:
                wb_m.close(); wb_s.close()
                return ("Fehler: " + _spalten_bericht(kopf_s, fehlend)
                        + "\nHinweis: eine Spalte muss in BEIDEN Tabellen "
                          "existieren – der Master gibt die Zielspalte vor.")
        else:
            # Vorgabe: alle gleichnamigen Spalten ausser den Schluesseln.
            for i_s, name in enumerate(kopf_s, start=1):
                if not name or name.strip().lower() in schluessel_namen:
                    continue
                im = _spalte_aufloesen(kopf_m, name)
                if im:
                    paare.append((i_s, im, name))
        if not paare:
            wb_m.close(); wb_s.close()
            return ("Fehler: keine zu uebernehmende Spalte gefunden. Master und "
                    "Slave haben ausser den Schluesseln keine gleichnamige "
                    "Spalte.\nMaster-Kopfzeile: "
                    + ", ".join(repr(k) for k in kopf_m if k)[:800]
                    + "\nSlave-Kopfzeile: "
                    + ", ".join(repr(k) for k in kopf_s if k)[:800]
                    + "\nGib 'spalten' ausdruecklich an, wenn die Namen abweichen.")

        # ── Slave nach Schluessel indizieren ──────────────────────────────
        def _norm(v) -> str:
            """Schluesselwert vergleichbar machen.

            '2004' aus einer Textspalte und 2004 aus einer Zahlenspalte sind
            derselbe Schluessel – ohne diese Normierung trifft in gemischt
            getippten Tabellen KEINE einzige Zeile, und der Lauf endet in
            'nichts zugeordnet', obwohl alles passt. Float mit ganzzahligem
            Wert wird bewusst wie die Ganzzahl behandelt (2004.0 == 2004).
            """
            if v is None:
                return ""
            if isinstance(v, float) and v.is_integer():
                return str(int(v))
            return str(v).strip().lower()

        index: dict[tuple, list] = {}
        doppelte = 0
        slave_zeilen = 0
        for row in ws_s.iter_rows(min_row=ab_s, values_only=True):
            if all(c is None for c in row):
                continue
            slave_zeilen += 1
            key = tuple(_norm(row[i - 1] if i - 1 < len(row) else None) for i in k_s)
            if all(t == "" for t in key):
                continue
            if key in index:
                doppelte += 1
            else:
                index[key] = row

        if not index:
            wb_m.close(); wb_s.close()
            return (f"Fehler: der Slave enthaelt keine auswertbaren Zeilen unter "
                    f"der Kopfzeile ({_kz_text(kz_s, auto_s, ab_s)}). Stimmt "
                    f"'kopfzeile' und "
                    f"'slave_blatt'?")

        # ── Master durchgehen und schreiben ───────────────────────────────
        getroffen = 0
        ohne_treffer = 0
        geschriebene_zellen = 0
        master_zeilen = 0
        benutzte_keys: set = set()
        beispiel_master: list[str] = []

        for r in range(ab_m, (ws_m.max_row or ab_m) + 1):
            werte = [ws_m.cell(row=r, column=i).value for i in k_m]
            if all(v is None for v in werte):
                continue
            master_zeilen += 1
            key = tuple(_norm(v) for v in werte)
            if len(beispiel_master) < 3:
                beispiel_master.append("|".join(key))
            treffer = index.get(key)
            if treffer is None:
                ohne_treffer += 1
                continue
            benutzte_keys.add(key)
            if art == "anfuegen":
                continue
            getroffen += 1
            for i_s, i_m, _n in paare:
                wert = treffer[i_s - 1] if i_s - 1 < len(treffer) else None
                if wert is None and not leere_uebernehmen:
                    continue
                ws_m.cell(row=r, column=i_m).value = wert
                geschriebene_zellen += 1

        # ── Nicht zugeordnete Slave-Zeilen anfuegen ───────────────────────
        angefuegt = 0
        if art in ("anfuegen", "beides"):
            ziel_zeile = (ws_m.max_row or ab_m) + 1
            for key, row in index.items():
                if key in benutzte_keys:
                    continue
                for i, i_m in enumerate(k_m):
                    ws_m.cell(row=ziel_zeile, column=i_m).value = (
                        row[k_s[i] - 1] if k_s[i] - 1 < len(row) else None)
                    geschriebene_zellen += 1
                for i_s, i_m, _n in paare:
                    wert = row[i_s - 1] if i_s - 1 < len(row) else None
                    if wert is None and not leere_uebernehmen:
                        continue
                    ws_m.cell(row=ziel_zeile, column=i_m).value = wert
                    geschriebene_zellen += 1
                ziel_zeile += 1
                angefuegt += 1

        # ── LAUTER Fehlschlag, wenn nichts passiert ist ───────────────────
        # Eine Datei zurueckzugeben, in der nichts steht, ist der Vorfall vom
        # 2026-08-19. Ohne einen einzigen Treffer stimmt fast immer die
        # Schluesselwahl nicht – und dann helfen BEISPIELWERTE aus beiden
        # Tabellen weiter als jede allgemeine Fehlermeldung.
        if getroffen == 0 and angefuegt == 0:
            beispiel_slave = ["|".join(k) for k in list(index)[:3]]
            wb_m.close(); wb_s.close()
            return (
                f"Fehler: kein einziger Schluessel hat getroffen – es wurde "
                f"NICHTS geschrieben und KEINE Datei erzeugt.\n"
                f"Master: {master_zeilen} Datenzeilen, Slave: {slave_zeilen}.\n"
                f"Schluesselwerte im Master (Beispiele): "
                f"{', '.join(beispiel_master) or '(keine)'}\n"
                f"Schluesselwerte im Slave  (Beispiele): "
                f"{', '.join(beispiel_slave) or '(keine)'}\n"
                f"Gelesene Kopfzeilen: Master {_kz_text(kz_m, auto_m, ab_m)}, "
                f"Slave {_kz_text(kz_s, auto_s, ab_s)}.\n"
                f"Pruefe mit xlsx_inspect, ob 'schluessel' in beiden Tabellen "
                f"dieselbe Bedeutung hat. Stimmt eine der Kopfzeilen nicht, gib "
                f"'kopfzeile' ausdruecklich mit."
            )

        verluste = _verluste(wb_m)

        # ── Speichern ─────────────────────────────────────────────────────
        from skills.office.main import _new_path, _ok  # noqa: PLC0415
        disk, fname, dl = _new_path(ziel, "xlsx")
        try:
            wb_m.save(str(disk))
        except Exception as e:  # noqa: BLE001
            wb_m.close(); wb_s.close()
            return f"Fehler beim Speichern: {e}"
        wb_m.close(); wb_s.close()

        # DIE BENUTZTEN KOPFZEILEN GEHOEREN IN DEN BERICHT. Die Erkennung
        # arbeitet automatisch – eine Automatik, deren Ergebnis niemand sieht,
        # ist nicht ueberpruefbar. Steht die Zahl da, faellt ein Fehlgriff sofort
        # auf und laesst sich mit 'kopfzeile' ueberstimmen.
        bericht = [
            f"Master '{p_m.name.split('__', 1)[-1]}' Blatt '{ws_m.title}': "
            f"{master_zeilen} Datenzeilen, Kopfzeile {_kz_text(kz_m, auto_m, ab_m)}.",
            f"Slave Blatt '{ws_s.title}': {slave_zeilen} Datenzeilen, "
            f"{len(index)} verschiedene Schluessel, "
            f"Kopfzeile {_kz_text(kz_s, auto_s, ab_s)}.",
            f"Uebernommene Spalten ({len(paare)}): "
            + ", ".join(n for _a, _b, n in paare[:25])
            + (f" … (+{len(paare) - 25})" if len(paare) > 25 else ""),
            f"Aktualisierte Master-Zeilen: {getroffen}",
            f"Angefuegte Zeilen: {angefuegt}",
            f"Master-Zeilen ohne Treffer im Slave: {ohne_treffer}",
            f"Geschriebene Zellen: {geschriebene_zellen}",
        ]
        if doppelte:
            bericht.append(
                f"⚠ {doppelte} Slave-Zeile(n) hatten einen bereits vergebenen "
                f"Schluessel – verwendet wurde jeweils die ERSTE. Wenn das "
                f"nicht gewollt ist, ist der Schluessel nicht eindeutig.")
        nicht_verwendet = len(index) - len(benutzte_keys)
        if nicht_verwendet and art == "aktualisieren":
            bericht.append(
                f"⚠ {nicht_verwendet} Slave-Schluessel kamen im Master nicht vor "
                f"und wurden NICHT uebernommen (modus='aktualisieren'). Mit "
                f"modus='beides' werden sie als neue Zeilen angefuegt.")

        return _ok(dl, fname, disk,
                   extra="\n".join(bericht) + _verlust_hinweis(verluste)
                   + _weiterarbeiten_hinweis(fname))


# ═══════════════════════════════════════════════════════════════════════════
# xlsx_edit – einzelne Zellen/Spalten schreiben, Layout bleibt
# ═══════════════════════════════════════════════════════════════════════════

_ZELLE_RE = re.compile(r"^([A-Za-z]{1,3})(\d+)$")


class EditTool(BaseTool):
    pfad_parameter = ("path",)
    ergebnis_max = 16000

    @property
    def name(self) -> str:
        return "xlsx_edit"

    @property
    def description(self) -> str:
        return (
            "Schreibt einzelne Zellen in eine BESTEHENDE .xlsx und liefert das "
            "Ergebnis als neue Datei. Die Originaldatei wird geoeffnet und "
            "beschrieben, Formeln und Layout bleiben erhalten. Fuer viele "
            "Datenzeilen aus einer zweiten Tabelle ist xlsx_merge der richtige "
            "Weg – hier werden die Werte einzeln angegeben."
        )

    def parameters_schema(self) -> dict:
        return {
            "type": "OBJECT",
            "properties": {
                "path": {"type": "STRING", "description": "Zu bearbeitende .xlsx (eine CSV kann nicht bearbeitet werden – sie hat kein Layout)."},
                "ziel": {"type": "STRING", "description": "Dateiname des Ergebnisses (ohne Pfad)."},
                "blatt": {"type": "STRING", "description": "Blattname (Standard: erstes Blatt)."},
                "aenderungen": {
                    "type": "ARRAY",
                    "items": {"type": "OBJECT"},
                    "description": "Liste von {\"zelle\": \"B5\", \"wert\": ...} – 'wert' darf auch eine Formel sein (\"=SUMME(B2:B4)\").",
                },
            },
            "required": ["path", "ziel", "aenderungen"],
        }

    async def execute(self, path: str = "", ziel: str = "", blatt: str = "",
                      aenderungen=None, **kwargs) -> str:
        unbekannt = _unbekannte(kwargs)
        if unbekannt:
            return unbekannt
        if not ziel:
            return "Fehler: 'ziel' ist Pflicht (Dateiname des Ergebnisses)."
        if not aenderungen or not isinstance(aenderungen, (list, tuple)):
            return ("Fehler: 'aenderungen' ist Pflicht und muss eine nicht-leere "
                    "Liste sein, z.B. [{\"zelle\": \"B5\", \"wert\": 42}].")

        try:
            wb, _p = _oeffnen(path, schreibend=True)
            ws = _blatt(wb, blatt)
        except TabellenFehler as e:
            return f"Fehler: {e}"

        geschrieben = 0
        fehler: list[str] = []
        for nr, a in enumerate(aenderungen, start=1):
            if not isinstance(a, dict):
                fehler.append(f"#{nr}: kein Objekt ({type(a).__name__})")
                continue
            zelle = str(a.get("zelle") or "").strip()
            m = _ZELLE_RE.match(zelle)
            if not m:
                fehler.append(f"#{nr}: 'zelle' war {zelle!r} – erwartet wird "
                              f"z.B. 'B5'")
                continue
            spalte = _buchstabe_zu_index(m.group(1))
            zeile = int(m.group(2))
            if not spalte or zeile < 1:
                fehler.append(f"#{nr}: {zelle!r} ist keine gueltige Zelle")
                continue
            try:
                ws.cell(row=zeile, column=spalte).value = a.get("wert")
                geschrieben += 1
            except Exception as e:  # noqa: BLE001
                fehler.append(f"#{nr}: {zelle} nicht beschreibbar ({e})")

        # LAUT scheitern, wenn nichts geschrieben wurde – kein leeres Ergebnis.
        if not geschrieben:
            wb.close()
            return ("Fehler: keine einzige Zelle konnte geschrieben werden, es "
                    "wurde KEINE Datei erzeugt.\n" + "\n".join(fehler[:20]))

        verluste = _verluste(wb)
        from skills.office.main import _new_path, _ok  # noqa: PLC0415
        disk, fname, dl = _new_path(ziel, "xlsx")
        try:
            wb.save(str(disk))
        except Exception as e:  # noqa: BLE001
            wb.close()
            return f"Fehler beim Speichern: {e}"
        wb.close()

        extra = f"{geschrieben} Zelle(n) in Blatt '{ws.title}' geschrieben."
        if fehler:
            # Teilerfolg wird BENANNT. Ein "erstellt" ueber einer halb
            # ausgefuehrten Aenderungsliste ist dieselbe stille Luege wie eine
            # leere Datei mit Erfolgsmeldung.
            extra += (f"\n⚠ {len(fehler)} Aenderung(en) wurden NICHT ausgefuehrt:\n"
                      + "\n".join(fehler[:20]))
        return _ok(dl, fname, disk,
                   extra=extra + _verlust_hinweis(verluste)
                   + _weiterarbeiten_hinweis(fname))


def get_tabellen_tools():
    return [InspectTool(), ReadRangeTool(), MergeTool(), EditTool()]
