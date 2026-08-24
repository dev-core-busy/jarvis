#!/usr/bin/env python3
"""Waechter fuer skills/office/xlsx_patch.py – Zellen im ZIP patchen.

LAEUFT OHNE openpyxl und ohne fastapi; ``backend.config`` wird NICHT importiert.
Das Modul selbst braucht nur ``zipfile`` und ``re``, deshalb laufen alle
Struktur-Pruefungen ueberall. Wo openpyxl vorhanden ist (venv auf DEV), kommen
zusaetzliche Pruefungen dazu, die die Datei mit einem ECHTEN Konsumenten lesen.

Die Pruefungen gegen die echte Produktionsdatei stehen NICHT hier – sie enthaelt
Kundendaten und gehoert nicht ins oeffentliche Repo. Das synthetische Material
hier bildet die Struktur nach, die dort gemessen wurde: mehrere Blaetter mit
eigenen Beziehungen, shared formulas, printerSettings, calcChain.
"""

import re
import shutil
import sys
import tempfile
import zipfile
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from skills.office import xlsx_patch as xp                  # noqa: E402

TMP = Path(tempfile.mkdtemp(prefix="xlsx_patch_"))

# ── SANDKASTEN-WAECHTER ─────────────────────────────────────────────────────
# Exit 2, wenn ein Pfad aus dem Wegwerf-Verzeichnis herauszeigt: "konnte nicht
# laufen" muss von "bestanden" unterscheidbar sein (Register).
if not str(TMP).startswith(tempfile.gettempdir()):
    print("ABBRUCH: Sandkasten liegt nicht in %s" % tempfile.gettempdir())
    sys.exit(2)

_ok = _fail = 0


def pruefe(bed, text):
    global _ok, _fail
    if bed:
        _ok += 1
        print("  OK   %s" % text)
    else:
        _fail += 1
        print("  FAIL %s" % text)


def wohlgeformt(datei: Path) -> list:
    """Namen der Teile, die KEIN gueltiges XML sind."""
    import xml.dom.minidom as md
    kaputt = []
    with zipfile.ZipFile(datei) as z:
        for i in z.infolist():
            if i.filename.endswith((".xml", ".rels")):
                try:
                    md.parseString(z.read(i.filename))
                except Exception:  # noqa: BLE001
                    kaputt.append(i.filename)
    return kaputt


# ── Material: eine Mappe, die NICHT von openpyxl stammt ─────────────────────
# Nachgebaut nach der echten Datei: zwei Blaetter, Beziehungen je Blatt,
# printerSettings, calcChain, shared formulas, sharedStrings.
_WB = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
       '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
       ' xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
       '<sheets><sheet name="Daten" sheetId="1" r:id="rId1"/>'
       '<sheet name="A &amp; B" sheetId="2" r:id="rId2"/></sheets>'
       '<calcPr calcId="191029"/></workbook>')
_RELS = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
         '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
         '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet7.xml"/>'
         '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet3.xml"/>'
         '<Relationship Id="rId5" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/calcChain" Target="calcChain.xml"/>'
         '<Relationship Id="rId6" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" Target="sharedStrings.xml"/>'
         '<Relationship Id="rId7" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>'
         '</Relationships>')
_SML = "application/vnd.openxmlformats-officedocument.spreadsheetml"
# VOLLSTAENDIGE Content-Types und eine Wurzel-Beziehung auf xl/workbook.xml:
# ohne beides ist die Mappe kein gueltiges xlsx, und openpyxl meldet "File
# contains no valid workbook part" – die erste Fassung dieses Tests hatte das
# und Abschnitt 9 konnte damit gar nichts beweisen.
_CT = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
       '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
       '<Default Extension="bin" ContentType="%s.printerSettings"/>'
       '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
       '<Default Extension="xml" ContentType="application/xml"/>'
       '<Override PartName="/xl/workbook.xml" ContentType="%s.sheet.main+xml"/>'
       '<Override PartName="/xl/worksheets/sheet7.xml" ContentType="%s.worksheet+xml"/>'
       '<Override PartName="/xl/worksheets/sheet3.xml" ContentType="%s.worksheet+xml"/>'
       '<Override PartName="/xl/styles.xml" ContentType="%s.styles+xml"/><Override PartName="/xl/sharedStrings.xml" ContentType="%s.sharedStrings+xml"/>'
       '<Override PartName="/xl/calcChain.xml" ContentType="%s.calcChain+xml"/>'
       '</Types>' % (_SML, _SML, _SML, _SML, _SML, _SML, _SML))
_ROOT_RELS = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
              '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
              '<Relationship Id="rId1" Target="xl/workbook.xml" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument"/>'
              '</Relationships>')
# Blatt 1: Zeilen 1,2,5 – Zelle B2 traegt die MASTER-Formel einer Gruppe,
# C2 ist ihr Folger. A5 hat einen Stil, den der Patch behalten muss.
_SHEET1 = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
           '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
           ' xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006"'
           ' mc:Ignorable="x14ac" xmlns:x14ac="http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac">'
           '<dimension ref="A1:D5"/><sheetData>'
           '<row r="1" spans="1:4"><c r="A1" s="1" t="s"><v>0</v></c>'
           '<c r="C1" s="2"><v>7</v></c></row>'
           '<row r="2"><c r="B2" s="1"><f t="shared" ref="B2:D2" si="0">SUM(A1:A1)</f><v>1</v></c>'
           '<c r="C2" s="1"><f t="shared" si="0"/><v>2</v></c></row>'
           '<row r="5"><c r="A5" s="2"><v>5</v></c></row>'
           '</sheetData><pageMargins left="0.7"/>'
           '<pageSetup paperSize="9" r:id="rId1"'
           ' xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"/>'
           '</worksheet>')
_SHEET2 = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
           '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
           '<dimension ref="A1:A1"/><sheetData/></worksheet>')
_SHARED = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
           '<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
           ' count="1" uniqueCount="1"><si><t>Kopf</t></si></sst>')
_CALC = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
         '<calcChain xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
         '<c r="B2" i="1"/></calcChain>')
# Ein Teil, den openpyxl nicht kennt – hier steht er fuer Diagramme/Pivots.
_FREMD = b"\x89PNG-nicht-wirklich-aber-binaer\x00\x01\x02"
# Drei cellXfs – die Blatt-XML benutzt s="1" und s="2". Ohne styles.xml wirft
# openpyxl beim Lesen IndexError; ein Testmaterial, das ein echter Konsument
# nicht oeffnet, kann nichts beweisen.
_STYLES = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
           '<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
           '<fonts count="1"><font><sz val="11"/><name val="Calibri"/></font></fonts>'
           '<fills count="1"><fill><patternFill patternType="none"/></fill></fills>'
           '<borders count="1"><border/></borders>'
           '<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>'
           '<cellXfs count="3">'
           '<xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>'
           '<xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>'
           '<xf numFmtId="4" fontId="0" fillId="0" borderId="0" xfId="0" applyNumberFormat="1"/>'
           '</cellXfs></styleSheet>')


def mappe(pfad: Path) -> Path:
    with zipfile.ZipFile(pfad, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("[Content_Types].xml", _CT)
        z.writestr("_rels/.rels", _ROOT_RELS)
        z.writestr("xl/workbook.xml", _WB)
        z.writestr("xl/_rels/workbook.xml.rels", _RELS)
        z.writestr("xl/worksheets/sheet7.xml", _SHEET1)
        z.writestr("xl/worksheets/sheet3.xml", _SHEET2)
        z.writestr("xl/worksheets/_rels/sheet7.xml.rels",
                   '<?xml version="1.0"?><Relationships '
                   'xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                   '<Relationship Id="rId1" Type="http://x/printerSettings" '
                   'Target="../printerSettings/printerSettings1.bin"/></Relationships>')
        z.writestr("xl/printerSettings/printerSettings1.bin", _FREMD)
        z.writestr("xl/charts/chart1.xml", '<?xml version="1.0"?><chartSpace/>')
        z.writestr("xl/styles.xml", _STYLES)
        z.writestr("xl/sharedStrings.xml", _SHARED)
        z.writestr("xl/calcChain.xml", _CALC)
    return pfad


QUELLE = mappe(TMP / "quelle.xlsx")


# ═══════════════════════════════════════════════════════════════════════════
print("1) Adressen – dieselbe Haerte wie in tabellen.py")
# ═══════════════════════════════════════════════════════════════════════════
for b, n in (("A", 1), ("Z", 26), ("AA", 27), ("AZ", 52), ("BA", 53),
             ("XFD", 16384), ("a", 1)):
    pruefe(xp.spalte_zu_index(b) == n, "Spalte %s -> %d" % (b, n))
for b, n in ((1, "A"), (26, "Z"), (27, "AA"), (16384, "XFD")):
    pruefe(xp.index_zu_spalte(b) == n, "Index %d -> %s" % (b, n))
# Genau die Faelle, die in tabellen.py Spalte 4.498.495.991.152 ergaben.
for b in ("Unbekannt", "ÄB", "ABCD", "XFE", "", "A1", " "):
    try:
        xp.spalte_zu_index(b)
        pruefe(False, "Spaltenangabe %r wird abgelehnt" % b)
    except xp.PatchFehler:
        pruefe(True, "Spaltenangabe %r wird abgelehnt" % b)
# "Ort" IST eine gueltige Spalte (10.628) – ablehnen waere falsch. Der Schutz
# ist deshalb ein HINWEIS im Bericht, kein Verbot (s. Abschnitt 4b).
pruefe(xp.spalte_zu_index("Ort") == 10628,
       "ein dreibuchstabiges Wort ist eine gueltige Spalte – das ist die Falle")
pruefe(xp.adresse_teilen("B12") == ("B", 12, 2), "B12 wird zerlegt")
pruefe(xp.adresse_teilen(" c3 ") == ("C", 3, 3), "Leerzeichen und Kleinschreibung")
for adr in ("B0", "12B", "B", "12", "B-1", "ÄB1", "B1048577"):
    try:
        xp.adresse_teilen(adr)
        pruefe(False, "Adresse %r wird abgelehnt" % adr)
    except xp.PatchFehler:
        pruefe(True, "Adresse %r wird abgelehnt" % adr)


# ═══════════════════════════════════════════════════════════════════════════
print("\n2) Zell-XML – Typen, Stil, Maskierung")
# ═══════════════════════════════════════════════════════════════════════════
pruefe(xp._zelle_xml("B2", 42, None) == '<c r="B2"><v>42</v></c>', "Ganzzahl")
pruefe('<v>1.5</v>' in xp._zelle_xml("B2", 1.5, None), "Kommazahl")
pruefe('t="b"><v>1</v>' in xp._zelle_xml("B2", True, None), "Wahrheitswert")
pruefe('t="inlineStr"><is><t>Hallo</t></is>' in xp._zelle_xml("B2", "Hallo", None),
       "Text als inlineStr (sharedStrings bleibt unangetastet)")
pruefe('xml:space="preserve"' in xp._zelle_xml("B2", " Rand ", None),
       "Text mit Randleerzeichen behaelt sie (xml:space)")
pruefe(xp._zelle_xml("B2", None, "7") == '<c r="B2" s="7"/>',
       "None leert die Zelle, der Stil bleibt")
pruefe(xp._zelle_xml("B2", "", None) == '<c r="B2"/>', "Leerstring leert die Zelle")
pruefe('<f>SUM(A1:A9)</f>' in xp._zelle_xml("B2", "=SUM(A1:A9)", None)
       and "<v>" not in xp._zelle_xml("B2", "=SUM(A1:A9)", None),
       "Formel ohne gecachten Wert (der waere eine Behauptung)")
pruefe(' s="42"' in xp._zelle_xml("B2", 1, "42"), "Stil wird uebernommen")
_x = xp._zelle_xml("B2", '<a href="x">&\'', None)
pruefe("&lt;" in _x and "&amp;" in _x and "&quot;" in _x and "&#39;" in _x
       and "<a" not in _x, "Text wird vollstaendig maskiert (auch \" und ')")
pruefe("\x07" not in xp._zelle_xml("B2", "a\x07b", None),
       "Steuerzeichen werden entfernt (sonst ist die XML ungueltig)")
try:
    xp._zelle_xml("B2", "=", None)
    pruefe(False, "leere Formel wird abgelehnt")
except xp.PatchFehler:
    pruefe(True, "leere Formel wird abgelehnt")


# ═══════════════════════════════════════════════════════════════════════════
print("\n3) Blatt finden – ueber workbook.xml, NICHT ueber den Dateinamen")
# ═══════════════════════════════════════════════════════════════════════════
k = xp.blaetter_von(QUELLE)
pruefe(list(k) == ["Daten", "A & B"], "Namen und Reihenfolge aus workbook.xml")
pruefe(k["Daten"] == "xl/worksheets/sheet7.xml",
       "Blatt 1 zeigt auf sheet7.xml – die Dateinummer sagt nichts")
pruefe(k["A & B"] == "xl/worksheets/sheet3.xml",
       "Blattname mit &-Entity wird korrekt gelesen")


# ═══════════════════════════════════════════════════════════════════════════
print("\n4) Zellen ersetzen, einsortieren, Zeilen anlegen")
# ═══════════════════════════════════════════════════════════════════════════
xml, b = xp.patch_blatt(_SHEET1, {"A1": 5})
pruefe('<c r="A1" s="1"><v>5</v></c>' in xml,
       "bestehende Zelle wird ersetzt und behaelt ihren Stil")
pruefe('t="s"' not in xml.split("<sheetData>")[1].split("</row>")[0],
       "das alte t=\"s\" ist weg (sonst zeigte Excel Muell)")

xml, b = xp.patch_blatt(_SHEET1, {"B1": 9})
_zeile1 = xml.split('<row r="1"')[1].split("</row>")[0]
_reihe = re.findall(r'<c r="([A-Z]+)1"', _zeile1)
pruefe(_reihe == ["A", "B", "C"],
       "neue Zelle wird in AUFSTEIGENDER Reihenfolge einsortiert (%s)" % _reihe)

xml, b = xp.patch_blatt(_SHEET1, {"D1": 1})
pruefe(re.findall(r'<c r="([A-Z]+)1"', xml.split('<row r="1"')[1].split("</row>")[0])
       == ["A", "C", "D"], "eine Zelle hinter der letzten wird angehaengt")

xml, b = xp.patch_blatt(_SHEET1, {"A3": 3})
_zeilen = re.findall(r'<row r="(\d+)"', xml)
pruefe(_zeilen == ["1", "2", "3", "5"],
       "fehlende Zeile wird an der richtigen Stelle angelegt (%s)" % _zeilen)
pruefe(b["neue_zeilen"] == [3], "und im Bericht genannt")

xml, b = xp.patch_blatt(_SHEET1, {"A9": 9})
pruefe(re.findall(r'<row r="(\d+)"', xml) == ["1", "2", "5", "9"],
       "eine Zeile hinter der letzten wird angehaengt")
pruefe('<dimension ref="A1:D9"/>' in xml,
       "dimension wird auf die neue Zeile erweitert")
xml, b = xp.patch_blatt(_SHEET1, {"F1": 1})
pruefe('<dimension ref="A1:F5"/>' in xml, "dimension wird auf die neue Spalte erweitert")
xml, b = xp.patch_blatt(_SHEET1, {"A1": 1})
pruefe('<dimension ref="A1:D5"/>' in xml, "dimension bleibt, wenn innerhalb geschrieben")

# Leeres <sheetData/> (Blatt 2) muss ein Endtag bekommen.
xml, b = xp.patch_blatt(_SHEET2, {"B2": 1})
pruefe("<sheetData><row r=\"2\">" in xml and "</sheetData>" in xml,
       "in ein leeres Blatt wird sheetData korrekt geoeffnet")

# Mehrere Zellen und Zeilen in EINEM Aufruf – Positionen duerfen nicht verrutschen.
xml, b = xp.patch_blatt(_SHEET1, {"A1": 1, "D1": 2, "A3": 3, "A5": 4, "B7": 5})
pruefe(len(b["geschrieben"]) == 5, "fuenf Zellen in einem Aufruf")
pruefe(re.findall(r'<row r="(\d+)"', xml) == ["1", "2", "3", "5", "7"],
       "Zeilenreihenfolge bleibt korrekt")
pruefe(re.findall(r'<c r="([A-Z]+)1"', xml.split('<row r="1"')[1].split("</row>")[0])
       == ["A", "C", "D"], "und die Spaltenreihenfolge in Zeile 1 auch")


print("\n4b) Weit ausserhalb der Tabelle geschrieben – benannt, nicht verboten")
# ═══════════════════════════════════════════════════════════════════════════
xml, b = xp.patch_blatt(_SHEET1, {"Ort1": 5})
pruefe(b["geschrieben"] == ["ORT1"],
       "die Zelle wird geschrieben (die Spalte gibt es) – Adresse normalisiert")
pruefe(b["weit_draussen"] == ["ORT1"],
       "aber als 'weit ausserhalb' vermerkt – sonst faellt der Vertipper nie auf")
xml, b = xp.patch_blatt(_SHEET1, {"E1": 5})
pruefe(b["weit_draussen"] == [], "eine Spalte direkt daneben ist kein Anlass")
xml, b = xp.patch_blatt(_SHEET1, {"A2000": 5})
pruefe(b["weit_draussen"] == ["A2000"], "auch eine weit entfernte ZEILE faellt auf")
xml, b = xp.patch_blatt(_SHEET1, {"A9": 5})
pruefe(b["weit_draussen"] == [], "ein paar Zeilen weiter nicht")
_qt = (ROOT / "skills" / "office" / "tabellen.py").read_text(encoding="utf-8")
pruefe("weit_draussen" in _qt, "xlsx_edit gibt den Hinweis auch aus")


# ═══════════════════════════════════════════════════════════════════════════
print("\n5) shared formulas – die MASTER-Zelle wird nicht ueberschrieben")
# ═══════════════════════════════════════════════════════════════════════════
xml, b = xp.patch_blatt(_SHEET1, {"B2": 99})
pruefe(b["abgelehnt"] == ["B2"] and b["geschrieben"] == [],
       "die Master-Zelle wird abgelehnt")
pruefe('<f t="shared" ref="B2:D2" si="0">SUM(A1:A1)</f>' in xml,
       "und ihre Formel steht unveraendert da")
xml, b = xp.patch_blatt(_SHEET1, {"C2": 99})
pruefe(b["geschrieben"] == ["C2"] and b["abgelehnt"] == [],
       "eine FOLGE-Zelle darf geschrieben werden")
pruefe('<c r="C2" s="1"><v>99</v></c>' in xml, "sie traegt danach den Wert")
pruefe('<f t="shared" ref="B2:D2" si="0">' in xml,
       "die Gruppendefinition bleibt trotzdem erhalten")


# ═══════════════════════════════════════════════════════════════════════════
print("\n6) patch_datei – alles andere bleibt BYTE-GLEICH")
# ═══════════════════════════════════════════════════════════════════════════
ziel = TMP / "ergebnis.xlsx"
b = xp.patch_datei(QUELLE, ziel, {"Daten": {"A1": "Neu", "A9": 9}})
pruefe(sorted(b["geschrieben"]) == ["Daten!A1", "Daten!A9"], "Bericht nennt die Zellen")
alt, neu = xp.teile(QUELLE), xp.teile(ziel)
pruefe(alt - neu == {"xl/calcChain.xml"},
       "genau EIN Teil fehlt: calcChain (bewusst) – verloren: %s" % (alt - neu))
pruefe(not (neu - alt), "und es kommt kein Teil hinzu")
with zipfile.ZipFile(QUELLE) as q, zipfile.ZipFile(ziel) as z:
    gleich = [t for t in sorted(neu) if q.read(t) == z.read(t)]
    anders = [t for t in sorted(neu) if q.read(t) != z.read(t)]
pruefe(anders == ["[Content_Types].xml", "xl/_rels/workbook.xml.rels",
                  "xl/workbook.xml", "xl/worksheets/sheet7.xml"],
       "nur die vier noetigen Teile sind veraendert (%s)" % anders)
pruefe("xl/charts/chart1.xml" in gleich and "xl/sharedStrings.xml" in gleich
       and "xl/printerSettings/printerSettings1.bin" in gleich
       and "xl/worksheets/_rels/sheet7.xml.rels" in gleich,
       "Diagramm, sharedStrings, Druckeinstellungen und Blatt-Beziehungen "
       "byte-gleich uebernommen")
pruefe("xl/worksheets/sheet3.xml" in gleich,
       "das NICHT bearbeitete Blatt ist unangetastet")
pruefe(wohlgeformt(ziel) == [], "alle XML-Teile sind wohlgeformt")

with zipfile.ZipFile(ziel) as z:
    wbx = z.read("xl/workbook.xml").decode("utf-8")
    ctx = z.read("[Content_Types].xml").decode("utf-8")
    rlx = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")
pruefe('fullCalcOnLoad="1"' in wbx,
       "fullCalcOnLoad gesetzt – sonst zeigt Excel alte Summen zu neuen Zahlen")
pruefe('calcId="191029"' in wbx, "und das vorhandene calcPr bleibt sonst erhalten")
pruefe("calcChain" not in ctx, "Content-Types-Eintrag der calcChain entfernt")
pruefe("calcChain" not in rlx, "Beziehung auf die calcChain entfernt")
pruefe('Id="rId1"' in rlx and 'Id="rId2"' in rlx,
       "die anderen Beziehungen bleiben (sonst faende Excel die Blaetter nicht)")
pruefe(not list(TMP.glob("*.patch.tmp")), "keine Zwischendatei bleibt liegen")

# Zweites Blatt patchen – Name mit Entity.
z2 = TMP / "blatt2.xlsx"
b2 = xp.patch_datei(QUELLE, z2, {"A & B": {"B2": "x"}})
pruefe(b2["geschrieben"] == ["A & B!B2"], "auch das zweite Blatt ist erreichbar")
pruefe(wohlgeformt(z2) == [], "und das Ergebnis ist wohlgeformt")

# calcPr fehlt ganz -> muss ergaenzt werden.
pruefe('fullCalcOnLoad="1"' in xp._voll_neuberechnen("<workbook></workbook>"),
       "fehlendes calcPr wird ergaenzt")
pruefe(xp._voll_neuberechnen('<calcPr fullCalcOnLoad="0"/>')
       == '<calcPr fullCalcOnLoad="1"/>', "ein vorhandenes 0 wird auf 1 gesetzt")


# ═══════════════════════════════════════════════════════════════════════════
print("\n7) Fehlerfaelle – laut und ohne halbe Datei")
# ═══════════════════════════════════════════════════════════════════════════
for kwargs, wort in (
        ({"Gibtsnicht": {"A1": 1}}, "gibt es nicht"),
        ({}, "Keine Aenderung"),
        ({"Daten": {}}, "Keine Aenderung")):
    ziel_f = TMP / "fehl.xlsx"
    try:
        xp.patch_datei(QUELLE, ziel_f, kwargs)
        pruefe(False, "%r wird abgelehnt" % (list(kwargs) or "leer"))
    except xp.PatchFehler as e:
        pruefe(wort.lower() in str(e).lower(),
               "%r wird abgelehnt: %s" % (list(kwargs) or "leer", str(e)[:60]))
    pruefe(not ziel_f.exists(), "und es entsteht KEINE Zieldatei")

keine = TMP / "keine.xlsx"
try:
    xp.patch_datei(keine, TMP / "x.xlsx", {"Daten": {"A1": 1}})
    pruefe(False, "fehlende Quelle wird gemeldet")
except xp.PatchFehler as e:
    pruefe("nicht gefunden" in str(e), "fehlende Quelle wird gemeldet")

keintext = TMP / "keintext.xlsx"
keintext.write_bytes(b"das ist kein ZIP")
try:
    xp.patch_datei(keintext, TMP / "y.xlsx", {"Daten": {"A1": 1}})
    pruefe(False, "eine umbenannte Fremddatei wird gemeldet")
except xp.PatchFehler as e:
    pruefe("kein ZIP" in str(e), "eine umbenannte Fremddatei wird gemeldet")

# Die Quelle darf NIE angefasst werden.
_vorher = QUELLE.read_bytes()
xp.patch_datei(QUELLE, TMP / "nochmal.xlsx", {"Daten": {"A1": 1}})
pruefe(QUELLE.read_bytes() == _vorher, "die Quelldatei bleibt unveraendert")


# ═══════════════════════════════════════════════════════════════════════════
print("\n8) Verdrahtung in xlsx_edit: Patch ist der REGELWEG")
# ═══════════════════════════════════════════════════════════════════════════
QT = (ROOT / "skills" / "office" / "tabellen.py").read_text(encoding="utf-8")


def ohne_kommentare(t: str) -> str:
    t = re.sub(r'""".*?"""', "", t, flags=re.DOTALL)
    return "\n".join(z.split("#", 1)[0] for z in t.split("\n"))


import ast                                                   # noqa: E402
_baum = ast.parse(QT)
_zeilen = QT.split("\n")
_edit = ""
for k in ast.walk(_baum):
    if isinstance(k, ast.ClassDef) and k.name == "EditTool":
        _edit = "\n".join(_zeilen[k.lineno - 1:k.end_lineno])
pruefe(bool(_edit), "EditTool gefunden")
_e = ohne_kommentare(_edit)
pruefe("xlsx_patch" in _e, "xlsx_edit benutzt das Patch-Modul")
_i_patch = _e.find("patch_datei")
_i_open = _e.find("_oeffnen(path, schreibend=True)")
pruefe(0 <= _i_patch < _i_open,
       "der Patch-Weg kommt VOR dem openpyxl-Rueckfall")
pruefe("PatchFehler" in _e, "ein Patch-Fehlschlag wird gefangen")
pruefe("HINWEIS_AN_NUTZER" in _e,
       "der Rueckfall ist NICHT still – er sagt, was verloren gehen kann")
pruefe("shared" in _e.lower(),
       "eine abgelehnte Master-Zelle wird dem Modell erklaert")
pruefe("_weiterarbeiten_hinweis" in _e,
       "der Hinweis auf den naechsten Schritt bleibt erhalten")
# Der Patch-Weg darf den Verlusthinweis NICHT setzen – dort geht nichts verloren.
_patchteil = _e[:_i_open] if _i_open > 0 else _e
pruefe("_verlust_hinweis" not in _patchteil,
       "im Patch-Weg wird kein Verlust behauptet (es geht keiner verloren)")


# ═══════════════════════════════════════════════════════════════════════════
print("\n9) Mit openpyxl (falls vorhanden): ein echter Konsument liest die Datei")
# ═══════════════════════════════════════════════════════════════════════════
try:
    from openpyxl import load_workbook
except Exception:  # noqa: BLE001
    print("  (openpyxl nicht vorhanden – uebersprungen)")
else:
    z3 = TMP / "gelesen.xlsx"
    xp.patch_datei(QUELLE, z3, {"Daten": {"A1": "Text", "B5": 12.5,
                                          "D9": "=SUM(A1:A2)", "C1": None}})
    wb = load_workbook(z3, data_only=False)
    ws = wb["Daten"]
    pruefe(ws["A1"].value == "Text", "Text kommt an")
    pruefe(ws["B5"].value == 12.5, "Kommazahl kommt an")
    pruefe(ws["D9"].value == "=SUM(A1:A2)", "Formel kommt an")
    pruefe(ws["C1"].value is None, "geleerte Zelle ist leer")
    pruefe(ws["A5"].value == 5, "unberuehrte Zelle unveraendert")
    pruefe(ws["B2"].value == "=SUM(A1:A1)", "die Master-Formel steht noch")
    pruefe(wb.sheetnames == ["Daten", "A & B"], "Blattnamen und Reihenfolge")
    # Der Stil der ersetzten Zelle muss erhalten sein – sonst wird aus einem
    # Datum eine fuenfstellige Zahl.
    q = load_workbook(QUELLE, data_only=False)["Daten"]
    pruefe(ws["A5"].number_format == q["A5"].number_format,
           "das Zahlenformat einer unberuehrten Zelle bleibt (%s)"
           % ws["A5"].number_format)
    z4 = TMP / "stil.xlsx"
    xp.patch_datei(QUELLE, z4, {"Daten": {"A5": 7}})
    ws4 = load_workbook(z4, data_only=False)["Daten"]
    pruefe(ws4["A5"].value == 7
           and ws4["A5"].number_format == q["A5"].number_format,
           "und auch das Format der GEAENDERTEN Zelle (%s)"
           % ws4["A5"].number_format)
    wb.close()


print("\n%s\n  %d bestanden, %d fehlgeschlagen  (Sandkasten: %s)\n%s"
      % ("=" * 62, _ok, _fail, TMP, "=" * 62))
shutil.rmtree(TMP, ignore_errors=True)
sys.exit(1 if _fail else 0)
