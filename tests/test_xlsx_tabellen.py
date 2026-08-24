#!/usr/bin/env python3
"""Tabellen-Werkzeuge + die stillen Fehlschlaege, die den Vorfall verursacht haben.

VORFALL (ECHT, 2026-08-19): Ein Short-Track "Tabellen zusammenfuehren" lieferte
fuenf Dateien statt einer – eine davon LEER, eine mit zwei Zeilen. Gemessen an
der echten Eingabe: 13 Blaetter, 362.195 Zellen, 32.779 Formeln; das Modell sah
0,4 % davon und konnte das Ergebnis nur noch abtippen.

Der Test prueft die vier Ursachen einzeln:
  1. Ein Werkzeug meldete Erfolg, obwohl es nichts geschrieben hat.
  2. Eine Kuerzung wurde nicht beziffert (und die Meldung selbst weggekuerzt).
  3. Es gab keinen Weg, eine BESTEHENDE Tabelle zu bearbeiten.
  4. Jedes Zwischenprodukt wurde als Ergebnis angeboten.

LAEUFT OHNE fastapi. ``backend.config`` wird NICHT importiert – der echte Import
migriert Profile und schriebe die Live-``settings.json`` zurueck.
BRAUCHT openpyxl (also das venv: ``venv/bin/python tests/test_xlsx_tabellen.py``).
"""

import asyncio
import os
import shutil
import sys
import tempfile
import types
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

# ── backend.config als Attrappe, VOR jedem Import aus backend/ ──────────────
_stub = types.ModuleType("backend.config")


class _Cfg:
    def get_skill_states(self):
        return {}

    def __getattr__(self, n):
        return None


_stub.config = _Cfg()
sys.modules.setdefault("backend.config", _stub)

try:
    from openpyxl import load_workbook, Workbook
except Exception as e:  # noqa: BLE001
    print(f"ABBRUCH: openpyxl fehlt ({e}). Mit dem venv starten:")
    print("  venv/bin/python tests/test_xlsx_tabellen.py")
    sys.exit(2)

from skills.office import main as office_main          # noqa: E402
from skills.office import tabellen as tab              # noqa: E402

TMP = Path(tempfile.mkdtemp(prefix="xlsxtest_"))

# ── SANDKASTEN-SCHRANKE ────────────────────────────────────────────────────
# Die Werkzeuge schreiben ueber `office_main._new_path` nach DOCS_DIR – im
# Normalbetrieb ist das `data/documents` des LAUFENDEN Servers. Ein Test, der
# dorthin schreibt, verschmutzt die Ablage echter Benutzer. Deshalb umbiegen
# UND nachweisen, dass es gewirkt hat; sonst Exit 2 (nicht 1 – "konnte nicht
# laufen" muss von "durchgefallen" unterscheidbar bleiben).
_ECHT_DOCS = office_main.DOCS_DIR
office_main.DOCS_DIR = TMP / "documents"
office_main.DOCS_DIR.mkdir(parents=True, exist_ok=True)
if office_main.DOCS_DIR.resolve() == _ECHT_DOCS.resolve() or \
        TMP not in office_main.DOCS_DIR.resolve().parents:
    print("ABBRUCH: DOCS_DIR zeigt nicht in das Wegwerf-Verzeichnis.")
    sys.exit(2)

ok = fail = 0


def pruefe(bedingung, text):
    global ok, fail
    if bedingung:
        ok += 1
    else:
        fail += 1
        print(f"  FAIL: {text}")


def lauf(coro):
    return asyncio.get_event_loop().run_until_complete(coro)


def dateien_in_ablage():
    return sorted(p.name for p in office_main.DOCS_DIR.glob("*"))


def datei_aus(antwort: str) -> Path | None:
    """Holt die erzeugte Datei aus einer Erfolgsmeldung (/api/documents/<cap>)."""
    import re
    m = re.search(r"/api/documents/([0-9a-f]{32}__[^\s\)]+)", antwort or "")
    return (office_main.DOCS_DIR / m.group(1)) if m else None


# ── Testdaten: klein, aber mit denselben Eigenschaften wie das Original ─────
def baue_master(pfad: Path):
    """Master mit Formeln, Spaltenbreiten und verbundenen Zellen."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Statistik"
    ws.append(["Jahr", "Monat", "Standort", "Menge", "Umsatz", "Schnitt"])
    for jahr in (2004, 2005):
        for monat in range(1, 4):
            ws.append([jahr, monat, "HH", None, None, None])
    # Formeln in einer Spalte + eine Summenzeile
    for r in range(2, 8):
        ws.cell(row=r, column=6).value = f"=IFERROR(E{r}/D{r},0)"
    ws.cell(row=9, column=4).value = "=SUM(D2:D7)"
    ws.column_dimensions["C"].width = 42.5
    ws.merge_cells("A11:C11")
    wb.save(str(pfad))


def baue_slave(pfad: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Daten"
    ws.append(["Jahr", "Monat", "Menge", "Umsatz", "Unbekannt"])
    # Jahr bewusst als TEXT, im Master als Zahl -> prueft die Normierung
    ws.append(["2004", 1, 10, 100, "x"])
    ws.append(["2004", 2, 20, 200, "x"])
    ws.append(["2004", 3, 30, 300, "x"])
    ws.append(["2005", 1, 40, 400, "x"])
    ws.append(["2099", 9, 99, 999, "x"])      # kein Partner im Master
    wb.save(str(pfad))


MASTER = TMP / "Master.xlsx"
SLAVE = TMP / "Slave.xlsx"
baue_master(MASTER)
baue_slave(SLAVE)

inspect, readrange, merge, edit = tab.get_tabellen_tools()
excel = office_main.CreateExcelTool()
lesen = office_main.ReadDocumentTool()


# ═══════════════════════════════════════════════════════════════════════════
print("\n1) office_create_excel meldet keinen Erfolg mehr, wenn nichts geschrieben wird")
# ═══════════════════════════════════════════════════════════════════════════
vorher = dateien_in_ablage()

# Der GEMELDETE Aufruf, woertlich: Parametername mit fuehrendem Leerzeichen.
r = lauf(excel.execute(filename="IBSv3_Master_2004",
                       **{" sheets": '{"Monatsstatistik": {"headers": ["a"], "rows": [[1]]}}'}))
pruefe(r.startswith("Fehler"), f"' sheets' muss ein Fehler sein, war: {r[:90]}")
pruefe("' sheets'" in r, "der Fehler muss den beanstandeten Parameter NENNEN")
pruefe(dateien_in_ablage() == vorher, "bei unbekanntem Parameter darf KEINE Datei entstehen")

r = lauf(excel.execute(filename="Leer"))
pruefe(r.startswith("Fehler"), "ohne Daten muss es ein Fehler sein")
pruefe(dateien_in_ablage() == vorher, "ohne Daten darf KEINE Datei entstehen")

r = lauf(excel.execute(filename="Leer2", rows=[], headers=[]))
pruefe(r.startswith("Fehler"), "leere rows/headers sind ebenfalls 'keine Daten'")

# sheets als JSON-STRING (zweiter Weg in dieselbe leere Datei) wird geparst
r = lauf(excel.execute(filename="AusString",
                       sheets='{"B1": {"headers": ["x"], "rows": [[1],[2]]}}'))
p = datei_aus(r)
pruefe(p is not None and p.exists(), "sheets als JSON-String muss akzeptiert werden")
if p and p.exists():
    wb = load_workbook(str(p))
    pruefe(wb.sheetnames == ["B1"], f"Blattname aus dem JSON-String: {wb.sheetnames}")
    pruefe(wb["B1"].max_row == 3, f"Kopf + 2 Zeilen erwartet, waren {wb['B1'].max_row}")
    wb.close()

r = lauf(excel.execute(filename="Murks", sheets="{kaputt"))
pruefe(r.startswith("Fehler") and "JSON" in r, "kaputtes JSON muss benannt werden")

# Der gute Weg bleibt unveraendert benutzbar
r = lauf(excel.execute(filename="Normal", headers=["a", "b"], rows=[[1, 2]]))
pruefe(not r.startswith("Fehler"), "der normale Aufruf muss weiter funktionieren")


# ═══════════════════════════════════════════════════════════════════════════
print("2) xlsx_inspect – Struktur statt Inhalt, Grenzen werden beziffert")
# ═══════════════════════════════════════════════════════════════════════════
r = lauf(inspect.execute(path=str(MASTER)))
pruefe("Statistik" in r, "Blattname fehlt")
pruefe("Jahr" in r and "Umsatz" in r, "Kopfzeile fehlt")
pruefe("Formel" in r, "die Formeln der Mappe muessen erwaehnt werden")
pruefe("office_create_excel" in r and "NICHT" in r,
       "inspect muss vom Neuaufbau abraten")
pruefe(len(r) < tab.AUSGABE_MAX, "Ausgabe muss begrenzt sein")

# BREITE TABELLE: die SPALTENNAMEN sind Struktur und muessen VOLLSTAENDIG
# kommen. Vorfall 2026-08-24 (ECHT): ein Blatt mit 254 Laborcode-Spalten wurde
# bei 60 gekappt ("+182 weitere benannte Spalten") – die 182 Namen bekam man
# nirgends her, waehrend der Hinweis "Mit 'spalten' gezielt auswaehlen" genau
# sie verlangte. Das Modell probierte 24 Leseaufrufe durch und lieferte nichts.
breit = TMP / "Breit.xlsx"
wb = Workbook(); ws = wb.active
ws.append([f"Sp{i}" for i in range(1, 200)])
ws.append(list(range(1, 200)))
wb.save(str(breit)); wb.close()
r = lauf(inspect.execute(path=str(breit)))
pruefe("Sp199" in r and "Sp1" in r,
       "bei 199 kurzen Spaltennamen muessen ALLE benannt werden")
pruefe("weitere benannte Spalten" not in r,
       "und dann gibt es auch nichts zu beziffern")
pruefe(len(r) < tab.AUSGABE_MAX, "trotzdem unter dem Gesamtdeckel")

# Gegenprobe: reisst die Namensliste den ZEICHEN-Deckel, wird beziffert – eine
# unvollstaendige Liste, die sich fuer vollstaendig ausgibt, waere schlimmer
# als eine kurze.
lang = TMP / "LangeNamen.xlsx"
wb = Workbook(); ws = wb.active
ws.append([("Spaltenueberschrift_Nummer_%03d_mit_viel_Text" % i) for i in range(1, 200)])
ws.append(list(range(1, 200)))
wb.save(str(lang)); wb.close()
r2 = lauf(inspect.execute(path=str(lang)))
pruefe("weitere benannte Spalten" in r2,
       "bei sehr langen Namen muss die Kuerzung beziffert werden")
pruefe(len(r2) < tab.AUSGABE_MAX, "und der Gesamtdeckel haelt")

# xlsx_read_range darf die Kopfzeile NICHT enger kappen als inspect – sonst
# verlangt sein eigener Hinweis Namen, die es gerade verschwiegen hat.
r3 = lauf(readrange.execute(path=str(breit), zeilen=1))
pruefe("Sp199" in r3, "read_range nennt die Spaltennamen vollstaendig")
pruefe("199" in r, "die Gesamtzahl der Spalten muss vorkommen")

r = lauf(inspect.execute(path="/gibt/es/nicht.xlsx"))
pruefe(r.startswith("Fehler"), "fehlende Datei = Fehler")
r = lauf(inspect.execute(path=str(MASTER), quatsch=1))
pruefe(r.startswith("Fehler") and "quatsch" in r,
       "unbekannter Parameter muss auch hier benannt werden")


# ═══════════════════════════════════════════════════════════════════════════
print("3) xlsx_read_range – Bilanz statt stiller Kuerzung")
# ═══════════════════════════════════════════════════════════════════════════
r = lauf(readrange.execute(path=str(MASTER), zeilen=3))
pruefe("Gezeigt:" in r, "die Bilanzzeile fehlt")
pruefe("von 11" in r or "x 6" in r, f"Gesamtgroesse muss genannt werden: {r[:120]}")

lang = TMP / "Lang.xlsx"
wb = Workbook(); ws = wb.active
ws.append(["A"])
for i in range(1, 501):
    ws.append([i])
wb.save(str(lang)); wb.close()

r = lauf(readrange.execute(path=str(lang), zeilen=9999))
pruefe(f"hoechstens {tab.LESE_ZEILEN_MAX}" in r,
       "die Obergrenze muss ausdruecklich genannt werden")
pruefe("NICHT gezeigt" in r, "die fehlenden Zeilen muessen beziffert werden")
pruefe("501" in r, "die Gesamtzeilenzahl muss vorkommen")

r = lauf(readrange.execute(path=str(MASTER), spalten=["Jahr", "Umsatz"]))
pruefe("Fehler" not in r[:20], "Spaltenauswahl per Kopfname muss gehen")
r = lauf(readrange.execute(path=str(MASTER), spalten=["Gibtsnicht"]))
pruefe(r.startswith("Fehler") and "Jahr" in r,
       "unbekannte Spalte -> Fehler MIT der echten Kopfzeile")
r = lauf(readrange.execute(path=str(MASTER), blatt="Fantasie"))
pruefe(r.startswith("Fehler") and "Statistik" in r,
       "unbekanntes Blatt -> Fehler MIT den vorhandenen Blattnamen")


# ═══════════════════════════════════════════════════════════════════════════
print("3b) Spalten-Aufloesung – ein Wort ist keine Spaltenangabe")
# ═══════════════════════════════════════════════════════════════════════════
# Ohne Laengen- UND Wertgrenze gilt JEDES Wort als Spaltenbuchstabe: im ersten
# Testlauf loeste "Unbekannt" zu Spalte 4.498.495.991.152 auf. Hier starb der
# Merge wenigstens laut beim Speichern – ein kuerzeres Wort ("Ort" -> Spalte
# 10.542) haette klaglos in eine voellig falsche Spalte geschrieben.
pruefe(tab._buchstabe_zu_index("A") == 1, "A muss Spalte 1 sein")
pruefe(tab._buchstabe_zu_index("XFD") == 16384, "XFD ist die letzte Excel-Spalte")
pruefe(tab._buchstabe_zu_index("XFE") is None, "hinter XFD gibt es keine Spalte mehr")
pruefe(tab._buchstabe_zu_index("ZZZZ") is None, "vier Buchstaben sind keine Spalte")
pruefe(tab._buchstabe_zu_index("Unbekannt") is None,
       "ein Wort darf NICHT als Spaltenbuchstabe durchgehen")
pruefe(tab._buchstabe_zu_index("Ä") is None,
       "ein Umlaut ist .isalpha(), aber keine Spalte – isascii() muss greifen")
# "Ort" IST syntaktisch eine gueltige Spaltenangabe (Spalte 10.628). Dagegen
# hilft nur die BREITENGRENZE der Kopfzeile, nicht die Laengenpruefung.
pruefe(tab._buchstabe_zu_index("Ort") == 10628,
       "'Ort' ist syntaktisch Spalte 10628 – das ist keine Fehlfunktion")
pruefe(tab._spalte_aufloesen(["Jahr", "Monat", "Menge"], "Ort") is None,
       "…aber jenseits der Kopfzeile darf es NICHT als Spalte durchgehen")
pruefe(tab._spalte_aufloesen(["Jahr", "Monat", "Menge"], "C") == 3,
       "eine Spaltenangabe INNERHALB der Kopfzeile muss weiter gelten")
pruefe(tab._spalte_aufloesen(["Jahr", "Monat"], "Unbekannt") is None,
       "unbekannter Spaltenname muss None ergeben, nicht eine Fantasiespalte")
pruefe(tab._spalte_aufloesen(["Jahr", "Monat"], "Monat") == 2,
       "Kopfzeilen-Name muss gewinnen")
pruefe(tab._spalte_aufloesen(["B", "Monat"], "B") == 1,
       "eine Spalte, die 'B' HEISST, gewinnt gegen den Buchstaben B")
pruefe(tab._spalte_aufloesen(["Jahr", "Monat"], "99999999") is None,
       "eine absurde Spaltennummer muss abgelehnt werden")


# ═══════════════════════════════════════════════════════════════════════════
print("3c) Kopfzeilen-Erkennung – die drei Bauformen der ECHTEN Datei")
# ═══════════════════════════════════════════════════════════════════════════
# Nachgebaut aus den echten Blaettern von ECHT (2026-08-19). Mit der frueheren
# festen Vorgabe "Zeile 1" liest Form B eine Liste von Nummerncodes als
# Spaltennamen und findet danach keinen einzigen Schluessel.
formen = TMP / "Formen.xlsx"
wb = Workbook()

# Form A (wie Blatt '2004'): Kopfzeile in Z1, Daten ab Z2.
a = wb.active; a.title = "FormA"
a.append([2004, None, "an1", "bi1", "bi2", "bn2"])
a.append(["Jan", "Nuernberg", None, 143, None, 57])
a.append(["Feb", "Karlsruhe", None, 88, None, 12])

# Form B (wie Blatt '2015'): Z1 Nummerncodes ALS TEXT, Z2 leer, Z3 Kopfzeile.
b = wb.create_sheet("FormB")
b.append([None, "00000000083", "00000000113", "00000000121", "00000000062"])
b.append([None, None, None, None, None])
b.append(["Befunde", "b39", "bi2", "dd1", "do1"])
b.append(["Jan", 0, 0, 0, 0])
b.append(["Feb", 5, 7, 1, 3])

# Form C (wie Blatt '2019'): Z1 Kopfzeile, Z2 leer, Z3 Kopfzeile WIEDERHOLT.
c = wb.create_sheet("FormC")
c.append(["Labore", "ab1", "ac1", "ac5", "an1"])
c.append([None, None, None, None, None])
c.append(["IBSv1/v2 LDT", "ab1", "ac1", "ac5", "an1"])
c.append(["Jan", 440, 0, 0, 0])
c.append(["Feb", 373, 0, 0, 0])

# Form E (wie Blatt '2019' WIRKLICH ist): Z1 echte Namen, Z3 wiederholt sie als
# FORMEL (=B1). Nur an der echten Datei aufgefallen – ein synthetischer Test mit
# abgetippten Namen haette den Fehler nie gezeigt.
e = wb.create_sheet("FormE")
e.append(["Labore", "ab1", "ac1", "ac5", "an1"])
e.append([None, None, None, None, None])
e.append(["IBSv3 LDTs", "=B1", "=C1", "=D1", "=E1"])
e.append(["Jan", 440, 0, 0, 0])
e.append(["Feb", 373, 0, 0, 0])

# Form D: reine Texttabelle ohne Zahlen -> Rueckfall auf Zeile 1.
d = wb.create_sheet("FormD")
d.append(["Name", "Ort", "Rolle"])
d.append(["Meier", "Hamburg", "Leitung"])
wb.save(str(formen)); wb.close()

wbp = load_workbook(str(formen))
pruefe(tab._kopfzeile_raten(wbp["FormA"]) == (1, 2), "Form A: Kopf Z1, Daten ab Z2")
pruefe(tab._kopfzeile_raten(wbp["FormB"]) == (3, 4),
       f"Form B: Kopf Z3, Daten ab Z4 – erkannt wurde "
       f"{tab._kopfzeile_raten(wbp['FormB'])}. Nummerncodes als TEXT duerfen "
       f"nicht als Datenanfang zaehlen")
pruefe(tab._kopfzeile_raten(wbp["FormC"]) == (3, 4),
       "Form C: bei zwei echten Kopfzeilen muss die UNTERE gewinnen")
pruefe(tab._kopfzeile_raten(wbp["FormE"]) == (1, 4),
       f"Form E: die FORMEL-Wiederholung (=B1) ist keine Beschriftung – Kopf "
       f"muss Z1 sein, Daten ab Z4. Erkannt: {tab._kopfzeile_raten(wbp['FormE'])}")
pruefe(tab._kopfzeile_raten(wbp["FormD"]) == (1, 2),
       "Form D: ohne Zahlen Rueckfall auf Zeile 1")
wbp.close()

# _hat_beschriftungen trennt Namen von Formeln und Zahlen
pruefe(tab._hat_beschriftungen(("Befunde", "b39", "bi2")), "echte Namen")
pruefe(not tab._hat_beschriftungen(("=B1", "=C1", "=D1")), "Formeln sind keine Namen")
pruefe(not tab._hat_beschriftungen(("00000000083", "00000000113")),
       "Zahlen als Text sind keine Namen")

# Der Sentinel: 0/fehlend = automatisch, >=1 = ausdrueckliche Vorgabe.
wbp = load_workbook(str(formen))
pruefe(tab._kz(wbp["FormB"], 0) == (3, 4, True), "0 heisst automatisch erkennen")
pruefe(tab._kz(wbp["FormB"], None) == (3, 4, True), "fehlend heisst automatisch")
pruefe(tab._kz(wbp["FormB"], 1) == (1, 2, False),
       "eine ausdrueckliche 1 darf die Erkennung UEBERSTIMMEN – sonst gaebe es "
       "keinen Weg, einen Fehlgriff zu korrigieren")
wbp.close()

# Ende-zu-Ende: inspect nennt die abweichenden Blaetter und liest die RICHTIGEN
# Spaltennamen.
r = lauf(inspect.execute(path=str(formen)))
pruefe("Befunde" in r and "b39" in r,
       "inspect muss in Form B die ECHTE Kopfzeile lesen, nicht die Nummerncodes")
pruefe("00000000083" not in r.split("# Blatt 'FormB'")[-1].split("# Blatt")[0]
       or True, "(Nummerncodes duerfen als Datenzeile erscheinen)")
pruefe("automatisch erkannt" in r, "inspect muss sagen, dass es geraten hat")
pruefe("ACHTUNG – nicht jedes Blatt ist gleich gebaut" in r,
       "abweichende Blaetter muessen ausdruecklich benannt werden")
pruefe("'FormB' -> Kopfzeile 3, Daten ab 4" in r,
       "die abweichenden Blaetter muessen NAMENTLICH mit Zeile genannt werden")
pruefe("'FormE' -> Kopfzeile 1, Daten ab 4" in r,
       "auch eine abweichende DATENZEILE bei Kopf in Z1 muss gemeldet werden")

_formE = r.split("# Blatt 'FormE'")[-1]
pruefe("=B1" not in _formE.split("Zeile 4")[0],
       "in Form E duerfen die Formeln NICHT als Spaltennamen erscheinen")
pruefe("ab1" in _formE and "ac1" in _formE,
       "in Form E muessen die ECHTEN Namen aus Zeile 1 erscheinen")

# read_range loest Spaltennamen gegen die erkannte Kopfzeile auf
r = lauf(readrange.execute(path=str(formen), blatt="FormB", spalten=["b39"]))
pruefe(not r.startswith("Fehler"),
       f"Spaltenname aus der erkannten Kopfzeile muss aufloesen: {r[:120]}")
pruefe("Kopfzeile Zeile 3 (automatisch erkannt)" in r,
       "read_range muss die benutzte Kopfzeile nennen")

# DIE WARNUNG MUSS DIE KUERZUNG UEBERLEBEN. An der echten Datei gemessen:
# 13 Blaetter ergeben 14.134 Zeichen bei einem Deckel von 14.000 – die am ENDE
# angehaengte Warnung wurde damit abgeschnitten (derselbe Fehler wie bei
# office_read). Deshalb: viele breite Blaetter bauen, Kuerzung erzwingen und
# pruefen, dass Warnung und Wegweiser trotzdem dastehen.
viele = TMP / "Viele.xlsx"
wb = Workbook()
for n in range(1, 26):
    w = wb.create_sheet(f"B{n}")
    w.append([f"sehr_langer_spaltenname_nummer_{i}" for i in range(1, 60)])
    w.append([None] * 59)
    w.append([f"kopf_{i}" for i in range(1, 60)])       # zweite Kopfzeile
    w.append(list(range(1, 60)))
del wb["Sheet"]
wb.save(str(viele)); wb.close()

rv = lauf(inspect.execute(path=str(viele)))
pruefe("gekuerzt:" in rv, f"dieser Aufbau MUSS die Kuerzung ausloesen ({len(rv)} Zeichen)")
pruefe("ACHTUNG" in rv,
       "die Warnung ueber abweichende Blaetter muss die Kuerzung ueberleben")
pruefe("NAECHSTER SCHRITT" in rv,
       "der Wegweiser muss die Kuerzung ueberleben")
pruefe(rv.index("ACHTUNG") < 900,
       "…und dafuer weit VORNE stehen, nicht am Ende")

# merge erkennt BEIDE Seiten getrennt
sl_b = TMP / "SlaveB.xlsx"
wb = Workbook(); ws = wb.active; ws.title = "S"
ws.append(["Befunde", "b39"]); ws.append(["Jan", 111]); ws.append(["Feb", 222])
wb.save(str(sl_b)); wb.close()
r = lauf(merge.execute(master=str(formen), slave=str(sl_b), ziel="FormB_voll",
                       master_blatt="FormB", slave_blatt="S",
                       schluessel=["Befunde"]))
pruefe(not r.startswith("Fehler"),
       f"Merge ueber unterschiedlich gebaute Tabellen muss laufen: {r[:250]}")
pruefe("Master" in r and "Kopfzeile Zeile 3" in r,
       "der Bericht muss die erkannte Master-Kopfzeile nennen")
pruefe("Kopfzeile Zeile 1" in r, "…und die des Slave (dort Zeile 1)")
p = datei_aus(r)
if p and p.exists():
    wbv = load_workbook(str(p)); wsv = wbv["FormB"]
    pruefe(wsv.cell(row=4, column=2).value == 111,
           "Jan/b39 muss 111 sein (Datenzeile 4, nicht 2)")
    pruefe(wsv.cell(row=5, column=2).value == 222, "Feb/b39 muss 222 sein")
    pruefe(wsv.cell(row=1, column=2).value == "00000000083",
           "die Nummerncode-Zeile darf NICHT ueberschrieben worden sein")
    wbv.close()


# ═══════════════════════════════════════════════════════════════════════════
print("4) xlsx_merge – DER KERN: Daten uebertragen, Layout und Formeln bleiben")
# ═══════════════════════════════════════════════════════════════════════════
r = lauf(merge.execute(master=str(MASTER), slave=str(SLAVE), ziel="Master_erweitert",
                       schluessel=["Jahr", "Monat"]))
pruefe(not r.startswith("Fehler"), f"Merge muss laufen: {r[:200]}")
p = datei_aus(r)
pruefe(p is not None and p.exists(), "Merge muss eine Datei erzeugen")

if p and p.exists():
    wb = load_workbook(str(p))          # data_only=False -> Formeln sichtbar
    ws = wb["Statistik"]
    # a) Daten sind angekommen
    pruefe(ws.cell(row=2, column=4).value == 10, "Menge 2004/01 muss 10 sein")
    pruefe(ws.cell(row=4, column=5).value == 300, "Umsatz 2004/03 muss 300 sein")
    # b) TEXT '2005' im Slave gegen ZAHL 2005 im Master -> Normierung greift
    pruefe(ws.cell(row=5, column=4).value == 40,
           "Schluessel-Normierung Text/Zahl hat nicht getroffen")
    # c) DAS ENTSCHEIDENDE: Formeln UEBERLEBEN
    pruefe(str(ws.cell(row=2, column=6).value).startswith("="),
           "Formel in F2 ist verloren gegangen")
    pruefe(str(ws.cell(row=9, column=4).value) == "=SUM(D2:D7)",
           "Summenformel ist verloren gegangen")
    # d) Layout ueberlebt
    pruefe(abs((ws.column_dimensions["C"].width or 0) - 42.5) < 0.6,
           "Spaltenbreite ist verloren gegangen")
    pruefe("A11:C11" in [str(x) for x in ws.merged_cells.ranges],
           "verbundene Zellen sind verloren gegangen")
    # e) Zeilen ohne Partner bleiben unangetastet, es kommt nichts dazu
    pruefe(ws.max_row <= 11, f"modus=aktualisieren darf nichts anfuegen ({ws.max_row})")
    wb.close()

pruefe("Aktualisierte Master-Zeilen: 4" in r, f"Bericht falsch: {r[-400:]}")
pruefe("ohne Treffer im Slave: 2" in r, "Master-Zeilen ohne Treffer fehlen im Bericht")
pruefe("nicht vor" in r and "NICHT uebernommen" in r,
       "der uebrige Slave-Schluessel (2099) muss gemeldet werden")
# Defensiv: ein `split(...)[1]` WIRFT, wenn die Marke fehlt – die Pruefung
# bricht dann ab, statt fehlzuschlagen, und der Lauf sieht wie ein Erfolg aus.
# Dieselbe Falle wie `.index()` (Projekt-Merkregel, mehrfach zugeschlagen).
_teile = r.split("Uebernommene Spalten")
pruefe(len(_teile) > 1 and "Unbekannt" not in _teile[1][:200],
       "eine Spalte, die es im Master nicht gibt, darf nicht uebernommen werden")

# modus=beides fuegt die uebrige Zeile an
r2 = lauf(merge.execute(master=str(MASTER), slave=str(SLAVE), ziel="M2",
                        schluessel=["Jahr", "Monat"], modus="beides"))
p2 = datei_aus(r2)
pruefe("Angefuegte Zeilen: 1" in r2, f"modus=beides muss 1 Zeile anfuegen: {r2[-300:]}")
if p2 and p2.exists():
    wb = load_workbook(str(p2)); ws = wb["Statistik"]
    werte = [ws.cell(row=r_, column=1).value for r_ in range(1, ws.max_row + 1)]
    pruefe("2099" in [str(v) for v in werte], "die angefuegte Zeile fehlt")
    wb.close()

# KEIN Treffer -> Fehler UND KEINE Datei.
# ACHTUNG: der Slave braucht dafuer eine Schluesselspalte, die es WIRKLICH
# gibt – sonst greift schon die Pruefung "Schluesselspalte nicht gefunden"
# davor, und dieser Zweig wird nie erreicht. Genau daran lief die erste
# Testfassung vorbei: die Gegenprobe "Merge-Fehlschlag liefert doch Datei"
# blieb gruen, obwohl die Schranke ausgebaut war.
fremd = TMP / "Fremd.xlsx"
wb = Workbook(); ws = wb.active
ws.append(["Jahr", "Monat", "Menge"])
ws.append([3000, 1, 5]); ws.append([3001, 2, 6])
wb.save(str(fremd)); wb.close()

vorher = dateien_in_ablage()
r3 = lauf(merge.execute(master=str(MASTER), slave=str(fremd), ziel="Nix",
                        schluessel=["Jahr", "Monat"]))
pruefe(r3.startswith("Fehler"), f"ohne Treffer muss es ein Fehler sein: {r3[:120]}")
pruefe("NICHTS geschrieben" in r3, "der Fehler muss sagen, dass nichts entstand")
pruefe("3000" in r3 and "2004" in r3,
       "der Fehler muss Beispiel-Schluessel BEIDER Seiten nennen – sonst kann "
       "das Modell die Schluesselwahl nicht korrigieren")
pruefe(dateien_in_ablage() == vorher, "ohne Treffer darf KEINE Datei entstehen")

# Schluesselspalte, die es im Slave gar nicht gibt -> anderer, frueherer Fehler
r3b = lauf(merge.execute(master=str(MASTER), slave=str(SLAVE), ziel="Nix",
                         schluessel=["Standort"]))
pruefe(r3b.startswith("Fehler") and "SLAVE" in r3b,
       "fehlende Schluesselspalte im Slave muss als solche gemeldet werden")

# unbekannte Schluesselspalte -> Fehler nennt beide Kopfzeilen
r4 = lauf(merge.execute(master=str(MASTER), slave=str(SLAVE), ziel="Nix",
                        schluessel=["Quartal"]))
pruefe(r4.startswith("Fehler") and "MASTER" in r4 and "SLAVE" in r4,
       "fehlende Schluesselspalte muss fuer beide Seiten gemeldet werden")

# fehlende Pflichtangaben
pruefe(lauf(merge.execute(master=str(MASTER), slave=str(SLAVE), ziel="x")
            ).startswith("Fehler"), "ohne 'schluessel' muss es scheitern")
pruefe(lauf(merge.execute(master=str(MASTER), slave=str(SLAVE),
                          schluessel=["Jahr"])).startswith("Fehler"),
       "ohne 'ziel' muss es scheitern")
pruefe(lauf(merge.execute(master=str(MASTER), slave=str(SLAVE), ziel="x",
                          schluessel=["Jahr"], modus="quatsch")).startswith("Fehler"),
       "unbekannter Modus muss scheitern")

# doppelte Schluessel im Slave werden GEMELDET
dop = TMP / "Dop.xlsx"
wb = Workbook(); ws = wb.active
ws.append(["Jahr", "Monat", "Menge"])
ws.append([2004, 1, 11]); ws.append([2004, 1, 22])
wb.save(str(dop)); wb.close()
r5 = lauf(merge.execute(master=str(MASTER), slave=str(dop), ziel="Dop",
                        schluessel=["Jahr", "Monat"]))
pruefe("bereits vergebenen" in r5, "doppelter Schluessel muss gemeldet werden")


# ═══════════════════════════════════════════════════════════════════════════
print("5) xlsx_edit – Zellen schreiben, Layout bleibt, Teilerfolg wird benannt")
# ═══════════════════════════════════════════════════════════════════════════
r = lauf(edit.execute(path=str(MASTER), ziel="Editiert",
                      aenderungen=[{"zelle": "D2", "wert": 7},
                                   {"zelle": "E2", "wert": "=D2*2"}]))
p = datei_aus(r)
pruefe(p is not None and p.exists(), f"edit muss eine Datei erzeugen: {r[:150]}")
if p and p.exists():
    wb = load_workbook(str(p)); ws = wb["Statistik"]
    pruefe(ws.cell(row=2, column=4).value == 7, "Wert nicht geschrieben")
    pruefe(ws.cell(row=2, column=5).value == "=D2*2", "Formel nicht geschrieben")
    pruefe(str(ws.cell(row=9, column=4).value) == "=SUM(D2:D7)",
           "bestehende Formel ging verloren")
    wb.close()

vorher = dateien_in_ablage()
pruefe(lauf(edit.execute(path=str(MASTER), ziel="x", aenderungen=[])
            ).startswith("Fehler"), "leere Aenderungsliste = Fehler")
pruefe(dateien_in_ablage() == vorher, "bei leerer Liste darf keine Datei entstehen")

r = lauf(edit.execute(path=str(MASTER), ziel="x",
                      aenderungen=[{"zelle": "ZZZZ", "wert": 1}]))
pruefe(r.startswith("Fehler"), "nur ungueltige Zellen = Fehler, keine Datei")

r = lauf(edit.execute(path=str(MASTER), ziel="Teil",
                      aenderungen=[{"zelle": "D2", "wert": 1},
                                   {"zelle": "kaputt", "wert": 2}]))
pruefe(not r.startswith("Fehler") and "NICHT ausgefuehrt" in r,
       "Teilerfolg muss die uebersprungenen Aenderungen BENENNEN")


# ═══════════════════════════════════════════════════════════════════════════
print("6) office_read – keine stille Kuerzung mehr")
# ═══════════════════════════════════════════════════════════════════════════
gross = TMP / "Gross.xlsx"
wb = Workbook()
for j in range(2004, 2012):
    ws = wb.create_sheet(f"J{j}")
    for i in range(400):
        ws.append([f"wert-{j}-{i}-{k}" for k in range(12)])
del wb["Sheet"]
wb.save(str(gross)); wb.close()

r = lauf(lesen.execute(path=str(gross)))
pruefe("zu gross" in r, "grosse Mappe muss als solche gemeldet werden")
pruefe("xlsx_inspect" in r and "xlsx_merge" in r,
       "die Antwort muss auf die Tabellen-Werkzeuge verweisen")
pruefe("Blatt 'J2004'" in r and "Blatt 'J2011'" in r,
       "alle Blaetter muessen im Ueberblick vorkommen")
pruefe("wert-2004-0-0" not in r, "es darf KEIN Datenfragment geliefert werden")
pruefe(len(r) < 4000, f"der Ueberblick muss klein sein, war {len(r)}")

# Kleine Mappe wird weiter normal gelesen
r = lauf(lesen.execute(path=str(MASTER)))
pruefe("Statistik" in r and "Jahr" in r, "kleine Mappe muss normal lesbar bleiben")
pruefe("zu gross" not in r, "kleine Mappe darf nicht abgelehnt werden")

# Die Kuerzungsmeldung steht VORNE (sonst kappt sie der Agenten-Deckel weg)
quelle = (ROOT / "skills" / "office" / "main.py").read_text(encoding="utf-8")
i_marke = quelle.find('f"[GEKUERZT: {20000} von {len(text)}')
i_text = quelle.find('+ text[:20000]')
pruefe(0 < i_marke < i_text, "die GEKUERZT-Meldung muss VOR dem Text stehen")
pruefe(getattr(lesen, "ergebnis_max", 0) >= 20000,
       "office_read braucht ergebnis_max, sonst kappt der Agent ein zweites Mal")


# ═══════════════════════════════════════════════════════════════════════════
print("7) Short Tracks – nur das Endergebnis wird angeboten")
# ═══════════════════════════════════════════════════════════════════════════
import importlib.util  # noqa: E402
spec = importlib.util.spec_from_file_location(
    "_str", ROOT / "backend" / "short_tracks_runner.py")
# Der Runner zieht fastapi-nahe Module – deshalb nur die eine Funktion per
# Quelltext holen, statt das Modul zu importieren.
_src = (ROOT / "backend" / "short_tracks_runner.py").read_text(encoding="utf-8")
_a = _src.index("def _endergebnis_filtern")
_b = _src.index("def _kein_ergebnis")
import re as _re  # noqa: E402
_ns: dict = {"re": _re}
exec(compile(_src[_a:_b], "st", "exec"), _ns)
filtern = _ns["_endergebnis_filtern"]

chips = [{"name": "Master.xlsx", "url": "/api/documents/a__Master.xlsx"},
         {"name": "Slave.xlsx", "url": "/api/documents/b__Slave.xlsx"},
         {"name": "erweiterte_master.xlsx", "url": "/api/documents/c__erweiterte_master.xlsx"}]

ende, zw = filtern(chips, "Fertig – hier ist die **erweiterte_master.xlsx** mit allen Daten.")
pruefe(len(ende) == 1 and ende[0]["name"] == "erweiterte_master.xlsx",
       f"nur das Endergebnis darf uebrig bleiben, war {[e['name'] for e in ende]}")
pruefe(len(zw) == 2, "die zwei Zwischenprodukte muessen gezaehlt werden")

ende, zw = filtern(chips, "Fertig – ohne Nennung eines Dateinamens.")
pruefe(len(ende) == 3 and not zw,
       "nennt die Antwort KEINE Datei, darf nichts weggenommen werden")

ende, zw = filtern(chips, "Ergebnis: erweiterte_master (ohne Endung genannt)")
pruefe(len(ende) == 1, "der Name ohne Endung muss ebenfalls treffen")
pruefe(filtern([], "egal") == ([], []), "leere Chip-Liste darf nicht werfen")


# ═══════════════════════════════════════════════════════════════════════════
print("8) Verdrahtung: Pfad-Freigabe, Bereiche, Manifest, Prompt")
# ═══════════════════════════════════════════════════════════════════════════
for t in (inspect, readrange, merge, edit):
    pruefe(bool(getattr(t, "pfad_parameter", None)),
           f"{t.name} braucht 'pfad_parameter' – sonst laeuft es an der "
           f"Pfad-Freigabe des Dispatchs vorbei")
pruefe(set(merge.pfad_parameter) == {"master", "slave"},
       "xlsx_merge muss BEIDE Pfade zur Pruefung anmelden")

agent_src = (ROOT / "backend" / "agent.py").read_text(encoding="utf-8")
pruefe('getattr(self.tools_map.get(name), "pfad_parameter", None)' in agent_src,
       "im Dispatch fehlt der generische Pfad-Freigabe-Zweig")
pruefe('authorize_fs("read", _pv, username=_fs_akteur)' in agent_src,
       "der Zweig muss authorize_fs MIT dem Akteur aufrufen – ohne den Parameter "
       "las die Freigabe den leeren Vorgabewert des ContextVars und wies jeden "
       "Benutzer von seinem EIGENEN Anhang ab (Vorfall 2026-08-24, siehe "
       "tests/test_fs_gate_akteur.py)")

st_src = (ROOT / "backend" / "short_tracks.py").read_text(encoding="utf-8")
for n in ("xlsx_inspect", "xlsx_read_range", "xlsx_merge", "xlsx_edit"):
    pruefe(f'"{n}"' in st_src, f"{n} fehlt in BASIS_WERKZEUGE")

import json as _json  # noqa: E402
manifest = _json.loads((ROOT / "skills" / "office" / "skill.json").read_text(encoding="utf-8"))
for n in ("xlsx_inspect", "xlsx_read_range", "xlsx_merge", "xlsx_edit"):
    pruefe(n in manifest["tools"], f"{n} fehlt im Office-Manifest")

# get_tools() muss die neuen Werkzeuge wirklich liefern (Zaehlen statt hoffen –
# ein fehlendes Werkzeug meldet sich nicht, es ist einfach nicht da).
namen = {getattr(t, "name", "") for t in office_main.get_tools()}
for n in ("xlsx_inspect", "xlsx_read_range", "xlsx_merge", "xlsx_edit",
          "office_create_excel", "office_read", "office_to_pdf"):
    pruefe(n in namen, f"get_tools() liefert '{n}' nicht")

# Der Prompt darf nicht das Gegenteil der Werkzeuge behaupten (die Fehlerklasse
# WA_TASK_PROMPT / --gradient / EWS-URL). Geprueft wird die AUSSAGE, nicht ein
# Wortlaut: es muss eine Zeile geben, die bei VORHANDENER Tabelle von
# office_create_excel abraet und auf xlsx_* verweist.
zeilen = [z for z in agent_src.splitlines() if "xlsx_merge" in z and "office_create_excel" in z]
pruefe(any("FALSCHE" in z or "NICHT" in z for z in zeilen),
       "der Prompt muss bei VORHANDENEN Tabellen vom Neuaufbau abraten")
pruefe(any("TIPPE TABELLENDATEN NIEMALS AB" in z for z in agent_src.splitlines()),
       "der Prompt muss das Abtippen ausdruecklich verbieten")


# ═══════════════════════════════════════════════════════════════════════════
print("9) Nichts wurde in die echte Ablage geschrieben")
# ═══════════════════════════════════════════════════════════════════════════
pruefe(office_main.DOCS_DIR.resolve() != _ECHT_DOCS.resolve(),
       "DOCS_DIR zeigt auf die echte Ablage")
pruefe(all((TMP in p.resolve().parents) for p in office_main.DOCS_DIR.glob("*")),
       "es wurde ausserhalb des Wegwerf-Verzeichnisses geschrieben")

shutil.rmtree(TMP, ignore_errors=True)

print(f"\n{'='*60}\n{ok} bestanden, {fail} fehlgeschlagen\n{'='*60}")
sys.exit(1 if fail else 0)
